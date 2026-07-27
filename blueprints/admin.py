# Admin Blueprint
# Provides stats, export (Excel/PDF), backup/restore, DB management,
# global search, and dashboard home routes.

import os
import shutil
import tempfile
import re
import logging
import html
import sqlite3
from datetime import datetime, timedelta
from io import BytesIO

from flask import (
    Blueprint, request, jsonify, send_file, Response,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from utils import safe_table, login_required, get_json_or_400
from database import get_db, row_to_dict, DATABASE_PATH
from labels import project_status_label, task_status_label

logger = logging.getLogger(__name__)

admin_bp = Blueprint('admin', __name__)

def _excel_safe(val):
    """Prevent formula injection in Excel exports."""
    if isinstance(val, str) and val and val[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + val
    return val

# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

@admin_bp.route('/api/stats', methods=['GET'])
@login_required
def get_stats():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status = 'in_lucru' THEN 1 ELSE 0 END) as active,
            SUM(CASE WHEN status = 'finalizat' THEN 1 ELSE 0 END) as finished
        FROM proiecte
    """)
    row = cursor.fetchone()
    conn.close()
    return jsonify({
        'total': row['total'] or 0,
        'active': row['active'] or 0,
        'finished': row['finished'] or 0
    })

@admin_bp.route('/api/stats/extended', methods=['GET'])
@login_required
def get_extended_stats():
    """Extended statistics for Chart.js dashboard"""
    conn = get_db()
    cursor = conn.cursor()

    # Projects by status
    cursor.execute("""
        SELECT status, COUNT(*) as count
        FROM proiecte
        GROUP BY status
    """)
    by_status = [{'status': row['status'], 'count': row['count']} for row in cursor.fetchall()]

    # Projects by manufacturer
    cursor.execute("""
        SELECT producator, COUNT(*) as count
        FROM proiecte
        GROUP BY producator
        ORDER BY count DESC
    """)
    by_manufacturer = [{'producator': row['producator'], 'count': row['count']} for row in cursor.fetchall()]

    # Projects per month (last 12 months)
    cursor.execute("""
        SELECT
            strftime('%Y-%m', created_at) as month,
            COUNT(*) as count
        FROM proiecte
        WHERE created_at >= date('now', '-12 months')
        GROUP BY month
        ORDER BY month ASC
    """)
    by_month = [{'month': row['month'], 'count': row['count']} for row in cursor.fetchall()]

    conn.close()

    return jsonify({
        'by_status': by_status,
        'by_manufacturer': by_manufacturer,
        'by_month': by_month,
    })

# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@admin_bp.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    """Export data to Excel format"""
    export_type = request.args.get('type', 'projects')

    conn = get_db()
    cursor = conn.cursor()

    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def auto_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    logger.warning(f"auto_width cell length compute failed: {e}")
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    if export_type == 'projects':
        ws = wb.active
        ws.title = 'Proiecte'

        headers = ['Nume', 'Client', 'Tip', 'Producător', 'Status', 'Data Start', 'Deadline', 'Locație']
        ws.append(headers)
        style_header(ws)

        cursor.execute('''
            SELECT nume, client, tip, producator, status, data_incepere, deadline, locatie
            FROM proiecte ORDER BY created_at DESC
        ''')
        for row in cursor.fetchall():
            # Map status to Romanian
            row_data = list(row)
            row_data[4] = project_status_label(row_data[4])
            ws.append([_excel_safe(v) for v in row_data])

        auto_width(ws)

    elif export_type == 'tasks':
        ws = wb.active
        ws.title = 'Task-uri'

        headers = ['Proiect', 'Task', 'Status', 'Prioritate', 'Data Scadență', 'Data Finalizare']
        ws.append(headers)
        style_header(ws)

        cursor.execute('''
            SELECT p.nume, t.titlu, t.status, t.prioritate, t.data_scadenta, t.data_finalizare
            FROM tasks t
            JOIN proiecte p ON t.proiect_id = p.id
            ORDER BY t.created_at DESC
        ''')
        for row in cursor.fetchall():
            # Map status to Romanian
            priority_map = {'urgent': 'Urgent', 'normal': 'Normal', 'minor': 'Minor'}
            row_data = list(row)
            row_data[2] = task_status_label(row_data[2])
            row_data[3] = priority_map.get(row_data[3], row_data[3])
            ws.append([_excel_safe(v) for v in row_data])

        auto_width(ws)

    conn.close()

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'pif_export_{export_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    logger.info(f"Excel export: {export_type} - {filename}")

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ---------------------------------------------------------------------------
# PDF export — palette & helpers
# ---------------------------------------------------------------------------

# PIF design palette (matches the web UI tokens 1:1)
_PIF_BG = colors.HexColor('#11161e')
_PIF_ACCENT = colors.HexColor('#58d1c9')
_PIF_ACCENT_SOFT = colors.HexColor('#1d3835')
_PIF_TEXT = colors.HexColor('#1f2937')
_PIF_TEXT_DIM = colors.HexColor('#5a6473')
_PIF_LINE = colors.HexColor('#dbe1ea')
_PIF_SUCCESS = colors.HexColor('#16a34a')
_PIF_DANGER = colors.HexColor('#dc2626')
_PIF_WARNING = colors.HexColor('#d97706')


def _pdf_safe_text(text):
    """Escape HTML special chars and normalize line breaks for ReportLab Paragraph.

    Mixed content from WYSIWYG: existing <br>/<div> tags are stripped, newlines
    become <br/>, and remaining HTML chars are escaped to literal text.
    """
    if not text:
        return ''
    t = re.sub(r'<br\s*/?>', '\n', text)
    t = re.sub(r'</?div[^>]*>', '', t)
    t = html.escape(t)
    return t.replace('\n', '<br/>')


def _pdf_make_styles():
    """ParagraphStyle palette used across the PDF report."""
    base = getSampleStyleSheet()
    return {
        'title': ParagraphStyle(
            'PIFTitle', parent=base['Heading1'],
            fontSize=18, leading=22, spaceAfter=4,
            textColor=_PIF_BG, alignment=TA_LEFT, fontName='Helvetica-Bold'),
        'subtitle': ParagraphStyle(
            'PIFSubtitle', parent=base['Normal'],
            fontSize=10, leading=14, textColor=_PIF_TEXT_DIM, spaceAfter=18, fontName='Helvetica'),
        'heading': ParagraphStyle(
            'PIFHeading', parent=base['Heading2'],
            fontSize=12, leading=16, spaceBefore=14, spaceAfter=8,
            textColor=_PIF_ACCENT, fontName='Helvetica-Bold'),
        'subheading': ParagraphStyle(
            'PIFSubHeading', parent=base['Heading3'],
            fontSize=10.5, leading=14, spaceBefore=8, spaceAfter=4,
            textColor=_PIF_BG, fontName='Helvetica-Bold'),
        'normal': ParagraphStyle(
            'PIFNormal', parent=base['Normal'],
            fontSize=9.5, leading=14, textColor=_PIF_TEXT, fontName='Helvetica'),
        'small': ParagraphStyle(
            'PIFSmall', parent=base['Normal'],
            fontSize=8, leading=11, textColor=_PIF_TEXT_DIM, fontName='Helvetica'),
    }


def _pdf_section_header(elements, project_dict, is_pif, styles):
    """Header band -- accent tip-label + project name + meta line."""
    tip_label = 'PIF' if is_pif else 'Service'
    elements.append(Paragraph(
        f"<font color='#58d1c9'>{tip_label}</font> — {project_dict.get('nume', '')}",
        styles['title']))
    meta_bits = []
    if project_dict.get('client'): meta_bits.append(project_dict['client'])
    if project_dict.get('locatie'): meta_bits.append(project_dict['locatie'])
    meta_bits.append(f"export {datetime.now().strftime('%d.%m.%Y')}")
    elements.append(Paragraph(' · '.join(meta_bits), styles['subtitle']))


def _pdf_section_admin(elements, project_dict, styles):
    """1. Detalii administrative -- fixed two-column info table."""
    elements.append(Paragraph("1. Detalii administrative", styles['heading']))
    project_info = [
        ['Client', project_dict.get('client') or '-'],
        ['Locație', project_dict.get('locatie') or '-'],
        ['Producător', project_dict.get('producator') or '-'],
        ['Echipament principal', project_dict.get('echipament_principal') or '-'],
        ['Status', project_status_label(project_dict.get('status', '')) or '-'],
        ['Data începere', project_dict.get('data_incepere') or '-'],
        ['Deadline', project_dict.get('deadline') or '-'],
        ['PM', project_dict.get('pm') or '-'],
        ['Nr. comandă', project_dict.get('nr_comanda') or '-'],
        ['Nr. contract', project_dict.get('nr_contract') or '-'],
        ['Cod proiect', project_dict.get('cod_proiect') or '-'],
    ]
    info_table = Table(project_info, colWidths=[4.5*cm, 11*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), _PIF_TEXT_DIM),
        ('TEXTCOLOR', (1, 0), (1, -1), _PIF_TEXT),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (0, 0), (-1, -2), 0.3, _PIF_LINE),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))


def _pdf_section_tech(elements, project_dict, is_pif, styles):
    """2. Conținut tehnic -- PIF observații or Service before/after."""
    if is_pif and project_dict.get('observatii'):
        elements.append(Paragraph("2. Observații tehnice", styles['heading']))
        elements.append(Paragraph(_pdf_safe_text(project_dict.get('observatii', '')), styles['normal']))
        elements.append(Spacer(1, 6))
    if not is_pif and (project_dict.get('service_before') or project_dict.get('service_after')):
        elements.append(Paragraph("2. Fișă intervenție", styles['heading']))
        if project_dict.get('service_before'):
            elements.append(Paragraph("Constatări înainte de intervenție", styles['subheading']))
            elements.append(Paragraph(_pdf_safe_text(project_dict.get('service_before', '')), styles['normal']))
            elements.append(Spacer(1, 4))
        if project_dict.get('service_after'):
            elements.append(Paragraph("Acțiuni efectuate și rezultat", styles['subheading']))
            elements.append(Paragraph(_pdf_safe_text(project_dict.get('service_after', '')), styles['normal']))
            elements.append(Spacer(1, 4))


def _pdf_section_tasks(elements, tasks, section_n, styles):
    """N. Listă taskuri -- table of project tasks."""
    elements.append(Paragraph(f"{section_n}. Listă taskuri", styles['heading']))
    task_data = [['#', 'Titlu', 'Status', 'Prioritate', 'Termen']]
    for i, t in enumerate(tasks, 1):
        task_data.append([
            str(i),
            t.get('titlu', '-'),
            task_status_label(t.get('status', '')) or '-',
            (t.get('prioritate') or 'Normal').capitalize(),
            t.get('data_scadenta') or '-',
        ])
    task_table = Table(task_data, colWidths=[0.8*cm, 8.5*cm, 2.2*cm, 1.8*cm, 2.2*cm])
    task_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), _PIF_ACCENT),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('LINEBELOW', (0, 0), (-1, -1), 0.25, _PIF_LINE),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(task_table)
    elements.append(Spacer(1, 10))


# ---------------------------------------------------------------------------
# PDF export routes
# ---------------------------------------------------------------------------

@admin_bp.route('/api/export/pdf', methods=['GET'])
@login_required
def export_pdf():
    """Export project to PDF format"""
    project_id = request.args.get('project_id')
    if not project_id:
        return jsonify({'error': 'project_id is required'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (project_id,))
    project = cursor.fetchone()
    if not project:
        conn.close()
        return jsonify({'error': 'Project not found'}), 404
    project_dict = row_to_dict(project)

    cursor.execute('SELECT * FROM tasks WHERE proiect_id = ? ORDER BY ordine ASC', (project_id,))
    tasks = [row_to_dict(row) for row in cursor.fetchall()]
    conn.close()

    is_pif = (project_dict.get('tip') == 'PIF')

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm, topMargin=2*cm, bottomMargin=2*cm,
        title=f"PIF Report — {re.sub(r'[<>]', '', project_dict.get('nume', '') or '')}",
        author='PIF Dashboard'
    )

    styles = _pdf_make_styles()
    elements = []

    _pdf_section_header(elements, project_dict, is_pif, styles)
    _pdf_section_admin(elements, project_dict, styles)
    _pdf_section_tech(elements, project_dict, is_pif, styles)

    n_tasks = 3

    if tasks:
        _pdf_section_tasks(elements, tasks, n_tasks, styles)

    # Footer
    elements.append(Spacer(1, 16))
    elements.append(Paragraph(
        f"<font color='#5a6473'>Document generat automat din PIF Dashboard · {datetime.now().strftime('%d.%m.%Y %H:%M')} · Ion Ursu</font>",
        styles['small']))

    doc.build(elements)

    buffer.seek(0)
    safe_name = (project_dict.get('nume', 'project') or 'project').replace(' ', '_').replace('/', '_')
    if project_dict.get('cod_proiect'):
        filename = f"{project_dict['cod_proiect']}_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    else:
        filename = f"pif_report_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    logger.info(f"PDF export: {filename}")

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@admin_bp.route('/api/export/ics', methods=['GET'])
@login_required
def export_ics():
    """Export deadline-uri proiecte + scadente taskuri ca fisier .ics (calendar).
    Abonabil din Google/Apple Calendar."""
    conn = get_db()
    cur = conn.cursor()
    now = datetime.now().strftime('%Y%m%dT%H%M%SZ')
    lines = ['BEGIN:VCALENDAR', 'VERSION:2.0', 'PRODID:-//PIF Dashboard//RO',
             'CALSCALE:GREGORIAN', 'X-WR-CALNAME:PIF Dashboard']

    def esc(s):
        return (str(s or '')).replace('\\', '\\\\').replace(';', '\\;').replace(',', '\\,').replace('\n', '\\n')

    def date_only(iso):
        m = re.match(r'(\d{4})-(\d{2})-(\d{2})', str(iso or ''))
        return (m.group(1) + m.group(2) + m.group(3)) if m else None

    def add_event(uid, dt, summary, desc=''):
        d = date_only(dt)
        if not d:
            return
        lines.extend(['BEGIN:VEVENT', f'UID:{uid}@pif.iupif.org', f'DTSTAMP:{now}',
                      f'DTSTART;VALUE=DATE:{d}', f'SUMMARY:{esc(summary)}'])
        if desc:
            lines.append(f'DESCRIPTION:{esc(desc)}')
        lines.append('END:VEVENT')

    try:
        cur.execute("SELECT id, nume, deadline FROM proiecte WHERE deadline IS NOT NULL "
                    "AND TRIM(deadline) <> '' AND status != 'finalizat'")
        for r in cur.fetchall():
            add_event(f"proj-{r['id']}", r['deadline'], f"Deadline: {r['nume']}", 'Deadline proiect PIF')

        cur.execute("SELECT t.id, t.titlu, t.data_scadenta, p.nume AS pnume FROM tasks t "
                    "JOIN proiecte p ON t.proiect_id = p.id WHERE t.data_scadenta IS NOT NULL "
                    "AND TRIM(t.data_scadenta) <> '' AND (t.data_finalizare IS NULL OR TRIM(t.data_finalizare) = '')")
        for r in cur.fetchall():
            add_event(f"task-{r['id']}", r['data_scadenta'], f"Scadenta: {r['titlu']}", f"Proiect: {r['pnume']}")

        cur.execute("SELECT id, titlu, data_scadenta FROM global_tasks WHERE data_scadenta IS NOT NULL "
                    "AND TRIM(data_scadenta) <> '' AND (data_finalizare IS NULL OR TRIM(data_finalizare) = '')")
        for r in cur.fetchall():
            add_event(f"gtask-{r['id']}", r['data_scadenta'], f"Scadenta: {r['titlu']}", 'Task global')
    finally:
        conn.close()

    lines.append('END:VCALENDAR')
    ics = '\r\n'.join(lines) + '\r\n'
    return Response(ics, mimetype='text/calendar',
                    headers={'Content-Disposition': 'attachment; filename="pif-calendar.ics"'})


@admin_bp.route('/api/export/pdf/all', methods=['GET'])
@login_required
def export_all_projects_pdf():
    """Export summary of all projects to PDF"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM proiecte ORDER BY created_at DESC')
    projects = [row_to_dict(row) for row in cursor.fetchall()]
    conn.close()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=20, alignment=TA_CENTER)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=12, spaceBefore=15, spaceAfter=10)

    elements = []
    elements.append(Paragraph("Raport Sumar Proiecte PIF", title_style))
    elements.append(Spacer(1, 10))

    # Summary table
    summary_data = [['#', 'Nume', 'Client', 'Tip', 'Status', 'Deadline']]
    for i, proj in enumerate(projects, 1):
        summary_data.append([
            str(i),
            proj.get('nume', '-')[:30],
            proj.get('client', '-')[:20],
            proj.get('tip', '-'),
            project_status_label(proj.get('status', '')) or '-',
            proj.get('deadline', '-')
        ])

    summary_table = Table(summary_data, colWidths=[1*cm, 5*cm, 3.5*cm, 2*cm, 3*cm, 2.5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(summary_table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"pif_all_projects_{datetime.now().strftime('%Y%m%d')}.pdf"
    logger.info(f"PDF export all projects: {filename}")

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# ---------------------------------------------------------------------------
# Backup / Restore
# ---------------------------------------------------------------------------

@admin_bp.route('/api/backup', methods=['GET'])
@login_required
def backup_database():
    conn = get_db()
    cursor = conn.cursor()

    backup = {}

    tables = ['proiecte', 'tasks', 'task_subtasks', 'task_dependencies',
                  'implementari', 'global_tasks', 'clienti', 'app_settings']
    # sarim tabelele absente
    # ca sa nu pice backup-ul cu 500.
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    existing = {row[0] for row in cursor.fetchall()}
    for table in tables:
        if table not in existing:
            backup[table] = []
            continue
        cursor.execute(f'SELECT * FROM {safe_table(table)}')
        rows = cursor.fetchall()
        backup[table] = [row_to_dict(row) for row in rows]

    conn.close()

    return jsonify(backup)


@admin_bp.route('/api/restore', methods=['POST'])
@login_required
def restore_database():
    data = get_json_or_400()

    conn = get_db()
    cursor = conn.cursor()

    try:
        # Run the whole clear + restore inside ONE transaction, so a bad payload
        # rolls the deletes back instead of leaving the database wiped.
        conn.execute('BEGIN TRANSACTION')
        # Clear existing data (skip tables absent on this deploy)
        tables = ['proiecte', 'tasks', 'task_subtasks', 'task_dependencies',
                  'implementari', 'global_tasks', 'clienti', 'app_settings']
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        existing = {row[0] for row in cursor.fetchall()}
        for table in tables:
            if table in existing:
                cursor.execute(f'DELETE FROM {safe_table(table)}')

        # Restore proiecte
        for p in data.get('proiecte', []):
            cursor.execute('''
                INSERT INTO proiecte (id, tip, nume, client, locatie, echipament_principal, producator,
                    cod_proiect, pm, folder_server, data_incepere, deadline, data_crearii,
                    status, observatii, nr_comanda, nr_contract, service_before, service_after,
                    confirmat_client, client_nume_confirmare, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                p.get('id'), p.get('tip'), p.get('nume'), p.get('client'), p.get('locatie'),
                p.get('echipament_principal'), p.get('producator'), p.get('cod_proiect'),
                p.get('pm'), p.get('folder_server'), p.get('data_incepere'), p.get('deadline'),
                p.get('data_crearii'), p.get('status'), p.get('observatii'), p.get('nr_comanda'),
                p.get('nr_contract'), p.get('service_before'), p.get('service_after'),
                p.get('confirmat_client', 0), p.get('client_nume_confirmare'),
                p.get('created_at'), p.get('updated_at')
            ))

        # Restore tasks (v8+ columns: descriere, recurenta, updated_at, ordine; v21: data_planificata, ordine_agenda)
        for t in data.get('tasks', []):
            cursor.execute('''
                INSERT INTO tasks (id, proiect_id, titlu, descriere, status, prioritate,
                    data_scadenta, data_finalizare, ordine, recurenta, created_at, updated_at,
                    data_planificata, ordine_agenda)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (t.get('id'), t.get('proiect_id'), t.get('titlu'), t.get('descriere'),
                  t.get('status'), t.get('prioritate'), t.get('data_scadenta'),
                  t.get('data_finalizare'), t.get('ordine', 0), t.get('recurenta'),
                  t.get('created_at'), t.get('updated_at'),
                  t.get('data_planificata'), t.get('ordine_agenda', 0)))

        # jurnal / timer_sessions din backup-uri vechi se ignora (v22 a scos featureul)
        # checklist_pif / checklist_categorii / project_templates din backup-uri vechi
        # se ignora (v23 a sters featureurile Checklist + Template)

        # Restore global_tasks (v9+ column: recurenta; v21: data_planificata, ordine_agenda)
        for gt in data.get('global_tasks', []):
            cursor.execute('''
                INSERT INTO global_tasks (id, titlu, descriere, prioritate, status, categorie,
                    data_scadenta, data_finalizare, recurenta, created_at, updated_at,
                    data_planificata, ordine_agenda)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (gt.get('id'), gt.get('titlu'), gt.get('descriere'), gt.get('prioritate'),
                  gt.get('status'), gt.get('categorie'), gt.get('data_scadenta'),
                  gt.get('data_finalizare'), gt.get('recurenta'),
                  gt.get('created_at'), gt.get('updated_at'),
                  gt.get('data_planificata'), gt.get('ordine_agenda', 0)))

        # Restore clienti
        for c in data.get('clienti', []):
            cursor.execute('''
                INSERT INTO clienti (id, nume, adresa, telefon, email, contact_principal, note, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (c.get('id'), c.get('nume'), c.get('adresa'), c.get('telefon'),
                  c.get('email'), c.get('contact_principal'), c.get('note'), c.get('created_at')))

        # Restore task_subtasks (schema: id, task_id, titlu, done, ordine, created_at)
        for s in data.get('task_subtasks', []):
            cursor.execute('''
                INSERT INTO task_subtasks (id, task_id, titlu, done, ordine, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (s.get('id'), s.get('task_id'), s.get('titlu'),
                  s.get('done', 0), s.get('ordine', 0), s.get('created_at')))

        # Restore implementari (perioadele de implementare — planificarea reala
        # a lui Ion). Lipseau din backup pana in 2026-07-27: un restore le pierdea
        # in tacere.
        for im in data.get('implementari', []):
            cursor.execute('''
                INSERT INTO implementari (id, proiect_id, data_start, data_sfarsit,
                    locatie, eticheta, ordine, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (im.get('id'), im.get('proiect_id'), im.get('data_start'),
                  im.get('data_sfarsit'), im.get('locatie') or 'site',
                  im.get('eticheta'), im.get('ordine', 0), im.get('created_at')))

        # Restore task_dependencies (dupa tasks — FK pe ambele capete)
        for dep in data.get('task_dependencies', []):
            cursor.execute('''
                INSERT INTO task_dependencies (id, proiect_id, predecessor_id,
                    successor_id, tip, lag, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (dep.get('id'), dep.get('proiect_id'), dep.get('predecessor_id'),
                  dep.get('successor_id'), dep.get('tip') or 'FS',
                  dep.get('lag', 0), dep.get('created_at')))

        # Restore app_settings (vault Obsidian, cheile de idempotenta debrief).
        # assistant_memory din backup-uri vechi se ignora (v23 a sters Hermes).
        for s in data.get('app_settings', []):
            cursor.execute('INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)',
                           (s.get('key'), s.get('value'), s.get('updated_at')))

        conn.commit()
        conn.close()

        logger.info("Database restored successfully")
        return jsonify({'message': 'Database restored successfully'})
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error(f"Error restoring database: {e}")
        return jsonify({'error': 'Restaurare esuata — modificarile au fost anulate.'}), 500

# ---------------------------------------------------------------------------
# Admin DB management
# ---------------------------------------------------------------------------

@admin_bp.route('/admin/db-upload', methods=['GET'])
@login_required
def admin_db_upload_page():
    """Minimal HTML form for uploading a local DB file."""
    return '''<!doctype html>
<html><head><meta charset="utf-8"><title>DB upload</title>
<style>
body{background:#0a0d12;color:#e3e8ef;font-family:system-ui,sans-serif;
     max-width:560px;margin:60px auto;padding:0 24px}
h1{font-size:20px;margin:0 0 16px}
.box{background:#161c26;border:1px solid #232a36;border-radius:10px;padding:24px}
input[type=file]{margin:16px 0;color:#e3e8ef}
button{background:#58d1c9;color:#0a0d12;border:0;padding:10px 18px;
       border-radius:8px;font-weight:600;cursor:pointer}
button:disabled{opacity:.5;cursor:wait}
.note{color:#9aa4b2;font-size:13px;margin-top:14px}
pre{background:#0a0d12;border:1px solid #232a36;border-radius:6px;
    padding:10px;font-size:12px;overflow:auto;max-height:280px}
.ok{color:#66d19e}.err{color:#f97066}
</style></head>
<body><div class="box">
<h1>Upload DB SQLite</h1>
<form id="f">
<input type="file" name="db" accept=".db" required>
<br><button id="submit" type="submit">Upload (inlocuieste DB serverului)</button>
</form>
<div class="note">DB-ul curent va fi salvat in <code>backups/</code> automat inainte de inlocuire.</div>
<pre id="out"></pre>
</div>
<script>
const f=document.getElementById('f'),btn=document.getElementById('submit'),out=document.getElementById('out');
f.addEventListener('submit',async e=>{
  e.preventDefault();btn.disabled=true;out.textContent='Uploading...';
  const fd=new FormData(f);
  try{
    const r=await fetch('/api/admin/db-upload',{method:'POST',body:fd});
    const j=await r.json();
    out.textContent=JSON.stringify(j,null,2);
    out.className=r.ok?'ok':'err';
  }catch(err){out.textContent=err.message;out.className='err';}
  btn.disabled=false;
});
</script>
</body></html>'''


@admin_bp.route('/api/admin/db-upload', methods=['POST'])
@login_required
def admin_db_upload():
    """Replace the server SQLite DB with an uploaded copy.
    Expects a multipart form field named 'db' with a .db file.
    Validates SQLite header, backs up the existing DB, then atomic-replaces it.
    """
    if 'db' not in request.files:
        return jsonify({'error': "missing form field 'db'"}), 400
    f = request.files['db']
    if not f or f.filename == '':
        return jsonify({'error': 'empty file'}), 400

    dir_name = os.path.dirname(DATABASE_PATH) or '.'
    fd, tmp_path = tempfile.mkstemp(prefix='dbupload_', suffix='.db', dir=dir_name)
    os.close(fd)
    try:
        f.save(tmp_path)
        with open(tmp_path, 'rb') as fh:
            magic = fh.read(16)
        if not magic.startswith(b'SQLite format 3'):
            os.unlink(tmp_path)
            return jsonify({'error': 'not a valid SQLite database'}), 400

        # Beyond the magic header: verify the file is a sound SQLite DB and
        # actually looks like a PIF database before replacing the live one.
        try:
            _chk = sqlite3.connect(tmp_path)
            _integrity = _chk.execute('PRAGMA integrity_check').fetchone()
            _has_core = _chk.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='proiecte'"
            ).fetchone()
            _chk.close()
        except (sqlite3.Error, OSError):
            os.unlink(tmp_path)
            return jsonify({'error': 'fisierul nu este o baza SQLite utilizabila'}), 400
        if not _integrity or _integrity[0] != 'ok':
            os.unlink(tmp_path)
            return jsonify({'error': 'baza incarcata a esuat integrity_check'}), 400
        if not _has_core:
            os.unlink(tmp_path)
            return jsonify({'error': 'baza incarcata nu contine structura PIF (tabelul proiecte)'}), 400

        backups_dir = os.path.join(dir_name, 'backups')
        os.makedirs(backups_dir, exist_ok=True)
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(backups_dir, f'pif_dashboard_pre_upload_{stamp}.db')
        if os.path.exists(DATABASE_PATH):
            shutil.copy2(DATABASE_PATH, backup_path)

        # Checkpoint WAL into the main DB file so the *-wal/*-shm files are
        # truncated. Without this, the OLD WAL can persist after replace and
        # corrupt reads against the NEW DB.
        try:
            from database import close_db
            conn = get_db()
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            conn.commit()
            close_db(None)
        except Exception as e:
            logger.warning(f"WAL checkpoint before db replace failed: {e}")

        os.replace(tmp_path, DATABASE_PATH)

        return jsonify({
            'status': 'ok',
            'backup': backup_path,
            'replaced': DATABASE_PATH,
            'size': os.path.getsize(DATABASE_PATH),
        })
    except Exception as e:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        logger.error(f"db-upload failed: {e}")
        return jsonify({'error': 'Incarcarea bazei a esuat.'}), 500


@admin_bp.route('/api/admin/db-dump', methods=['GET'])
@login_required
def admin_db_dump():
    """Stream o copie consistenta a DB-ului SQLite pentru audit local.
    Foloseste sqlite3 backup API ca sa nu blocheze scrieri concurente.
    """
    import sqlite3 as _sql3
    from flask import after_this_request

    if not os.path.exists(DATABASE_PATH):
        return jsonify({'error': 'DB not found on server'}), 404

    # Hot backup la un fisier temporar (safe vs WAL)
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp.close()
    src = _sql3.connect(DATABASE_PATH)
    dst = _sql3.connect(tmp.name)
    try:
        src.backup(dst)
    finally:
        dst.close()
        src.close()

    # Schedule cleanup of the temp file once the response is sent -- without
    # this the temp dir slowly fills up (~40 MB per download).
    @after_this_request
    def _cleanup(response):
        try:
            os.unlink(tmp.name)
        except OSError:
            pass
        return response

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=f'pif_dashboard_{timestamp}.db',
        mimetype='application/octet-stream',
    )

# ---------------------------------------------------------------------------
# Global search
# ---------------------------------------------------------------------------

def _search_snippet(text, query, width=80):
    """A short context window around the first match of `query` in `text`."""
    if not text:
        return ''
    text = str(text)
    low = text.lower()
    idx = low.find(query.lower())
    if idx < 0:
        return text[:width].strip()
    start = max(0, idx - 32)
    end = min(len(text), idx + len(query) + width)
    return ('…' if start > 0 else '') + text[start:end].strip() + ('…' if end < len(text) else '')


@admin_bp.route('/api/search', methods=['GET'])
@login_required
def global_search():
    """Unified search across everything in the app for the command palette."""
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'results': [], 'query': q})
    like = f'%{q}%'
    results = []
    conn = get_db()
    cur = conn.cursor()

    cur.execute('SELECT id, nume, client FROM proiecte '
                'WHERE nume LIKE ? OR client LIKE ? OR cod_proiect LIKE ? OR locatie LIKE ? LIMIT 8',
                (like, like, like, like))
    for r in cur.fetchall():
        results.append({'type': 'proiect', 'id': r['id'], 'title': r['nume'],
                        'subtitle': r['client'] or '', 'snippet': '', 'proiect_id': r['id']})

    cur.execute('SELECT id, nume, observatii FROM proiecte WHERE observatii LIKE ? LIMIT 8', (like,))
    for r in cur.fetchall():
        results.append({'type': 'observatie', 'id': r['id'], 'title': f"Observații — {r['nume']}",
                        'subtitle': '', 'snippet': _search_snippet(r['observatii'], q), 'proiect_id': r['id']})

    cur.execute('SELECT t.id, t.titlu, t.descriere, t.proiect_id, p.nume AS pnume FROM tasks t '
                'JOIN proiecte p ON t.proiect_id = p.id '
                'WHERE t.titlu LIKE ? OR t.descriere LIKE ? LIMIT 12', (like, like))
    for r in cur.fetchall():
        results.append({'type': 'task', 'id': r['id'], 'title': r['titlu'],
                        'subtitle': r['pnume'], 'snippet': _search_snippet(r['descriere'], q),
                        'proiect_id': r['proiect_id']})

    cur.execute('SELECT id, titlu, descriere, categorie FROM global_tasks '
                'WHERE titlu LIKE ? OR descriere LIKE ? LIMIT 10', (like, like))
    for r in cur.fetchall():
        results.append({'type': 'global_task', 'id': r['id'], 'title': r['titlu'],
                        'subtitle': r['categorie'] or 'Task zilnic', 'snippet': _search_snippet(r['descriere'], q)})

    cur.execute('SELECT id, nume, telefon FROM clienti WHERE nume LIKE ? OR contact_principal LIKE ? LIMIT 6',
                (like, like))
    for r in cur.fetchall():
        results.append({'type': 'client', 'id': r['id'], 'title': r['nume'],
                        'subtitle': 'Client', 'snippet': r['telefon'] or ''})

    conn.close()

    return jsonify({'results': results, 'query': q, 'count': len(results)})

@admin_bp.route('/api/calendar', methods=['GET'])
@login_required
def calendar_view():
    """Calendarul personal: unde esti in fiecare zi.

    Ion e o singura persoana, iar planificarea lui reala sunt PERIOADELE de
    implementare (14 pe 12 proiecte), nu deadline-urile (2 din 20 de proiecte).
    Deci intrebarea la care raspunde ecranul asta e „unde sunt marti", si tot
    aici stau si deciziile — nu intr-o lista separata.

    `necesita_decizie` = perioada s-a terminat, dar proiectul n-a fost mutat.
    Ori s-a facut si trebuie inchis, ori a alunecat si trebuie replanificat.

    `neplanificate` = proiecte active fara nicio perioada viitoare. Ele stau in
    banda laterala, de unde se trag pe o zi ca sa devina perioade.
    """
    start = (request.args.get('start') or '').strip()
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', start):
        start = datetime.now().date().replace(day=1).isoformat()
    try:
        zile = int(request.args.get('zile') or 42)
    except (TypeError, ValueError):
        zile = 42
    zile = max(7, min(zile, 200))
    start_d = datetime.strptime(start, '%Y-%m-%d').date()
    end_s = (start_d + timedelta(days=zile)).isoformat()

    conn = get_db()
    cursor = conn.cursor()

    # O perioada intra in fereastra daca se intersecteaza cu ea, nu doar daca
    # incepe in ea — altfel un bloc de 4 zile care trece peste 1 ale lunii dispare.
    cursor.execute("""
        SELECT i.id, i.data_start, i.data_sfarsit, i.eticheta, i.locatie,
               p.id AS proiect_id, p.nume, p.client, p.locatie AS locatie_proiect,
               p.status, p.tip,
               (SELECT COUNT(*) FROM tasks t
                 WHERE t.proiect_id = p.id AND t.status != 'done') AS taskuri_deschise,
               (CASE WHEN p.status NOT IN ('finalizat', 'anulat')
                      AND date(COALESCE(NULLIF(i.data_sfarsit, ''), i.data_start)) < date('now')
                     THEN 1 ELSE 0 END) AS necesita_decizie
        FROM implementari i JOIN proiecte p ON p.id = i.proiect_id
        WHERE p.status != 'anulat'
          AND date(i.data_start) < date(?)
          AND date(COALESCE(NULLIF(i.data_sfarsit, ''), i.data_start)) >= date(?)
        ORDER BY i.data_start, p.client, p.nume
    """, (end_s, start))
    perioade = [dict(r) for r in cursor.fetchall()]

    # Tot ce cere o decizie, chiar daca a ramas in urma ferestrei afisate —
    # altfel navighezi pe luna viitoare si semnalul dispare.
    cursor.execute("""
        SELECT i.id, i.data_start, i.data_sfarsit, i.eticheta, i.locatie,
               p.id AS proiect_id, p.nume, p.client, p.status
        FROM implementari i JOIN proiecte p ON p.id = i.proiect_id
        WHERE p.status NOT IN ('finalizat', 'anulat')
          AND date(COALESCE(NULLIF(i.data_sfarsit, ''), i.data_start)) < date('now')
        ORDER BY i.data_start
    """)
    de_decis = [dict(r) for r in cursor.fetchall()]

    cursor.execute("""
        SELECT id AS proiect_id, nume, client, status, tip
        FROM proiecte p
        WHERE p.status NOT IN ('finalizat', 'anulat')
          AND NOT EXISTS (
            SELECT 1 FROM implementari i
            WHERE i.proiect_id = p.id
              AND date(COALESCE(NULLIF(i.data_sfarsit, ''), i.data_start)) >= date('now')
          )
        ORDER BY p.status DESC, p.nume
    """)
    neplanificate = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return jsonify({
        'start': start,
        'zile': zile,
        'today': datetime.now().date().isoformat(),
        'perioade': perioade,
        'de_decis': de_decis,
        'neplanificate': neplanificate,
    })


# ---------------------------------------------------------------------------
# Dashboard home
# ---------------------------------------------------------------------------

@admin_bp.route('/api/dashboard/home', methods=['GET'])
@login_required
def dashboard_home():
    conn = get_db()
    cursor = conn.cursor()

    # Active projects count
    cursor.execute("SELECT COUNT(*) FROM proiecte WHERE status NOT IN ('finalizat', 'anulat')")
    active_projects = cursor.fetchone()[0]

    # Total projects
    cursor.execute("SELECT COUNT(*) FROM proiecte")
    total_projects = cursor.fetchone()[0]

    # Tasks completed in the last 7 days (project tasks + global tasks).
    # data_finalizare is set on completion; the timer-based weekly hours card
    # was removed in v22 (orele se ponteaza in e100).
    _done_where = """
        data_finalizare IS NOT NULL AND TRIM(data_finalizare) <> ''
    """
    cursor.execute(f"""
        SELECT COUNT(*) FROM (
            SELECT data_finalizare FROM tasks WHERE {_done_where}
            UNION ALL
            SELECT data_finalizare FROM global_tasks WHERE {_done_where}
        )
        WHERE date(data_finalizare) >= date('now', '-6 days')
    """)
    weekly_done = cursor.fetchone()[0]

    # Previous 7-day window for delta
    cursor.execute(f"""
        SELECT COUNT(*) FROM (
            SELECT data_finalizare FROM tasks WHERE {_done_where}
            UNION ALL
            SELECT data_finalizare FROM global_tasks WHERE {_done_where}
        )
        WHERE date(data_finalizare) >= date('now', '-13 days')
          AND date(data_finalizare) < date('now', '-6 days')
    """)
    weekly_done_delta = weekly_done - cursor.fetchone()[0]

    # Completed count per day, last 7 calendar days (oldest -> newest)
    cursor.execute(f"""
        SELECT CAST(julianday(date('now')) - julianday(date(data_finalizare)) AS INTEGER) AS days_ago,
               COUNT(*) AS cnt
        FROM (
            SELECT data_finalizare FROM tasks WHERE {_done_where}
            UNION ALL
            SELECT data_finalizare FROM global_tasks WHERE {_done_where}
        )
        WHERE date(data_finalizare) >= date('now', '-6 days')
        GROUP BY days_ago
    """)
    _spark = {r['days_ago']: r['cnt'] for r in cursor.fetchall()}
    weekly_spark = [_spark.get(i, 0) for i in range(6, -1, -1)]

    # Urgent tasks — global + project-level (UNION)
    cursor.execute("""
        SELECT * FROM (
            SELECT id, titlu, prioritate, data_scadenta, categorie,
                   NULL as proiect_id, NULL as proiect_nume
            FROM global_tasks
            WHERE LOWER(prioritate) = 'urgent' AND status != 'done'

            UNION ALL

            SELECT t.id, t.titlu, t.prioritate, t.data_scadenta, '' as categorie,
                   t.proiect_id, p.nume as proiect_nume
            FROM tasks t JOIN proiecte p ON t.proiect_id = p.id
            WHERE LOWER(t.prioritate) = 'urgent' AND t.status != 'done'
              AND p.status != 'anulat'
        )
        ORDER BY data_scadenta IS NULL, data_scadenta
        LIMIT 50
    """)
    urgent_tasks = [dict(r) for r in cursor.fetchall()]
    urgent_count = len(urgent_tasks)

    # Upcoming project deadlines (next 7 days)
    cursor.execute("""
        SELECT id, nume, client, deadline
        FROM proiecte
        WHERE deadline IS NOT NULL
          AND deadline >= date('now')
          AND deadline <= date('now', '+7 days')
          AND status NOT IN ('finalizat', 'anulat')
        ORDER BY deadline LIMIT 5
    """)
    upcoming_deadlines = [dict(r) for r in cursor.fetchall()]
    deadline_count = len(upcoming_deadlines)

    # ---- Risc, calculat pe PERIOADE DE IMPLEMENTARE, nu pe deadline ----------
    # Doar 2 din 20 de proiecte au deadline, dar 12 au perioade planificate —
    # perioadele sunt planificarea reala, deci semnalele se citesc de acolo.
    #
    #  1. perioada trecuta, status nemiscat -> ori s-a facut si nu ai inchis-o,
    #     ori a alunecat si trebuie replanificata
    #  2. perioada in urmatoarele 7 zile pe un proiect fara niciun task
    #  3. proiect in lucru fara nicio perioada viitoare
    cursor.execute("""
        SELECT i.id, i.data_start, i.data_sfarsit, i.eticheta, i.locatie,
               p.id AS proiect_id, p.nume, p.client, p.status
        FROM implementari i JOIN proiecte p ON p.id = i.proiect_id
        WHERE p.status NOT IN ('finalizat', 'anulat')
          AND date(COALESCE(NULLIF(i.data_sfarsit, ''), i.data_start)) < date('now')
        ORDER BY i.data_start DESC LIMIT 20
    """)
    risc_perioade_trecute = [dict(r) for r in cursor.fetchall()]

    cursor.execute("""
        SELECT p.id AS proiect_id, p.nume, p.client, p.status,
               MIN(i.data_start) AS data_start, COUNT(*) AS n_perioade
        FROM implementari i JOIN proiecte p ON p.id = i.proiect_id
        WHERE p.status NOT IN ('finalizat', 'anulat')
          AND date(i.data_start) >= date('now')
          AND date(i.data_start) <= date('now', '+7 days')
          AND NOT EXISTS (SELECT 1 FROM tasks t WHERE t.proiect_id = p.id)
        GROUP BY p.id ORDER BY data_start LIMIT 20
    """)
    risc_fara_taskuri = [dict(r) for r in cursor.fetchall()]

    cursor.execute("""
        SELECT id AS proiect_id, nume, client, status
        FROM proiecte p
        WHERE p.status = 'in_lucru'
          AND NOT EXISTS (
            SELECT 1 FROM implementari i
            WHERE i.proiect_id = p.id
              AND date(COALESCE(NULLIF(i.data_sfarsit, ''), i.data_start)) >= date('now')
          )
        ORDER BY nume LIMIT 20
    """)
    risc_fara_perioada = [dict(r) for r in cursor.fetchall()]

    risc = {
        'perioade_trecute': risc_perioade_trecute,
        'fara_taskuri': risc_fara_taskuri,
        'fara_perioada': risc_fara_perioada,
    }
    risc_count = sum(len(v) for v in risc.values())

    # Today's tasks — open tasks (not done), with due-today/overdue surfaced first,
    # then by priority. Tasks without a scadenta still show (the user rarely sets one).
    cursor.execute("""
        SELECT id, titlu, status, prioritate, categorie FROM global_tasks
        WHERE status != 'done'
          AND NOT (
            recurenta IS NOT NULL AND TRIM(recurenta) <> ''
            AND data_scadenta IS NOT NULL AND date(data_scadenta) > date('now')
          )
        ORDER BY
            CASE WHEN data_scadenta IS NOT NULL AND date(data_scadenta) <= date('now') THEN 0 ELSE 1 END,
            CASE LOWER(prioritate) WHEN 'urgent' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END,
            data_scadenta IS NULL, data_scadenta,
            created_at DESC
        LIMIT 5
    """)
    todays_tasks = [dict(r) for r in cursor.fetchall()]

    conn.close()

    return jsonify({
        'stats': {
            'active_projects': active_projects,
            'total_projects': total_projects,
            'weekly_done': weekly_done,
            'weekly_done_delta': weekly_done_delta,
            'weekly_spark': weekly_spark,
            'urgent_count': urgent_count,
            'deadline_count': deadline_count,
            'risc_count': risc_count
        },
        'urgent_tasks': urgent_tasks,
        'upcoming_deadlines': upcoming_deadlines,
        'risc': risc,
        'todays_tasks': todays_tasks
    })

