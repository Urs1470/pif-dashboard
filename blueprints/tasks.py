import calendar
import logging
import re
from datetime import datetime, timedelta
from io import BytesIO

from flask import Blueprint, jsonify, request, send_file

from database import get_db, row_to_dict
from utils import generate_uuid, login_required, get_json_or_400

tasks_bp = Blueprint('tasks', __name__)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _skip_weekend(d):
    """Ion nu lucreaza in weekend: o scadenta care cade sambata sau duminica se
    muta pe lunea urmatoare (inainte, nu inapoi). weekday(): luni=0 .. duminica=6."""
    if d.weekday() == 5:        # sambata -> +2 zile = luni
        return d + timedelta(days=2)
    if d.weekday() == 6:        # duminica -> +1 zi = luni
        return d + timedelta(days=1)
    return d


def _next_recurrence_date(base_str, recurenta):
    """base_str: 'YYYY-MM-DD' (flatpickr format) or empty. Returns the next
    occurrence date as 'YYYY-MM-DD'. Falls back to today when base is unparsable.
    Sare peste weekend: o aparitie care ar cadea sambata/duminica se muta luni
    (ex. un task zilnic terminat vineri revine luni, nu sambata)."""
    try:
        base = datetime.strptime((base_str or '')[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        base = datetime.now().date()
    if recurenta == 'zilnic':
        nxt = base + timedelta(days=1)
    elif recurenta == 'saptamanal':
        nxt = base + timedelta(days=7)
    elif recurenta == 'lunar':
        m = base.month + 1
        y = base.year + (1 if m > 12 else 0)
        if m > 12:
            m -= 12
        d = min(base.day, calendar.monthrange(y, m)[1])
        nxt = datetime(y, m, d).date()
    else:
        return base_str or ''
    return _skip_weekend(nxt).isoformat()


def _spawn_recurring_task(cursor, existing, recurenta):
    """Create the next occurrence of a recurring task that was just completed.
    Copies title/priority/description/recurrence and fresh (unchecked) subtasks.
    `existing` is the sqlite Row of the completed task. Returns the new id."""
    new_id = generate_uuid()
    now = datetime.now().isoformat()
    next_scad = _next_recurrence_date(existing['data_scadenta'] or '', recurenta)
    cursor.execute('SELECT MAX(ordine) FROM tasks WHERE proiect_id = ?', (existing['proiect_id'],))
    max_ordine = cursor.fetchone()[0] or 0
    # ordine_agenda e deliberat NECOPIAT: urmatoarea
    # occurrence is born unplanned and surfaces on the Astazi board later via its
    # future data_scadenta, not the moment the current one is completed.
    cursor.execute('''
        INSERT INTO tasks (id, proiect_id, titlu, status, data_scadenta,
                           data_finalizare, ordine, created_at, descriere, recurenta, updated_at)
        VALUES (?, ?, ?, 'to_do', ?, '', ?, ?, ?, ?, ?)
    ''', (new_id, existing['proiect_id'], existing['titlu'],
          next_scad, max_ordine + 1, now, existing['descriere'] or '', recurenta, now))
    # Carry the subtasks over, all unchecked -- a recurring checklist repeats clean.
    cursor.execute('SELECT titlu, ordine FROM task_subtasks WHERE task_id = ? ORDER BY ordine', (existing['id'],))
    for srow in cursor.fetchall():
        cursor.execute(
            'INSERT INTO task_subtasks (id, task_id, titlu, done, ordine, created_at) VALUES (?, ?, ?, 0, ?, ?)',
            (generate_uuid(), new_id, srow['titlu'], srow['ordine'], now)
        )
    return new_id


def _spawn_recurring_global_task(cursor, existing, recurenta):
    """Spawn the next occurrence of a completed recurring daily task. Copies
    title/priority/category/description + fresh subtasks. Returns the new id."""
    new_id = generate_uuid()
    now = datetime.now().isoformat()
    next_scad = _next_recurrence_date(existing['data_scadenta'] or '', recurenta)
    # ordine_agenda deliberat necopiat (vezi _spawn_recurring_task).
    cursor.execute('''
        INSERT INTO global_tasks (id, titlu, descriere, status, categorie,
                                  data_scadenta, data_finalizare, created_at, updated_at, recurenta)
        VALUES (?, ?, ?, 'to_do', ?, ?, '', ?, ?, ?)
    ''', (new_id, existing['titlu'], existing['descriere'] or '',
          existing['categorie'], next_scad, now, now, recurenta))
    cursor.execute('SELECT titlu, ordine FROM task_subtasks WHERE task_id = ? ORDER BY ordine', (existing['id'],))
    for srow in cursor.fetchall():
        cursor.execute(
            'INSERT INTO task_subtasks (id, task_id, titlu, done, ordine, created_at) VALUES (?, ?, ?, 0, ?, ?)',
            (generate_uuid(), new_id, srow['titlu'], srow['ordine'], now)
        )
    return new_id

# ---------------------------------------------------------------------------
# Project tasks CRUD
# ---------------------------------------------------------------------------

@tasks_bp.route('/api/proiecte/<project_id>/tasks', methods=['GET'])
@login_required
def get_tasks(project_id):
    conn = get_db()
    cursor = conn.cursor()

    # 1) Fetch all tasks for this project (single query).
    cursor.execute('''
        SELECT t.* FROM tasks t WHERE t.proiect_id = ?
        ORDER BY t.ordine ASC, t.created_at DESC
    ''', (project_id,))
    rows = cursor.fetchall()
    if not rows:
        conn.close()
        return jsonify([])

    task_ids = [r['id'] for r in rows]
    placeholders = ','.join('?' * len(task_ids))

    # 2) Batch-fetch subtask counts (one query for all tasks).
    cursor.execute(f'''
        SELECT task_id,
               COUNT(*) AS subtask_total,
               SUM(CASE WHEN done = 1 THEN 1 ELSE 0 END) AS subtask_done
        FROM task_subtasks
        WHERE task_id IN ({placeholders})
        GROUP BY task_id
    ''', task_ids)
    subtask_map = {}
    for r in cursor.fetchall():
        subtask_map[r['task_id']] = {'subtask_total': r['subtask_total'],
                                      'subtask_done': r['subtask_done']}

    conn.close()

    # 4) Merge results in Python.
    result = []
    for row in rows:
        d = row_to_dict(row)
        sc = subtask_map.get(d['id'], {})
        d['subtask_total'] = sc.get('subtask_total', 0)
        d['subtask_done'] = sc.get('subtask_done', 0)
        result.append(d)

    return jsonify(result)


@tasks_bp.route('/api/proiecte/<project_id>/tasks', methods=['POST'])
@login_required
def create_task(project_id):
    data = get_json_or_400()
    conn = get_db()
    cursor = conn.cursor()

    now = datetime.now().isoformat()
    task_id = data.get('id') or generate_uuid()

    # Get max ordine for this project
    cursor.execute('SELECT MAX(ordine) as max_ordine FROM tasks WHERE proiect_id = ?', (project_id,))
    result = cursor.fetchone()
    max_ordine = result['max_ordine'] if result and result['max_ordine'] is not None else 0

    cursor.execute('''
        INSERT INTO tasks (id, proiect_id, titlu, status, data_scadenta,
                           data_finalizare, ordine, created_at, descriere, recurenta, updated_at,
                           ordine_agenda, data_start, progres, is_milestone)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        task_id,
        project_id,
        data.get('titlu', ''),
        data.get('status', 'to_do'),
        data.get('data_scadenta', ''),
        data.get('data_finalizare', ''),
        max_ordine + 1,
        now,
        data.get('descriere', ''),
        data.get('recurenta', ''),
        now,
        data.get('ordine_agenda', 0),
        data.get('data_start', ''),
        data.get('progres', 0),
        1 if data.get('is_milestone') else 0
    ))

    conn.commit()
    conn.close()

    return jsonify({'id': task_id}), 201


@tasks_bp.route('/api/tasks/<task_id>', methods=['PUT'])
@login_required
def update_task(task_id):
    data = get_json_or_400()
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM tasks WHERE id = ?', (task_id,))
    existing = cursor.fetchone()
    if existing is None:
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    old_status = existing['status']

    # Tranzitia la 'done' e atomica (WHERE status != 'done'): doua PUT-uri
    # concurente (dublu-click / 2 workere) nu mai spawneaza doua recurente.
    transitioned_to_done = False
    if data.get('status') == 'done' and old_status != 'done':
        cursor.execute("UPDATE tasks SET status = 'done' WHERE id = ? AND status != 'done'", (task_id,))
        transitioned_to_done = cursor.rowcount == 1

    cursor.execute('''
        UPDATE tasks SET
            titlu = COALESCE(?, titlu),
            status = COALESCE(?, status),
            data_scadenta = COALESCE(?, data_scadenta),
            data_finalizare = COALESCE(?, data_finalizare),
            ordine = COALESCE(?, ordine),
            descriere = COALESCE(?, descriere),
            recurenta = COALESCE(?, recurenta),
            ordine_agenda = COALESCE(?, ordine_agenda),
            data_start = COALESCE(?, data_start),
            progres = COALESCE(?, progres),
            is_milestone = COALESCE(?, is_milestone),
            updated_at = ?
        WHERE id = ?
    ''', (
        data.get('titlu'),
        data.get('status'),
        data.get('data_scadenta'),
        data.get('data_finalizare'),
        data.get('ordine'),
        data.get('descriere'),
        data.get('recurenta'),
        data.get('ordine_agenda'),
        data.get('data_start'),
        data.get('progres'),
        (1 if data.get('is_milestone') else 0) if data.get('is_milestone') is not None else None,
        datetime.now().isoformat(),
        task_id
    ))

    # A recurring task just completed -> spawn the next occurrence.
    spawned_id = None
    next_scad = None
    if transitioned_to_done and (existing['recurenta'] or '').strip():
        recurenta = existing['recurenta'].strip()
        spawned_id = _spawn_recurring_task(cursor, existing, recurenta)
        next_scad = _next_recurrence_date(existing['data_scadenta'] or '', recurenta)

    conn.commit()
    conn.close()

    resp = {'message': 'Task updated'}
    if spawned_id:
        resp['recurring_spawned'] = spawned_id
        resp['recurring_next'] = next_scad
    return jsonify(resp)


@tasks_bp.route('/api/tasks/<task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM task_subtasks WHERE task_id = ?', (task_id,))
        cursor.execute('DELETE FROM task_dependencies WHERE predecessor_id = ? OR successor_id = ?', (task_id, task_id))
        cursor.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
        deleted = cursor.rowcount
        conn.commit()
        if deleted == 0:
            return jsonify({'error': 'Task not found'}), 404
        return jsonify({'message': 'Task deleted'})
    finally:
        conn.close()

# ---------------------------------------------------------------------------
# Project Gantt: schedule (start/end/progress/milestone) + dependencies
# ---------------------------------------------------------------------------

def _effective_progress(task_row, sub_total, sub_done):
    """Display % for a task: subtasks ratio when it has them, else the manual
    `progres` column, else status (done => 100). Keeps the bar honest."""
    status = task_row.get('status') or 'to_do'
    if status in ('done', 'finalizat'):
        return 100
    if sub_total and sub_total > 0:
        return int(round(100 * (sub_done or 0) / sub_total))
    return int(task_row.get('progres') or 0)


@tasks_bp.route('/api/proiecte/<project_id>/gantt', methods=['GET'])
@login_required
def get_project_gantt(project_id):
    """Everything the per-project Gantt needs: project meta, its tasks with the
    effective schedule (start/end/progress/milestone + subtask counts) and the
    dependency links between them."""
    data = _collect_gantt(project_id)
    if data is None:
        return jsonify({'error': 'Proiect inexistent'}), 404
    return jsonify(data)


def _collect_gantt(project_id):
    """Assemble the Gantt payload (proiect, tasks[], dependencies[]) or None if the
    project doesn't exist. Shared by the JSON endpoint and the PDF/Excel exports."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (project_id,))
    proj = cursor.fetchone()
    if proj is None:
        conn.close()
        return None
    proj = row_to_dict(proj)

    cursor.execute('SELECT * FROM tasks WHERE proiect_id = ? ORDER BY ordine ASC, created_at ASC', (project_id,))
    rows = [row_to_dict(r) for r in cursor.fetchall()]

    sub_map = {}
    if rows:
        ids = [r['id'] for r in rows]
        ph = ','.join('?' * len(ids))
        cursor.execute(f'''SELECT task_id, COUNT(*) AS tot, SUM(CASE WHEN done=1 THEN 1 ELSE 0 END) AS done
                           FROM task_subtasks WHERE task_id IN ({ph}) GROUP BY task_id''', ids)
        for r in cursor.fetchall():
            sub_map[r['task_id']] = (r['tot'], r['done'])

    tasks = []
    for r in rows:
        tot, done = sub_map.get(r['id'], (0, 0))
        start = (r.get('data_start') or '').strip()[:10]
        end = (r.get('data_scadenta') or '').strip()[:10]
        if not start and end:
            start = end
        if not end and start:
            end = start
        tasks.append({
            'id': r['id'], 'titlu': r.get('titlu') or '', 'status': r.get('status') or 'to_do',
            'data_start': start, 'data_scadenta': end,
            'is_milestone': bool(r.get('is_milestone')),
            'progres': _effective_progress(r, tot, done),
            'progres_manual': int(r.get('progres') or 0),
            'subtask_total': tot or 0, 'subtask_done': done or 0,
            'ordine': r.get('ordine') or 0,
            'recurenta': r.get('recurenta') or '',
        })

    cursor.execute('SELECT * FROM task_dependencies WHERE proiect_id = ?', (project_id,))
    deps = [{'id': d['id'], 'predecessor_id': d['predecessor_id'], 'successor_id': d['successor_id'],
             'tip': d['tip'] or 'FS', 'lag': d['lag'] or 0} for d in [row_to_dict(x) for x in cursor.fetchall()]]

    cursor.execute('SELECT * FROM implementari WHERE proiect_id = ? ORDER BY data_start ASC, ordine ASC', (project_id,))
    # `faza` merge la client aici din acelasi motiv ca in `/api/plan`: Gantt-ul
    # taie golul de pregatire DOAR pe etapele de implementare, iar o zi de
    # pregatire blocata explicit nu e etapa.
    implementari = [{'id': d['id'], 'data_start': (d.get('data_start') or '')[:10],
                     'data_sfarsit': (d.get('data_sfarsit') or '')[:10],
                     'locatie': d.get('locatie') or 'site', 'eticheta': d.get('eticheta') or '',
                     'faza': d.get('faza') or 'implementare'}
                    for d in [row_to_dict(x) for x in cursor.fetchall()]]

    conn.close()
    return {
        'proiect': {'id': proj['id'], 'nume': proj.get('nume'), 'client': proj.get('client'),
                    'tip': proj.get('tip'), 'cod_proiect': proj.get('cod_proiect'),
                    'status': proj.get('status'), 'locatie': proj.get('locatie')},
        'tasks': tasks, 'dependencies': deps, 'implementari': implementari,
    }


def _would_cycle(cursor, project_id, pred, succ):
    """True if adding pred->succ creates a cycle (succ already reaches pred)."""
    # Walk successors of `succ`; if we reach `pred`, it's a cycle.
    seen = set()
    stack = [succ]
    while stack:
        cur = stack.pop()
        if cur == pred:
            return True
        if cur in seen:
            continue
        seen.add(cur)
        cursor.execute('SELECT successor_id FROM task_dependencies WHERE predecessor_id = ?', (cur,))
        stack.extend(row['successor_id'] for row in cursor.fetchall())
    return False


@tasks_bp.route('/api/proiecte/<project_id>/dependencies', methods=['POST'])
@login_required
def create_dependency(project_id):
    data = get_json_or_400()
    pred = (data.get('predecessor_id') or '').strip()
    succ = (data.get('successor_id') or '').strip()
    tip = (data.get('tip') or 'FS').strip().upper()
    if tip not in ('FS', 'SS', 'FF', 'SF'):
        tip = 'FS'
    try:
        lag = int(data.get('lag') or 0)
    except (TypeError, ValueError):
        lag = 0
    if not pred or not succ or pred == succ:
        return jsonify({'error': 'Dependenta invalida (predecesor = succesor).'}), 400

    conn = get_db()
    cursor = conn.cursor()
    # Both tasks must belong to this project.
    cursor.execute('SELECT id FROM tasks WHERE id IN (?, ?) AND proiect_id = ?', (pred, succ, project_id))
    if len({row['id'] for row in cursor.fetchall()}) != 2:
        conn.close()
        return jsonify({'error': 'Ambele taskuri trebuie sa fie din acest proiect.'}), 400
    # No duplicates (either direction of the same pair is redundant).
    cursor.execute('''SELECT id FROM task_dependencies
                      WHERE proiect_id = ? AND ((predecessor_id = ? AND successor_id = ?)
                                             OR (predecessor_id = ? AND successor_id = ?))''',
                   (project_id, pred, succ, succ, pred))
    if cursor.fetchone():
        conn.close()
        return jsonify({'error': 'Dependenta exista deja.'}), 409
    if _would_cycle(cursor, project_id, pred, succ):
        conn.close()
        return jsonify({'error': 'Dependenta ar crea un ciclu.'}), 400

    dep_id = generate_uuid()
    cursor.execute('''INSERT INTO task_dependencies (id, proiect_id, predecessor_id, successor_id, tip, lag, created_at)
                      VALUES (?, ?, ?, ?, ?, ?, ?)''',
                   (dep_id, project_id, pred, succ, tip, lag, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'id': dep_id, 'predecessor_id': pred, 'successor_id': succ, 'tip': tip, 'lag': lag}), 201


@tasks_bp.route('/api/dependencies/<dep_id>', methods=['DELETE'])
@login_required
def delete_dependency(dep_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM task_dependencies WHERE id = ?', (dep_id,))
    deleted = cursor.rowcount
    conn.commit()
    conn.close()
    if deleted == 0:
        return jsonify({'error': 'Dependenta inexistenta'}), 404
    return jsonify({'message': 'ok'})


def _reschedule_project(cursor, project_id):
    """Forward-pass (as-soon-as-possible) scheduler: push each successor so it
    respects its dependencies (FS/SS/FF/SF + lag), preserving durations. Only
    moves tasks LATER (never earlier). Returns how many tasks were shifted."""
    from collections import deque
    cursor.execute('SELECT id, data_start, data_scadenta, is_milestone FROM tasks WHERE proiect_id = ?', (project_id,))
    tasks = {}
    for r in [row_to_dict(x) for x in cursor.fetchall()]:
        tasks[r['id']] = {'start': _pdate(r['data_start']), 'end': _pdate(r['data_scadenta'])}
    cursor.execute('SELECT predecessor_id, successor_id, tip, lag FROM task_dependencies WHERE proiect_id = ?', (project_id,))
    deps = [row_to_dict(x) for x in cursor.fetchall()]

    succ_of, indeg, preds_of = {}, {t: 0 for t in tasks}, {t: [] for t in tasks}
    for d in deps:
        p, s = d['predecessor_id'], d['successor_id']
        if p not in tasks or s not in tasks:
            continue
        succ_of.setdefault(p, []).append(s)
        preds_of[s].append(d)
        indeg[s] += 1

    q = deque([t for t in tasks if indeg[t] == 0])
    order = []
    while q:
        n = q.popleft()
        order.append(n)
        for s in succ_of.get(n, []):
            indeg[s] -= 1
            if indeg[s] == 0:
                q.append(s)

    now = datetime.now().isoformat()
    changed = 0
    for tid in order:
        t = tasks[tid]
        if not t['start'] or not t['end']:
            continue
        dur = (t['end'] - t['start']).days
        req = t['start']
        for d in preds_of[tid]:
            p = tasks.get(d['predecessor_id'])
            if not p or not p['start'] or not p['end']:
                continue
            lag = int(d['lag'] or 0)
            tip = (d['tip'] or 'FS').upper()
            if tip == 'SS':
                cand = p['start'] + timedelta(days=lag)
            elif tip == 'FF':
                cand = p['end'] + timedelta(days=lag) - timedelta(days=dur)
            elif tip == 'SF':
                cand = p['start'] + timedelta(days=lag) - timedelta(days=dur)
            else:  # FS (default): successor starts the day after predecessor finishes
                cand = p['end'] + timedelta(days=1 + lag)
            if cand > req:
                req = cand
        if req > t['start']:
            delta = (req - t['start']).days
            t['start'] = req
            t['end'] = t['end'] + timedelta(days=delta)
            cursor.execute('UPDATE tasks SET data_start = ?, data_scadenta = ?, updated_at = ? WHERE id = ?',
                           (t['start'].isoformat(), t['end'].isoformat(), now, tid))
            changed += 1
    return changed


@tasks_bp.route('/api/proiecte/<project_id>/reschedule', methods=['POST'])
@login_required
def reschedule_project(project_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM proiecte WHERE id = ?', (project_id,))
    if cursor.fetchone() is None:
        conn.close()
        return jsonify({'error': 'Proiect inexistent'}), 404
    changed = _reschedule_project(cursor, project_id)
    conn.commit()
    conn.close()
    return jsonify({'message': 'ok', 'changed': changed})


# ---------------------------------------------------------------------------
# Gantt export — client-ready PDF (reportlab) + Excel (openpyxl), server-side
# so the output is consistent and branded, independent of the browser.
# ---------------------------------------------------------------------------

_STATUS_RO = {'done': 'Finalizat', 'finalizat': 'Finalizat', 'in_progress': 'In lucru',
              'in_lucru': 'In lucru', 'to_do': 'De facut', 'in_asteptare': 'In asteptare',
              'blocat': 'Blocat'}
# Colors chosen to read on white paper / Excel (client-facing).
_STATUS_HEX = {'done': '4FA874', 'finalizat': '4FA874', 'in_progress': 'E0912E',
               'in_lucru': 'E0912E', 'to_do': 'B9B2A6', 'blocat': 'D65A4A', 'in_asteptare': '8B7FC0'}


def _pdate(s):
    try:
        return datetime.strptime((s or '')[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _gantt_window(data):
    """Padded [start, end] date window covering all tasks + project dates."""
    ds = []
    for t in data['tasks']:
        for k in ('data_start', 'data_scadenta'):
            d = _pdate(t.get(k))
            if d:
                ds.append(d)
    if not ds:
        today = datetime.now().date()
        return today, today + timedelta(days=14)
    lo, hi = min(ds), max(ds)
    return lo - timedelta(days=2), hi + timedelta(days=2)


def _fmt_ro(d):
    return d.strftime('%d.%m.%Y') if d else '—'


def _pred_titles(data):
    """{successor_id: [predecessor titlu, ...]} for the 'Depinde de' column."""
    by_id = {t['id']: t['titlu'] for t in data['tasks']}
    out = {}
    for dep in data['dependencies']:
        out.setdefault(dep['successor_id'], []).append(by_id.get(dep['predecessor_id'], '?'))
    return out


def _draw_logo(c, x, y, s):
    """Draw the PIF 'ramp area-chart' mark (amber tile + white accelerating curve
    + operating-point dot) in an s×s box with bottom-left at (x, y)."""
    from reportlab.lib.colors import HexColor, white
    c.setFillColor(HexColor('#ffb454'))
    c.roundRect(x, y, s, s, s * 0.22, fill=1, stroke=0)

    def P(gx, gy):  # 64-grid SVG coords (y-down) -> reportlab (y-up)
        return (x + gx / 64.0 * s, y + s - gy / 64.0 * s)

    c.setStrokeColor(white)
    c.setLineWidth(max(1.2, s * 0.085))
    c.setLineCap(1)
    c.setLineJoin(1)
    p = c.beginPath()
    p.moveTo(*P(14, 48))
    p.curveTo(*P(27, 48), *P(30, 40), *P(33, 30))
    p.curveTo(*P(36, 20), *P(42, 16), *P(50, 16))
    c.drawPath(p, stroke=1, fill=0)
    dx, dy = P(50, 16)
    c.setFillColor(white)
    c.circle(dx, dy, max(1.4, s * 0.06), fill=1, stroke=0)


@tasks_bp.route('/api/proiecte/<project_id>/gantt.pdf', methods=['GET'])
@login_required
def export_gantt_pdf(project_id):
    data = _collect_gantt(project_id)
    if data is None:
        return jsonify({'error': 'Proiect inexistent'}), 404

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors

    proj = data['proiect']
    win_start, win_end = _gantt_window(data)
    total = max((win_end - win_start).days, 1)
    today = datetime.now().date()

    # Gruparea pe faze (WBS) a plecat in v32: coloana tasks.faza nu putea fi
    # completata din nicio interfata, deci antetele de faza nu s-au randat niciodata.
    render_rows = [('task', t) for t in data['tasks']]
    # Implementation periods (Site / Sediu EGB) as distinct bands at the top.
    impl_rows = [('impl', im) for im in data.get('implementari', [])]
    render_rows = impl_rows + render_rows

    buf = BytesIO()
    W, H = landscape(A4)
    c = canvas.Canvas(buf, pagesize=(W, H))
    m = 1.3 * cm
    lbl_w = 5.6 * cm
    x0 = m + lbl_w            # timeline left
    x1 = W - m               # timeline right
    tw = x1 - x0

    def xfor(d):
        if not d:
            return x0
        frac = max(0.0, min(1.0, (d - win_start).days / total))
        return x0 + frac * tw

    # --- title block (logo + project + meta) ---
    y = H - m
    logo_sz = 24
    _draw_logo(c, m, y - logo_sz + 2, logo_sz)
    tx = m + logo_sz + 9
    c.setFillColor(colors.HexColor('#1a1206'))
    c.setFont('Helvetica-Bold', 15)
    c.drawString(tx, y - 4, f"{proj.get('nume') or 'Proiect'}")
    c.setFont('Helvetica', 9)
    c.setFillColor(colors.HexColor('#555555'))
    meta = []
    if proj.get('client'):
        meta.append(f"Client: {proj['client']}")
    if proj.get('cod_proiect'):
        meta.append(f"Cod: {proj['cod_proiect']}")
    if proj.get('tip'):
        meta.append(proj['tip'])
    meta.append(f"Perioada: {_fmt_ro(win_start)} – {_fmt_ro(win_end)}")
    c.drawString(tx, y - 20, "   ·   ".join(meta))
    c.setFont('Helvetica-Bold', 7.5)
    c.setFillColor(colors.HexColor('#8a5300'))
    c.drawRightString(x1, y - 3, "PIF DASHBOARD")
    c.setFont('Helvetica', 7.5)
    c.setFillColor(colors.HexColor('#777777'))
    c.drawRightString(x1, y - 14, f"Generat: {_fmt_ro(today)}")
    top = y - 34          # chart top

    # --- chart geometry ---
    header_h = 16
    avail = top - m - 26           # leave room for legend at bottom
    n = max(len(render_rows), 1)
    row_h = min(20, max(10, (avail - header_h) / n))
    chart_h = header_h + row_h * len(render_rows)

    # week gridlines + labels
    c.setFont('Helvetica', 6.5)
    d = win_start
    # advance to Monday
    d = d - timedelta(days=(d.weekday()))
    while d <= win_end:
        gx = xfor(d)
        c.setStrokeColor(colors.HexColor('#E2E0DA'))
        c.setLineWidth(0.5)
        c.line(gx, top - header_h, gx, top - chart_h)
        c.setFillColor(colors.HexColor('#888888'))
        if gx < x1 - 12:
            c.drawString(gx + 2, top - 11, d.strftime('%d.%m'))
        d += timedelta(days=7)
    # frame + header separator
    c.setStrokeColor(colors.HexColor('#CFCabf'))
    c.setLineWidth(0.8)
    c.rect(x0, top - chart_h, tw, chart_h, stroke=1, fill=0)
    c.line(m, top - header_h, x1, top - header_h)
    c.line(x0, top - header_h, x0, top - chart_h)  # label/timeline divider

    # today line
    if win_start <= today <= win_end:
        tx = xfor(today)
        c.setStrokeColor(colors.HexColor('#D64A3A'))
        c.setLineWidth(1)
        c.line(tx, top - header_h, tx, top - chart_h)

    # rows (implementari + tasks)
    pos = {}
    _IMPL_HEX = {'site': '3F9DC4', 'sediu': 'C99A3A'}
    for i, (kind, val) in enumerate(render_rows):
        ry = top - header_h - (i + 1) * row_h
        cy = ry + row_h / 2
        if kind == 'impl':
            s, e = _pdate(val.get('data_start')), _pdate(val.get('data_sfarsit'))
            loc = val.get('locatie') or 'site'
            lab = ('Sediu EGB' if loc == 'sediu' else 'Site') + (f" · {val['eticheta']}" if val.get('eticheta') else '')
            c.setFillColor(colors.HexColor('#222222'))
            c.setFont('Helvetica-Bold', 7.5)
            c.drawString(m + 4, cy - 3, lab[:44])
            if s and e:
                bx, bw = xfor(s), max(xfor(e) - xfor(s), 2)
                c.setFillColor(colors.HexColor('#' + _IMPL_HEX.get(loc, '3F9DC4')))
                c.roundRect(bx, cy - 5, bw, 10, 2, stroke=0, fill=1)
                c.setFillColor(colors.HexColor('#10130f'))
                c.setFont('Helvetica-Bold', 6)
                if bw > 40:
                    c.drawString(bx + 4, cy - 2, lab[:26])
            continue
        t = val
        pos[t['id']] = i
        # zebra
        if i % 2 == 1:
            c.setFillColor(colors.HexColor('#FAF7F1'))
            c.rect(m, ry, x1 - m, row_h, stroke=0, fill=1)
        c.setFillColor(colors.HexColor('#222222'))
        c.setFont('Helvetica', 8)
        label = t['titlu'][:44]
        c.drawString(m + 4, cy - 3, ('◆ ' if t['is_milestone'] else '') + label)
        # bar / milestone
        s, e = _pdate(t['data_start']), _pdate(t['data_scadenta'])
        hexc = _STATUS_HEX.get(t['status'], 'B9B2A6')
        if t['is_milestone'] and s:
            mx, my, r = xfor(s), cy, 4
            c.setFillColor(colors.HexColor('#B45309'))
            p = c.beginPath(); p.moveTo(mx, my + r); p.lineTo(mx + r, my); p.lineTo(mx, my - r); p.lineTo(mx - r, my); p.close()
            c.drawPath(p, fill=1, stroke=0)
        elif s and e:
            bx, bw = xfor(s), max(xfor(e) - xfor(s), 2)
            bh = row_h * 0.5
            by = cy - bh / 2
            # base bar (light)
            c.setFillColor(colors.HexColor('#' + hexc))
            c.setFillAlpha(0.28)
            c.roundRect(bx, by, bw, bh, 2, stroke=0, fill=1)
            # progress fill
            c.setFillAlpha(1)
            pw = bw * (t['progres'] / 100.0)
            if pw > 0:
                c.roundRect(bx, by, max(pw, 1.5), bh, 2, stroke=0, fill=1)
            c.setFillAlpha(1)
            c.setFillColor(colors.HexColor('#333333'))
            c.setFont('Helvetica', 6)
            c.drawString(bx + bw + 3, cy - 2, f"{t['progres']}%")

    # dependency lines (thin)
    # `stasks` era lista de taskuri sortata pe faze; a plecat cu `tasks.faza` in v32,
    # dar referinta a rămas si exportul PDF al Ganttului dadea 500 de atunci. Randurile
    # desenate sunt in `render_rows`, iar cele de tip perioada nu au dependinte.
    by_id = {val['id']: val for kind, val in render_rows if kind != 'impl'}
    c.setStrokeColor(colors.HexColor('#9A9488'))
    c.setLineWidth(0.6)
    for dep in data['dependencies']:
        pi, si = pos.get(dep['predecessor_id']), pos.get(dep['successor_id'])
        if pi is None or si is None:
            continue
        p, s = by_id[dep['predecessor_id']], by_id[dep['successor_id']]
        pe = _pdate(p['data_scadenta']); ss = _pdate(s['data_start'])
        if not pe or not ss:
            continue
        y1 = top - header_h - (pi + 1) * row_h + row_h / 2
        y2 = top - header_h - (si + 1) * row_h + row_h / 2
        xa, xb = xfor(pe), xfor(ss)
        c.line(xa, y1, xa + 6, y1); c.line(xa + 6, y1, xa + 6, y2); c.line(xa + 6, y2, xb, y2)
        c.setFillColor(colors.HexColor('#9A9488'))
        ah = c.beginPath(); ah.moveTo(xb, y2); ah.lineTo(xb - 4, y2 + 2.5); ah.lineTo(xb - 4, y2 - 2.5); ah.close()
        c.drawPath(ah, fill=1, stroke=0)

    # legend
    ly = m + 6
    lx = m
    c.setFont('Helvetica', 7)
    for key, lab in (('done', 'Finalizat'), ('in_progress', 'In lucru'), ('to_do', 'De facut')):
        c.setFillColor(colors.HexColor('#' + _STATUS_HEX[key])); c.roundRect(lx, ly, 14, 7, 1, stroke=0, fill=1)
        c.setFillColor(colors.HexColor('#555555')); c.drawString(lx + 18, ly + 1, lab)
        lx += 70
    c.setFillColor(colors.HexColor('#B45309'))
    p = c.beginPath(); p.moveTo(lx + 4, ly + 7); p.lineTo(lx + 8, ly + 3.5); p.lineTo(lx + 4, ly); p.lineTo(lx, ly + 3.5); p.close()
    c.drawPath(p, fill=1, stroke=0)
    c.setFillColor(colors.HexColor('#555555')); c.drawString(lx + 14, ly + 1, 'Milestone')

    # signature block (bottom-right) — for client sign-off
    c.setStrokeColor(colors.HexColor('#AAAAAA'))
    c.setLineWidth(0.6)
    c.setFont('Helvetica', 8)
    sy = m + 8
    sig_w = 150
    c2x = x1 - sig_w
    c1x = c2x - sig_w - 40
    for label, cx in (('Intocmit', c1x), ('Aprobat client', c2x)):
        c.line(cx, sy + 16, cx + sig_w, sy + 16)
        c.setFillColor(colors.HexColor('#777777'))
        c.drawString(cx, sy + 5, label)
        c.drawRightString(cx + sig_w, sy + 5, 'Data: __________')

    c.showPage()
    c.save()
    buf.seek(0)
    safe = (proj.get('nume') or 'proiect').replace('/', '-').replace(' ', '_')[:40]
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'Gantt_{safe}.pdf')


@tasks_bp.route('/api/proiecte/<project_id>/gantt.xlsx', methods=['GET'])
@login_required
def export_gantt_xlsx(project_id):
    data = _collect_gantt(project_id)
    if data is None:
        return jsonify({'error': 'Proiect inexistent'}), 404

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    proj = data['proiect']
    tasks = list(data['tasks'])
    win_start, win_end = _gantt_window(data)

    # Day columns when the span is short enough; weekly for long projects (else
    # a 6-month plan would blow out to 180 columns).
    total_days = (win_end - win_start).days + 1
    daily = total_days <= 92
    if daily:
        units = [win_start + timedelta(days=i) for i in range(total_days)]
    else:
        u0 = win_start - timedelta(days=win_start.weekday())
        units, dd = [], u0
        while dd <= win_end:
            units.append(dd)
            dd += timedelta(days=7)

    def unit_index(d):
        if d < units[0]:
            return 0
        if daily:
            return min((d - win_start).days, len(units) - 1)
        return min((d - units[0]).days // 7, len(units) - 1)

    # bar shades: (done-portion, remaining-portion) per status
    SHADES = {
        'done': ('4FA874', '4FA874'), 'finalizat': ('4FA874', '4FA874'),
        'in_progress': ('D9822B', 'F3D9AE'), 'in_lucru': ('D9822B', 'F3D9AE'),
        'to_do': ('B9B2A6', 'DED9D0'), 'blocat': ('D65A4A', 'ECC4BE'),
        'in_asteptare': ('8B7FC0', 'D9D3EE'),
    }
    today = datetime.now().date()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gantt'

    thin = Side(style='thin', color='E2DED6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    # title block
    ws['A1'] = proj.get('nume') or 'Proiect'
    ws['A1'].font = Font(bold=True, size=14)
    info = []
    if proj.get('client'):
        info.append(f"Client: {proj['client']}")
    if proj.get('cod_proiect'):
        info.append(f"Cod: {proj['cod_proiect']}")
    info.append(f"Perioada: {_fmt_ro(win_start)} – {_fmt_ro(win_end)}")
    info.append(f"Generat: {_fmt_ro(today)}")
    ws['A2'] = "   ·   ".join(info)
    ws['A2'].font = Font(size=9, color='666666')

    INFO = ['#', 'Task', 'Start', 'Sfarsit', 'Zile', '%']
    ncol_info = len(INFO)
    grid0 = ncol_info + 1
    HR_MONTH, HR_DAY, FIRST = 4, 5, 6

    hdr_fill = PatternFill('solid', fgColor='1A1206')
    hdr_font = Font(bold=True, color='FFFFFF')
    # info headers on the day-header row
    for ci, val in enumerate(INFO, start=1):
        c = ws.cell(HR_DAY, ci, val)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

    # month band (merged) + day/week header
    month_start = 0
    def flush_month(a, b):
        if b < a:
            return
        ca, cb = grid0 + a, grid0 + b
        if cb > ca:
            ws.merge_cells(start_row=HR_MONTH, start_column=ca, end_row=HR_MONTH, end_column=cb)
        _romon = ['Ian', 'Feb', 'Mar', 'Apr', 'Mai', 'Iun', 'Iul', 'Aug', 'Sep', 'Oct', 'Noi', 'Dec']
        mc = ws.cell(HR_MONTH, ca, f"{_romon[units[a].month - 1]} {units[a].year}")
        mc.alignment = center; mc.font = Font(size=8, bold=True, color='555555')
    for k, u in enumerate(units):
        col = grid0 + k
        is_we = daily and u.weekday() >= 5
        is_today = daily and u == today
        dc = ws.cell(HR_DAY, col, u.day if daily else u.strftime('%d.%m'))
        dc.alignment = center; dc.border = border
        dc.font = Font(size=7, color='C0392B' if is_today else ('999999' if is_we else '333333'), bold=is_today)
        if is_today:
            dc.fill = PatternFill('solid', fgColor='FCE9C8')
        elif is_we:
            dc.fill = PatternFill('solid', fgColor='F1EEE8')
        if k > 0 and (units[k].month != units[month_start].month):
            flush_month(month_start, k - 1); month_start = k
    flush_month(month_start, len(units) - 1)

    # implementation-period rows (Site / Sediu EGB) at the top
    impl = data.get('implementari', [])
    _IMPL_HEX = {'site': '3F9DC4', 'sediu': 'C99A3A'}
    for j, im in enumerate(impl):
        r = FIRST + j
        s, e = _pdate(im.get('data_start')), _pdate(im.get('data_sfarsit'))
        loc = im.get('locatie') or 'site'
        lab = ('Sediu EGB' if loc == 'sediu' else 'Site (santier)') + (f" · {im['eticheta']}" if im.get('eticheta') else '')
        cell = ws.cell(r, 2, lab)
        cell.font = Font(bold=True, color=_IMPL_HEX.get(loc, '3F9DC4'))
        for ci in range(1, ncol_info + 1):
            ws.cell(r, ci).border = border
        for k in range(len(units)):
            ws.cell(r, grid0 + k).border = border
        if s and e:
            si, ei = unit_index(s), unit_index(e)
            for k in range(si, ei + 1):
                ws.cell(r, grid0 + k).fill = PatternFill('solid', fgColor=_IMPL_HEX.get(loc, '3F9DC4'))
    task_first = FIRST + len(impl)

    # task rows
    for i, t in enumerate(tasks):
        r = task_first + i
        s, e = _pdate(t['data_start']), _pdate(t['data_scadenta'])
        zile = ((e - s).days + 1) if (s and e) else ''
        info_vals = [i + 1, t['titlu'], _fmt_ro(s), _fmt_ro(e), '◆' if t['is_milestone'] else zile, t['progres']]
        for ci, val in enumerate(info_vals, start=1):
            c = ws.cell(r, ci, val)
            c.border = border
            if ci not in (2, 3):
                c.alignment = center
        # grid cells (light borders everywhere for the chart look)
        for k in range(len(units)):
            ws.cell(r, grid0 + k).border = border
        if not s or not e:
            continue
        si, ei = unit_index(s), unit_index(e)
        done_hex, rem_hex = SHADES.get(t['status'], ('B9B2A6', 'DED9D0'))
        if t['is_milestone']:
            mc = ws.cell(r, grid0 + si, '◆')
            mc.alignment = center; mc.font = Font(color='B45309', bold=True)
            continue
        barlen = ei - si + 1
        done_cells = round(barlen * (t['progres'] / 100.0))
        for k in range(si, ei + 1):
            filled = (k - si) < done_cells
            ws.cell(r, grid0 + k).fill = PatternFill('solid', fgColor=done_hex if filled else rem_hex)

    # widths + freeze
    for col, w in {1: 4, 2: 30, 3: 15, 4: 11, 5: 11, 6: 6, 7: 6}.items():
        ws.column_dimensions[ws.cell(HR_DAY, col).column_letter].width = w
    gw = 3.0 if daily else 6.0
    for k in range(len(units)):
        ws.column_dimensions[ws.cell(HR_DAY, grid0 + k).column_letter].width = gw
    ws.freeze_panes = ws.cell(FIRST, grid0)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    safe = (proj.get('nume') or 'proiect').replace('/', '-').replace(' ', '_')[:40]
    return send_file(out, as_attachment=True, download_name=f'Gantt_{safe}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------------------------------------------------------------------------
# Task subtasks (lightweight checklist under a task)
# ---------------------------------------------------------------------------

@tasks_bp.route('/api/tasks/<task_id>/subtasks', methods=['GET'])
@login_required
def get_subtasks(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM task_subtasks WHERE task_id = ? ORDER BY ordine ASC, created_at ASC', (task_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@tasks_bp.route('/api/tasks/<task_id>/subtasks', methods=['POST'])
@login_required
def create_subtask(task_id):
    data = get_json_or_400()
    titlu = (data.get('titlu') or '').strip()
    if not titlu:
        return jsonify({'error': 'Titlu required'}), 400
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT COALESCE(MAX(ordine), -1) + 1 FROM task_subtasks WHERE task_id = ?', (task_id,))
        next_ordine = cursor.fetchone()[0]
        sid = generate_uuid()
        cursor.execute(
            'INSERT INTO task_subtasks (id, task_id, titlu, done, ordine, created_at) VALUES (?, ?, ?, 0, ?, ?)',
            (sid, task_id, titlu, next_ordine, datetime.now().isoformat())
        )
        conn.commit()
        return jsonify({'id': sid}), 201
    finally:
        conn.close()


@tasks_bp.route('/api/subtasks/<subtask_id>', methods=['PUT'])
@login_required
def update_subtask(subtask_id):
    data = get_json_or_400()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'UPDATE task_subtasks SET titlu = COALESCE(?, titlu), done = COALESCE(?, done) WHERE id = ?',
        (data.get('titlu'), data.get('done'), subtask_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@tasks_bp.route('/api/subtasks/<subtask_id>', methods=['DELETE'])
@login_required
def delete_subtask(subtask_id):
    conn = get_db()
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM task_subtasks WHERE id = ?', (subtask_id,))
        deleted = cursor.rowcount
        conn.commit()
        if deleted == 0:
            return jsonify({'error': 'Subtask not found'}), 404
        return jsonify({'ok': True})
    finally:
        conn.close()

# ---------------------------------------------------------------------------
# Global tasks CRUD
# ---------------------------------------------------------------------------

@tasks_bp.route('/api/global-tasks', methods=['GET'])
@login_required
def get_global_tasks():
    conn = get_db()
    cursor = conn.cursor()

    status = request.args.get('status')

    categorie = request.args.get('categorie')
    arhiva = request.args.get('arhiva')

    # 1) Fetch the global tasks (single query, no correlated subqueries).
    query = 'SELECT g.* FROM global_tasks g WHERE 1=1'
    params = []

    if arhiva == 'true':
        query += " AND status = 'done'"
    else:
        query += " AND status != 'done'"

    if status and arhiva != 'true':
        query += ' AND status = ?'
        params.append(status)
    if categorie:
        query += ' AND categorie = ?'
        params.append(categorie)

    query += ' ORDER BY created_at DESC'

    cursor.execute(query, params)
    rows = cursor.fetchall()
    if not rows:
        conn.close()
        return jsonify([])

    task_ids = [r['id'] for r in rows]
    placeholders = ','.join('?' * len(task_ids))

    # 2) Batch-fetch subtask counts (one query for all tasks).
    cursor.execute(f'''
        SELECT task_id,
               COUNT(*) AS subtask_total,
               SUM(CASE WHEN done = 1 THEN 1 ELSE 0 END) AS subtask_done
        FROM task_subtasks
        WHERE task_id IN ({placeholders})
        GROUP BY task_id
    ''', task_ids)
    subtask_map = {}
    for r in cursor.fetchall():
        subtask_map[r['task_id']] = {'subtask_total': r['subtask_total'],
                                      'subtask_done': r['subtask_done']}

    conn.close()

    # 4) Merge results in Python — response shape identical to the old
    #    correlated-subquery version (same keys, same 0 defaults).
    result = []
    for row in rows:
        d = row_to_dict(row)
        sc = subtask_map.get(d['id'], {})
        d['subtask_total'] = sc.get('subtask_total', 0)
        d['subtask_done'] = sc.get('subtask_done', 0)
        result.append(d)

    return jsonify(result)


@tasks_bp.route('/api/global-tasks', methods=['POST'])
@login_required
def create_global_task():
    data = get_json_or_400()
    conn = get_db()
    cursor = conn.cursor()

    now = datetime.now().isoformat()
    task_id = data.get('id') or generate_uuid()

    cursor.execute('''
        INSERT INTO global_tasks (id, titlu, descriere, status, categorie,
                                  data_scadenta, data_finalizare, created_at, updated_at, recurenta,
                                  ordine_agenda)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        task_id,
        data.get('titlu', ''),
        data.get('descriere', ''),
        data.get('status', 'to_do'),
        data.get('categorie', 'General'),
        data.get('data_scadenta', ''),
        data.get('data_finalizare', ''),
        now,
        now,
        data.get('recurenta', ''),
        data.get('ordine_agenda', 0)
    ))

    conn.commit()
    conn.close()

    return jsonify({'id': task_id}), 201


@tasks_bp.route('/api/global-tasks/<task_id>', methods=['GET'])
@login_required
def get_global_task(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM global_tasks WHERE id = ?', (task_id,))
    row = cursor.fetchone()
    conn.close()

    if row is None:
        return jsonify({'error': 'Task not found'}), 404

    return jsonify(row_to_dict(row))


@tasks_bp.route('/api/global-tasks/<task_id>', methods=['PUT'])
@login_required
def update_global_task(task_id):
    data = get_json_or_400()
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM global_tasks WHERE id = ?', (task_id,))
    existing = cursor.fetchone()
    if existing is None:
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    old_status = existing['status']
    now = datetime.now().isoformat()

    # Tranzitie atomica la 'done' — vezi update_task (guard anti-dublu-spawn).
    transitioned_to_done = False
    if data.get('status') == 'done' and old_status != 'done':
        cursor.execute("UPDATE global_tasks SET status = 'done' WHERE id = ? AND status != 'done'", (task_id,))
        transitioned_to_done = cursor.rowcount == 1

    cursor.execute('''
        UPDATE global_tasks SET
            titlu = COALESCE(?, titlu),
            descriere = COALESCE(?, descriere),
            status = COALESCE(?, status),
            categorie = COALESCE(?, categorie),
            data_scadenta = COALESCE(?, data_scadenta),
            data_finalizare = COALESCE(?, data_finalizare),
            recurenta = COALESCE(?, recurenta),
            ordine_agenda = COALESCE(?, ordine_agenda),
            updated_at = ?
        WHERE id = ?
    ''', (
        data.get('titlu'),
        data.get('descriere'),
        data.get('status'),
        data.get('categorie'),
        data.get('data_scadenta'),
        data.get('data_finalizare'),
        data.get('recurenta'),
        data.get('ordine_agenda'),
        now,
        task_id
    ))

    # A recurring daily task just completed -> spawn the next occurrence.
    spawned_id = None
    next_scad = None
    if transitioned_to_done and (existing['recurenta'] or '').strip():
        recurenta = existing['recurenta'].strip()
        spawned_id = _spawn_recurring_global_task(cursor, existing, recurenta)
        next_scad = _next_recurrence_date(existing['data_scadenta'] or '', recurenta)

    conn.commit()
    conn.close()

    resp = {'message': 'Task updated'}
    if spawned_id:
        resp['recurring_spawned'] = spawned_id
        resp['recurring_next'] = next_scad
    return jsonify(resp)


@tasks_bp.route('/api/global-tasks/<task_id>', methods=['DELETE'])
@login_required
def delete_global_task(task_id):
    conn = get_db()
    try:
        cursor = conn.cursor()
        # Clean up orphan subtasks: task_subtasks has no FK so DELETE on global_tasks
        # doesn't cascade. Sessions cascade via FK ON DELETE CASCADE (v11).
        cursor.execute('DELETE FROM task_subtasks WHERE task_id = ?', (task_id,))
        cursor.execute('DELETE FROM global_tasks WHERE id = ?', (task_id,))
        deleted = cursor.rowcount
        conn.commit()
        if deleted == 0:
            return jsonify({'error': 'Task not found'}), 404
        return jsonify({'message': 'Task deleted'})
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Agenda — "Astazi", boardul cu care Ion isi incepe ziua
#
# UN TASK ARE O SINGURA DATA (v33). Ion: „mutarea este practic un deadline (…)
# trebuie sa fie adaugat ca deadline pur, sa nu mai dublam atat notiunile."
# Deci boardul de azi = ce e scadent azi SAU restant. A pune un task pe azi
# inseamna a-i da termenul de azi; a-l muta pe alta zi ii muta termenul.
# Boardul uneste taskurile globale (tip='global') si cele de proiect ('proiect').
# ---------------------------------------------------------------------------

_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')


def _resolve_today():
    """Local 'today' from the client (?today=YYYY-MM-DD); SQLite date('now') is
    UTC and can be off near midnight. Falls back to the server's local date."""
    t = (request.args.get('today') or '').strip()
    if _DATE_RE.match(t):
        return t
    return datetime.now().strftime('%Y-%m-%d')


def _agenda_item(d, tip, today):
    """Normalize a task row (dict) into the flat shape the board consumes."""
    scad = (d.get('data_scadenta') or '').strip()[:10]
    return {
        'tip': tip,
        'id': d['id'],
        'titlu': d.get('titlu') or '',
        'status': d.get('status') or 'to_do',
        'data_scadenta': d.get('data_scadenta') or '',
        'data_finalizare': d.get('data_finalizare') or '',
        'ordine_agenda': d.get('ordine_agenda') or 0,
        'recurenta': d.get('recurenta') or '',
        'categorie': (d.get('categorie') or '') if tip == 'global' else '',
        'proiect_id': d.get('proiect_id') if tip == 'proiect' else None,
        'proiect_nume': d.get('proiect_nume') if tip == 'proiect' else None,
        'is_scadent_azi': bool(scad) and scad == today,
        'is_restant': bool(scad) and scad < today,
    }


# Ce intra pe boardul de azi: taskurile deschise scadente AZI sau RESTANTE. Cu o
# singura data, regula are o singura linie — inainte avea trei ramuri si o exceptie,
# fiindca planul si termenul se puteau contrazice. Recurentele viitoare raman
# ascunse, ca o ocurenta abia generata sa nu apara inainte de vreme.
_AGENDA_WHERE = '''
        {alias}.status != 'done'
        AND {alias}.data_scadenta IS NOT NULL AND TRIM({alias}.data_scadenta) <> ''
        AND date({alias}.data_scadenta) <= date(:today)
        AND NOT (
            {alias}.recurenta IS NOT NULL AND TRIM({alias}.recurenta) <> ''
            AND {alias}.data_scadenta IS NOT NULL AND date({alias}.data_scadenta) > date(:today)
        )
'''


@tasks_bp.route('/api/agenda/today', methods=['GET'])
@login_required
def get_agenda_today():
    today = _resolve_today()
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        'SELECT g.* FROM global_tasks g WHERE ' + _AGENDA_WHERE.format(alias='g'),
        {'today': today})
    items = [_agenda_item(row_to_dict(r), 'global', today) for r in cursor.fetchall()]

    cursor.execute(
        '''SELECT t.*, p.nume AS proiect_nume
           FROM tasks t JOIN proiecte p ON t.proiect_id = p.id
           WHERE p.status != 'anulat' AND ''' + _AGENDA_WHERE.format(alias='t'),
        {'today': today})
    items += [_agenda_item(row_to_dict(r), 'proiect', today) for r in cursor.fetchall()]

    conn.close()

    # Restante first, then by board order (unordered = 0 sinks to the bottom), then title.
    items.sort(key=lambda d: (
        0 if d['is_restant'] else 1,
        d['ordine_agenda'] if d['ordine_agenda'] else 1_000_000,
        (d['titlu'] or '').lower(),
    ))
    return jsonify({'today': today, 'items': items})


@tasks_bp.route('/api/agenda/candidates', methods=['GET'])
@login_required
def get_agenda_candidates():
    """Not-done tasks (global + project) that can be added to today, excluding
    those already planned for today. Optional ?q= title search."""
    today = _resolve_today()
    q = (request.args.get('q') or '').strip()
    conn = get_db()
    cursor = conn.cursor()

    # Candidatii = ce NU e deja pe board. Cu o singura data asta inseamna: fara
    # termen, sau cu termen in viitor. Restantele sunt deja pe board.
    gq = '''SELECT g.* FROM global_tasks g
            WHERE g.status != 'done'
              AND (g.data_scadenta IS NULL OR TRIM(g.data_scadenta) = ''
                   OR date(g.data_scadenta) > date(:today))
              AND NOT (
                g.recurenta IS NOT NULL AND TRIM(g.recurenta) <> ''
                AND g.data_scadenta IS NOT NULL AND date(g.data_scadenta) > date(:today)
              )'''
    gp = {'today': today}
    if q:
        gq += ' AND g.titlu LIKE :q'
        gp['q'] = f'%{q}%'
    gq += ' ORDER BY g.titlu COLLATE NOCASE LIMIT 100'
    cursor.execute(gq, gp)
    items = [_agenda_item(row_to_dict(r), 'global', today) for r in cursor.fetchall()]

    tq = '''SELECT t.*, p.nume AS proiect_nume
            FROM tasks t JOIN proiecte p ON t.proiect_id = p.id
            WHERE t.status != 'done' AND p.status != 'anulat'
              AND (t.data_scadenta IS NULL OR TRIM(t.data_scadenta) = ''
                   OR date(t.data_scadenta) > date(:today))
              AND NOT (
                t.recurenta IS NOT NULL AND TRIM(t.recurenta) <> ''
                AND t.data_scadenta IS NOT NULL AND date(t.data_scadenta) > date(:today)
              )'''
    tp = {'today': today}
    if q:
        tq += ' AND t.titlu LIKE :q'
        tp['q'] = f'%{q}%'
    tq += ' ORDER BY t.titlu COLLATE NOCASE LIMIT 100'
    cursor.execute(tq, tp)
    items += [_agenda_item(row_to_dict(r), 'proiect', today) for r in cursor.fetchall()]

    conn.close()
    return jsonify({'today': today, 'items': items})


@tasks_bp.route('/api/agenda/reorder', methods=['POST'])
@login_required
def reorder_agenda():
    """Persist the board order for a mixed list of global + project tasks.
    Body: {"order": [{"tip": "global"|"proiect", "id": "..."}, ...]}.
    Writes ordine_agenda = position+1 (1-based, so 0 stays 'unordered')."""
    data = get_json_or_400()
    order = data.get('order') or []
    conn = get_db()
    cursor = conn.cursor()
    try:
        for i, item in enumerate(order):
            tip = (item or {}).get('tip')
            tid = (item or {}).get('id')
            if not tid:
                continue
            if tip == 'global':
                cursor.execute('UPDATE global_tasks SET ordine_agenda = ? WHERE id = ?', (i + 1, tid))
            elif tip == 'proiect':
                cursor.execute('UPDATE tasks SET ordine_agenda = ? WHERE id = ?', (i + 1, tid))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.exception("reorder_agenda a esuat: %s", e)
        return jsonify({'error': 'Reordonarea a esuat.'}), 500
    conn.close()
    return jsonify({'message': 'ok', 'count': len(order)})


def _span_intersects(d1, d2, start_s, end_s):
    """True if the day-span [min(d1,d2), max(d1,d2)] intersects the window
    [start_s, end_s) (all 'YYYY-MM-DD'; end exclusive). Empty dates ignored;
    a single present date is treated as a 1-day span."""
    ds = [x for x in ((d1 or '')[:10], (d2 or '')[:10]) if x]
    if not ds:
        return False
    lo, hi = min(ds), max(ds)
    return lo < end_s and hi >= start_s


@tasks_bp.route('/api/plan', methods=['GET'])
@login_required
def get_plan():
    """Operational 14-day planner ("Planificator"): project lanes (each with its
    overall interval data_incepere -> ultima perioada planificata) containing their
    tasks, plus a "Globale" lane. A task appears where its span
    data_scadenta cade in fereastra. Din v33 taskul are o singura data, deci
    fiecare task e un semn de o zi, nu un interval — intervalul plan->termen era
    o distinctie pe care nimeni n-o folosea (3 randuri din 37, toate de o zi).

    Deadline-ul de proiect a plecat in v30 — nu se lua nimeni dupa el. Capatul
    din dreapta al benzii e acum ultima zi pe care chiar ai planificat-o."""
    today = _resolve_today()
    start = (request.args.get('start') or '').strip()
    if not _DATE_RE.match(start):
        start = today
    try:
        days = int(request.args.get('days') or 14)
    except (TypeError, ValueError):
        days = 14
    days = max(1, min(days, 370))  # up to ~1 an; UI merge pana la 6 luni (180)
    show_done = (request.args.get('done') or '').lower() in ('1', 'true', 'yes')
    start_d = datetime.strptime(start, '%Y-%m-%d').date()
    start_s = start_d.isoformat()
    end_s = (start_d + timedelta(days=days)).isoformat()  # exclusive

    # A task belongs in the window if its plan->due span intersects it; a finished
    # task with no plan/due can still qualify via its completion date.
    def _in_window(item):
        if _span_intersects(item['data_scadenta'], item['data_scadenta'], start_s, end_s):
            return True
        fin = (item.get('data_finalizare') or '')[:10]
        return bool(fin) and start_s <= fin < end_s

    conn = get_db()
    cursor = conn.cursor()

    status_clause = '' if show_done else "{alias}.status != 'done' AND "

    # Candidate project lanes (skip cancelled / finished).
    cursor.execute(
        "SELECT id, nume, tip, status FROM proiecte "
        "WHERE status NOT IN ('anulat', 'finalizat')")
    projects = [row_to_dict(r) for r in cursor.fetchall()]

    # Project tasks (exclude future recurrences, same idiom as the agenda).
    cursor.execute(
        '''SELECT t.* FROM tasks t JOIN proiecte p ON t.proiect_id = p.id
           WHERE ''' + status_clause.format(alias='t') + '''p.status NOT IN ('anulat', 'finalizat')
             AND NOT (
               t.recurenta IS NOT NULL AND TRIM(t.recurenta) <> ''
               AND t.data_scadenta IS NOT NULL AND date(t.data_scadenta) > date(:today)
             )''',
        {'today': today})
    tasks_by_project = {}
    for r in cursor.fetchall():
        item = _agenda_item(row_to_dict(r), 'proiect', today)
        if _in_window(item):
            tasks_by_project.setdefault(item['proiect_id'], []).append(item)

    def _task_sort_key(t):
        cand = [x for x in ((t['data_scadenta'] or '')[:10],) if x]
        return (min(cand) if cand else '9999-99-99', (t['titlu'] or '').lower())

    # Implementation periods (Site / Sediu EGB) per project, for lane bands.
    impl_by_project = {}
    proj_ids = [p['id'] for p in projects]
    if proj_ids:
        ph = ','.join('?' * len(proj_ids))
        cursor.execute(f'SELECT * FROM implementari WHERE proiect_id IN ({ph})', proj_ids)
        for r in [row_to_dict(x) for x in cursor.fetchall()]:
            # `faza` merge la client, nu doar `locatie`: Planificatorul are nevoie
            # de ea ca sa nu numere o zi de PREGATIRE ca etapa de implementare —
            # o astfel de zi nu rupe golul de pregatire, se deseneaza peste el.
            impl_by_project.setdefault(r['proiect_id'], []).append({
                'id': r['id'], 'data_start': (r.get('data_start') or '')[:10],
                'data_sfarsit': (r.get('data_sfarsit') or '')[:10],
                'locatie': r.get('locatie') or 'site', 'eticheta': r.get('eticheta') or '',
                'faza': r.get('faza') or 'implementare'})

    lanes = []
    for proj in projects:
        pid = proj['id']
        ptasks = sorted(tasks_by_project.get(pid, []), key=_task_sort_key)
        # Banda proiectului se intinde intre PERIOADE, nu de la un camp scris cu
        # mana: `data_incepere` a plecat in v36 (5 randuri din 18, si dubla prima
        # perioada). Capatul de sus e prima zi planificata, cel de jos ultima —
        # niciunul nu e un deadline impus (vezi v30).
        toate_impl = impl_by_project.get(pid, [])
        inc = min((im.get('data_start') or '' for im in toate_impl if im.get('data_start')),
                  default='')
        ultima = max((im.get('data_sfarsit') or im.get('data_start') or ''
                      for im in toate_impl), default='')
        band = _span_intersects(inc, ultima, start_s, end_s)
        pimpl = [im for im in toate_impl
                 if _span_intersects(im['data_start'], im['data_sfarsit'], start_s, end_s)]
        if not ptasks and not band and not pimpl:
            continue
        lanes.append({
            'tip': 'proiect',
            'id': pid,
            'nume': proj.get('nume') or '',
            'tip_proiect': proj.get('tip') or '',
            'status': proj.get('status') or '',
            'prima_zi': inc,
            'ultima_zi': ultima,
            'tasks': ptasks,
            'implementari': pimpl,
        })

    # Global tasks lane.
    cursor.execute(
        '''SELECT g.* FROM global_tasks g
           WHERE ''' + status_clause.format(alias='g') + '''NOT (
               g.recurenta IS NOT NULL AND TRIM(g.recurenta) <> ''
               AND g.data_scadenta IS NOT NULL AND date(g.data_scadenta) > date(:today)
             )''',
        {'today': today})
    gtasks = []
    for r in cursor.fetchall():
        item = _agenda_item(row_to_dict(r), 'global', today)
        if _in_window(item):
            gtasks.append(item)
    if gtasks:
        lanes.append({
            'tip': 'global', 'id': '__global__', 'nume': 'Globale',
            'tip_proiect': '', 'status': '', 'prima_zi': '', 'ultima_zi': '',
            'tasks': sorted(gtasks, key=_task_sort_key), 'implementari': [],
        })

    # Backlog: open tasks with NO plan and NO deadline — nowhere to place on the
    # timeline, so they surface in a side rail from which they can be scheduled.
    def _backlog_item(d, tip):
        return {
            'tip': tip, 'id': d['id'], 'titlu': d.get('titlu') or '',
                'categorie': (d.get('categorie') or '') if tip == 'global' else '',
            'proiect_id': d.get('proiect_id') if tip == 'proiect' else None,
            'proiect_nume': d.get('proiect_nume') if tip == 'proiect' else None,
        }

    undated = '''({a}.data_scadenta IS NULL OR TRIM({a}.data_scadenta) = '')'''
    cursor.execute(
        '''SELECT t.*, p.nume AS proiect_nume FROM tasks t JOIN proiecte p ON t.proiect_id = p.id
           WHERE t.status != 'done' AND p.status NOT IN ('anulat', 'finalizat')
             AND ''' + undated.format(a='t') + '''
           ORDER BY t.titlu COLLATE NOCASE LIMIT 300''')
    backlog = [_backlog_item(row_to_dict(r), 'proiect') for r in cursor.fetchall()]
    cursor.execute(
        '''SELECT g.* FROM global_tasks g WHERE g.status != 'done'
             AND ''' + undated.format(a='g') + '''
           ORDER BY g.titlu COLLATE NOCASE LIMIT 300''')
    backlog += [_backlog_item(row_to_dict(r), 'global') for r in cursor.fetchall()]

    conn.close()

    # Lane order: projects by earliest in-window activity, Globale last.
    def _lane_key(l):
        dates = []
        for t in l['tasks']:
            for k in ('data_scadenta',):
                v = (t.get(k) or '')[:10]
                if v:
                    dates.append(v)
        if l['prima_zi']:
            dates.append(l['prima_zi'])
        return (1 if l['tip'] == 'global' else 0, min(dates) if dates else '9999-99-99')

    lanes.sort(key=_lane_key)
    return jsonify({'start': start_s, 'days': days, 'today': today, 'lanes': lanes, 'backlog': backlog})
