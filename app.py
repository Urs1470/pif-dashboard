import os
import uuid
import json
import functools
import time
import logging
import hashlib
import hmac
import subprocess
import threading
import html
import re

from datetime import datetime, timedelta
from logging.handlers import RotatingFileHandler
from io import BytesIO
from flask import Flask, request, jsonify, send_file, render_template, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import sqlite3
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from database import get_db, init_db, row_to_dict, close_db
from blueprints.budget import budget_bp
from scripts.parse_params import parse_for_producator

VALID_TABLES = {'proiecte', 'tasks', 'task_subtasks', 'checklist_pif', 'jurnal', 'timer_sessions',
                'atasamente', 'echipamente', 'clienti', 'global_tasks', 'global_task_sessions',
                'checklist_categorii', 'fault_codes', 'project_templates', 'budget_state',
                'budget_audit', 'parametri_master'}

def safe_table(table_name):
    if table_name not in VALID_TABLES:
        raise ValueError(f"Invalid table name: {table_name}")
    return table_name

app = Flask(__name__)
app.register_blueprint(budget_bp)

# Persistent secret key — supravietuieste restart-urilor
SECRET_KEY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.secret_key')

def get_or_create_secret_key():
    env_key = os.environ.get('SECRET_KEY')
    if env_key:
        return env_key.encode() if isinstance(env_key, str) else env_key
    if os.path.exists(SECRET_KEY_FILE):
        with open(SECRET_KEY_FILE, 'rb') as f:
            return f.read()
    key = os.urandom(32)
    with open(SECRET_KEY_FILE, 'wb') as f:
        f.write(key)
    return key

app.secret_key = get_or_create_secret_key()
# CORS intentionally NOT enabled: the dashboard is a same-origin app, and a
# wildcard CORS policy would let any website call the authenticated API.
app.teardown_appcontext(close_db)

# ============ VERSION HASH ============

def file_hash(filepath):
    """Returnează primele 8 caractere din SHA256 al fișierului."""
    try:
        with open(filepath, 'rb') as f:
            return hashlib.sha256(f.read()).hexdigest()[:8]
    except FileNotFoundError:
        return 'dev'

@app.context_processor
def inject_version():
    return {
        'js_version': file_hash('static/mobile.js'),
        'sw_version': file_hash('static/service-worker.js'),
        'app_version': file_hash('static/app.js'),
        'core_version': file_hash('static/core.js')
    }

# ============ SESSION CONFIG ============

app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
# Allow large uploads for /api/admin/db-upload (DB ~40 MB now, may grow)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200 MB

# Session cookie hardening. SameSite=Lax is what stops a malicious site from
# making authenticated state-changing calls to the API (CSRF): cross-site
# POST/PUT/DELETE no longer carry the session cookie. Secure = HTTPS-only;
# HttpOnly = not readable from JavaScript.
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

@app.before_request
def make_session_permanent():
    session.permanent = True

# ============ PHASE 2a: STRUCTURED LOGGING ============
LOGS_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
os.makedirs(LOGS_FOLDER, exist_ok=True)

def setup_logging():
    logger = logging.getLogger('pif_dashboard')
    logger.setLevel(logging.INFO)
    
    # Avoid duplicate handlers
    if logger.handlers:
        return logger
    
    # Rotating file handler - 10 files, 1MB each
    handler = RotatingFileHandler(
        os.path.join(LOGS_FOLDER, 'app.log'),
        maxBytes=1024 * 1024,  # 1MB
        backupCount=10
    )
    handler.setLevel(logging.INFO)
    
    # Format: [YYYY-MM-DD HH:MM:SS] LEVEL - message
    formatter = logging.Formatter('[%(asctime)s] %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    handler.setFormatter(formatter)
    
    logger.addHandler(handler)
    return logger

logger = setup_logging()

# ============ PHASE 2a: API RATE LIMITING ============
rate_limit_store = {}  # {ip: [(timestamp, count), ...]}
RATE_LIMIT = 60  # requests per minute
RATE_WINDOW = 60  # seconds

def check_rate_limit():
    """Simple token-bucket rate limiter. Returns True if allowed, False if exceeded."""
    client_ip = request.remote_addr or '127.0.0.1'
    now = time.time()
    
    # Clean old entries
    if client_ip in rate_limit_store:
        rate_limit_store[client_ip] = [
            (ts, count) for ts, count in rate_limit_store[client_ip]
            if now - ts < RATE_WINDOW
        ]
    
    # Count requests in current window
    request_count = 0
    if client_ip in rate_limit_store:
        request_count = sum(count for ts, count in rate_limit_store[client_ip])
    
    if request_count >= RATE_LIMIT:
        return False
    
    # Record this request
    if client_ip not in rate_limit_store:
        rate_limit_store[client_ip] = []
    rate_limit_store[client_ip].append((now, 1))
    
    return True

# Phase 2c: One-time startup initialization
_startup_initialized = False
_startup_lock = threading.Lock()

@app.before_request
def before_request_func():
    global _startup_initialized

    # Run one-time initialization (migrations, templates)
    if not _startup_initialized:
        with _startup_lock:
            # Double-checked: only the first of the parallel requests an app
            # load fires actually runs the migrations.
            if not _startup_initialized:
                with app.app_context():
                    init_db()
                    try:
                        init_default_templates()
                    except:
                        pass  # Templates may already exist
                if not os.environ.get('PIF_DASHBOARD_PIN'):
                    logger.warning("PIF_DASHBOARD_PIN nu este setat — aplicatia foloseste PIN-ul implicit. Seteaza variabila de mediu pe server.")
                _startup_initialized = True
                logger.info("PIF Dashboard initialized")

    # Rate-limit all /api/* routes, plus the login endpoints (the latter
    # throttles PIN brute-force attempts).
    _rl_api = request.path.startswith('/api/') and request.path not in ('/api/login', '/api/healthz')
    _rl_login = request.path in ('/login', '/login-hash') and request.method == 'POST'
    if _rl_api or _rl_login:
        if not check_rate_limit():
            logger.warning(f"Rate limit exceeded for IP: {request.remote_addr} on {request.path}")
            return jsonify({'error': 'Rate limit exceeded. Maximum 60 requests per minute.', 'retry_after': RATE_WINDOW}), 429
    
    # Log all requests
    if request.path.startswith('/api/'):
        logger.info(f"{request.method} {request.path} - IP: {request.remote_addr}")

@app.after_request
def after_request_func(response):
    # Log API responses
    if request.path.startswith('/api/'):
        logger.info(f"{request.method} {request.path} - Status: {response.status_code}")
    return response

# ============ END PHASE 2a SETUP ============

# PIN configuration
def get_hashed_pin():
    """Get hashed PIN from environment variable"""
    pin = os.environ.get('PIF_DASHBOARD_PIN')
    if not pin:
        import secrets
        pin = secrets.token_hex(4)
        logger.critical(f"PIF_DASHBOARD_PIN nesetat — PIN generat automat: {pin}. Seteaza PIF_DASHBOARD_PIN in environment!")
    # Generate hash on first call if not cached
    if not hasattr(get_hashed_pin, '_hash'):
        get_hashed_pin._hash = generate_password_hash(pin)
    return get_hashed_pin._hash

# Login required decorator
def login_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'authenticated' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

# ============ HEALTHCHECK ============

@app.route('/api/healthz')
def healthz():
    """Ultra-light healthcheck pentru online/offline detection pe mobile.
    Răspunde instant (sub 1ms) — fără DB, fără auth."""
    return jsonify({'status': 'ok', 'timestamp': int(time.time())})

# ============ END HEALTHCHECK ============

# Serve the login page
@app.route('/login')
def login_page():
    if session.get('authenticated'):
        return redirect(url_for('index'))
    return render_template('login.html')

# Login endpoint
@app.route('/login', methods=['POST'])
def login():
    data = request.json
    pin = data.get('pin', '')
    
    if check_password_hash(get_hashed_pin(), pin):
        session['authenticated'] = True
        logger.info(f"Login successful for IP: {request.remote_addr}")
        return jsonify({'success': True})
    logger.warning(f"Login failed for IP: {request.remote_addr}")
    return jsonify({'error': 'Invalid PIN'}), 401

# Logout endpoint
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# Auto-login cu PIN hash (Remember Me)
@app.route('/login-hash', methods=['POST'])
def login_hash():
    """Auto-login cu hash SHA-256 al PIN-ului (pentru 'Remember Me' pe mobil)."""
    data = request.json
    pin_hash = data.get('pin_hash', '')
    if not pin_hash:
        return jsonify({'success': False, 'error': 'Missing pin_hash'}), 400
    
    # Get stored PIN (plaintext for hashing)
    pin = os.environ.get('PIF_DASHBOARD_PIN')
    if not pin:
        return jsonify({'success': False, 'error': 'Server PIN not configured'}), 500
    expected_hash = hashlib.sha256(pin.encode()).hexdigest()
    
    if hmac.compare_digest(str(pin_hash), expected_hash):
        session['authenticated'] = True
        return jsonify({'success': True})

    return jsonify({'success': False, 'error': 'Invalid hash'}), 401

# Serve the frontend
@app.route('/')
@login_required
def index():
    return render_template('index.html')

# SPA catch-all routes for bookmark/SEO
@app.route('/parametri')
@app.route('/notite')
@app.route('/administrativ')
@login_required
def spa_catchall():
    return render_template('index.html')

# Redirect old mobile route
@app.route('/m')
def redirect_mobile():
    return redirect(url_for('index'))

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def generate_uuid():
    return str(uuid.uuid4())

# ============ PROJECTS ============

@app.route('/api/proiecte', methods=['GET'])
@login_required
def get_proiecte():
    conn = get_db()
    cursor = conn.cursor()
    
    status = request.args.get('status')
    tip = request.args.get('tip')
    producator = request.args.get('producator')
    
    # Pagination parameters
    limit = request.args.get('limit', 100, type=int)
    offset = request.args.get('offset', 0, type=int)
    
    query = 'SELECT * FROM proiecte WHERE 1=1'
    params = []
    
    if status:
        query += ' AND status = ?'
        params.append(status)
    if tip:
        query += ' AND tip = ?'
        params.append(tip)
    if producator:
        query += ' AND producator = ?'
        params.append(producator)
    
    query += ' ORDER BY created_at DESC'
    
    # Apply pagination
    query += ' LIMIT ? OFFSET ?'
    params.extend([limit, offset])
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/proiecte', methods=['POST'])
@login_required
def create_proiect():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    project_id = data.get('id') or generate_uuid()
    
    cursor.execute('''
        INSERT INTO proiecte (
            id, tip, nume, client, locatie, echipament_principal, producator,
            cod_proiect, pm, folder_server, data_incepere, deadline, data_crearii,
            status, observatii, nr_comanda, nr_contract, service_before, service_after,
            confirmat_client, client_nume_confirmare, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        project_id,
        data.get('tip', 'PIF'),
        data.get('nume', ''),
        data.get('client', ''),
        data.get('locatie', ''),
        data.get('echipament_principal', ''),
        data.get('producator', 'Altul'),
        data.get('cod_proiect', ''),
        data.get('pm', ''),
        data.get('folder_server', ''),
        data.get('data_incepere', ''),
        data.get('deadline', ''),
        data.get('data_crearii', now[:10]),
        data.get('status', 'in_lucru'),
        data.get('observatii', ''),
        data.get('nr_comanda', ''),
        data.get('nr_contract', ''),
        data.get('service_before', ''),
        data.get('service_after', ''),
        data.get('confirmat_client', 0),
        data.get('client_nume_confirmare', ''),
        now,
        now
    ))
    
    conn.commit()
    conn.close()
    
    logger.info(f"Project created: {project_id} - {data.get('nume', '')}")
    return jsonify({'id': project_id, 'message': 'Project created'}), 201

@app.route('/api/proiecte/<project_id>', methods=['GET'])
@login_required
def get_proiect(project_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (project_id,))
    row = cursor.fetchone()
    conn.close()
    
    if row is None:
        return jsonify({'error': 'Project not found'}), 404
    
    return jsonify(row_to_dict(row))

@app.route('/api/proiecte/<project_id>', methods=['PUT'])
@login_required
def update_proiect(project_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    
    # Get current project status
    cursor.execute('SELECT status, nume FROM proiecte WHERE id = ?', (project_id,))
    current = cursor.fetchone()
    old_status = current['status'] if current else None
    project_name = current['nume'] if current else ''
    
    cursor.execute('''
        UPDATE proiecte SET
            tip = COALESCE(?, tip),
            nume = COALESCE(?, nume),
            client = COALESCE(?, client),
            locatie = COALESCE(?, locatie),
            echipament_principal = COALESCE(?, echipament_principal),
            producator = COALESCE(?, producator),
            cod_proiect = COALESCE(?, cod_proiect),
            pm = COALESCE(?, pm),
            folder_server = COALESCE(?, folder_server),
            data_incepere = COALESCE(?, data_incepere),
            deadline = COALESCE(?, deadline),
            status = COALESCE(?, status),
            observatii = COALESCE(?, observatii),
            nr_comanda = COALESCE(?, nr_comanda),
            nr_contract = COALESCE(?, nr_contract),
            service_before = COALESCE(?, service_before),
            service_after = COALESCE(?, service_after),
            confirmat_client = COALESCE(?, confirmat_client),
            client_nume_confirmare = COALESCE(?, client_nume_confirmare),
            updated_at = ?
        WHERE id = ?
    ''', (
        data.get('tip'),
        data.get('nume'),
        data.get('client'),
        data.get('locatie'),
        data.get('echipament_principal'),
        data.get('producator'),
        data.get('cod_proiect'),
        data.get('pm'),
        data.get('folder_server'),
        data.get('data_incepere'),
        data.get('deadline'),
        data.get('status'),
        data.get('observatii'),
        data.get('nr_comanda'),
        data.get('nr_contract'),
        data.get('service_before'),
        data.get('service_after'),
        data.get('confirmat_client'),
        data.get('client_nume_confirmare'),
        now,
        project_id
    ))
    
    conn.commit()

    conn.close()
    
    logger.info(f"Project updated: {project_id}")
    return jsonify({'message': 'Project updated'})

@app.route('/api/proiecte/<project_id>', methods=['DELETE'])
@login_required
def delete_proiect(project_id):
    conn = get_db()
    cursor = conn.cursor()
    try:
        # Subtasks are keyed by task_id (not proiect_id) — remove them first.
        cursor.execute(
            'DELETE FROM task_subtasks WHERE task_id IN (SELECT id FROM tasks WHERE proiect_id = ?)',
            (project_id,)
        )
        tables = ['tasks', 'checklist_pif', 'checklist_categorii', 'jurnal',
                  'timer_sessions', 'atasamente', 'echipamente']
        for table in tables:
            cursor.execute(f'DELETE FROM {safe_table(table)} WHERE proiect_id = ?', (project_id,))
        cursor.execute('DELETE FROM proiecte WHERE id = ?', (project_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error(f"Error deleting project {project_id}: {e}")
        return jsonify({'error': 'Stergerea proiectului a esuat.'}), 500
    conn.close()
    # Remove the project's uploaded files from disk (orphans otherwise).
    try:
        import shutil
        shutil.rmtree(os.path.join(UPLOAD_FOLDER, project_id), ignore_errors=True)
    except Exception:
        pass
    logger.info(f"Project deleted: {project_id}")
    return jsonify({'message': 'Project deleted'})

# ============ BATCH OPERATIONS ============

@app.route('/api/proiecte/batch', methods=['POST'])
@login_required
def batch_proiecte():
    """Batch update or delete multiple projects"""
    data = request.json
    action = data.get('action')  # 'update_status' or 'delete'
    project_ids = data.get('project_ids', [])
    
    if not project_ids:
        return jsonify({'error': 'No projects selected'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        if action == 'update_status':
            new_status = data.get('status')
            if not new_status:
                return jsonify({'error': 'Status required for update'}), 400
            
            now = datetime.now().isoformat()
            for pid in project_ids:
                cursor.execute('UPDATE proiecte SET status = ?, updated_at = ? WHERE id = ?', (new_status, now, pid))
            
            conn.commit()
            logger.info(f"Batch updated {len(project_ids)} projects to status: {new_status}")
            return jsonify({'message': f'{len(project_ids)} projects updated'})
        
        elif action == 'delete':
            import shutil
            for pid in project_ids:
                # Delete related data first. Subtasks are keyed by task_id.
                cursor.execute(
                    'DELETE FROM task_subtasks WHERE task_id IN (SELECT id FROM tasks WHERE proiect_id = ?)',
                    (pid,)
                )
                tables = ['tasks', 'checklist_pif', 'checklist_categorii', 'jurnal',
                          'timer_sessions', 'atasamente', 'echipamente']
                for table in tables:
                    cursor.execute(f'DELETE FROM {safe_table(table)} WHERE proiect_id = ?', (pid,))
                cursor.execute('DELETE FROM proiecte WHERE id = ?', (pid,))
            
            conn.commit()
            for pid in project_ids:
                shutil.rmtree(os.path.join(UPLOAD_FOLDER, pid), ignore_errors=True)
            logger.info(f"Batch deleted {len(project_ids)} projects")
            return jsonify({'message': f'{len(project_ids)} projects deleted'})
        
        else:
            return jsonify({'error': 'Invalid action'}), 400
    
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error(f"Batch operation error: {e}")
        return jsonify({'error': str(e)}), 500

# ============ TASKS ============

@app.route('/api/proiecte/<project_id>/tasks', methods=['GET'])
@login_required
def get_tasks(project_id):
    conn = get_db()
    cursor = conn.cursor()
    # Sort by ordine for drag-and-drop. Subtask counts + tracked time joined in
    # so the todo row can show progress badges without N extra requests.
    cursor.execute('''
        SELECT t.*,
            (SELECT COUNT(*) FROM task_subtasks s WHERE s.task_id = t.id) AS subtask_total,
            (SELECT COUNT(*) FROM task_subtasks s WHERE s.task_id = t.id AND s.done = 1) AS subtask_done,
            (SELECT COALESCE(SUM(durata_secunde), 0) FROM timer_sessions ts
             WHERE ts.task_id = t.id AND ts.durata_secunde IS NOT NULL) AS timp_secunde
        FROM tasks t WHERE t.proiect_id = ?
        ORDER BY t.ordine ASC, t.created_at DESC
    ''', (project_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/proiecte/<project_id>/tasks', methods=['POST'])
@login_required
def create_task(project_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    task_id = data.get('id') or generate_uuid()
    
    # Get max ordine for this project
    cursor.execute('SELECT MAX(ordine) as max_ordine FROM tasks WHERE proiect_id = ?', (project_id,))
    result = cursor.fetchone()
    max_ordine = result['max_ordine'] if result and result['max_ordine'] is not None else 0
    
    cursor.execute('''
        INSERT INTO tasks (id, proiect_id, titlu, status, prioritate, data_scadenta,
                           data_finalizare, ordine, created_at, descriere, recurenta, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        task_id,
        project_id,
        data.get('titlu', ''),
        data.get('status', 'to_do'),
        data.get('prioritate', 'normal'),
        data.get('data_scadenta', ''),
        data.get('data_finalizare', ''),
        max_ordine + 1,
        now,
        data.get('descriere', ''),
        data.get('recurenta', ''),
        now
    ))

    conn.commit()
    conn.close()

    return jsonify({'id': task_id}), 201


def _next_recurrence_date(base_str, recurenta):
    """base_str: 'YYYY-MM-DD' (flatpickr format) or empty. Returns the next
    occurrence date as 'YYYY-MM-DD'. Falls back to today when base is unparsable."""
    try:
        base = datetime.strptime((base_str or '')[:10], '%Y-%m-%d').date()
    except (ValueError, TypeError):
        base = datetime.now().date()
    if recurenta == 'zilnic':
        nxt = base + timedelta(days=1)
    elif recurenta == 'saptamanal':
        nxt = base + timedelta(days=7)
    elif recurenta == 'lunar':
        import calendar
        m = base.month + 1
        y = base.year + (1 if m > 12 else 0)
        if m > 12:
            m -= 12
        d = min(base.day, calendar.monthrange(y, m)[1])
        nxt = datetime(y, m, d).date()
    else:
        return base_str or ''
    return nxt.isoformat()


def _spawn_recurring_task(cursor, existing, recurenta):
    """Create the next occurrence of a recurring task that was just completed.
    Copies title/priority/description/recurrence and fresh (unchecked) subtasks.
    `existing` is the sqlite Row of the completed task. Returns the new id."""
    new_id = generate_uuid()
    now = datetime.now().isoformat()
    next_scad = _next_recurrence_date(existing['data_scadenta'] or '', recurenta)
    cursor.execute('SELECT MAX(ordine) FROM tasks WHERE proiect_id = ?', (existing['proiect_id'],))
    max_ordine = cursor.fetchone()[0] or 0
    cursor.execute('''
        INSERT INTO tasks (id, proiect_id, titlu, status, prioritate, data_scadenta,
                           data_finalizare, ordine, created_at, descriere, recurenta, updated_at)
        VALUES (?, ?, ?, 'to_do', ?, ?, '', ?, ?, ?, ?, ?)
    ''', (new_id, existing['proiect_id'], existing['titlu'], existing['prioritate'],
          next_scad, max_ordine + 1, now, existing['descriere'] or '', recurenta, now))
    # Carry the subtasks over, all unchecked — a recurring checklist repeats clean.
    cursor.execute('SELECT titlu, ordine FROM task_subtasks WHERE task_id = ? ORDER BY ordine', (existing['id'],))
    for srow in cursor.fetchall():
        cursor.execute(
            'INSERT INTO task_subtasks (id, task_id, titlu, done, ordine, created_at) VALUES (?, ?, ?, 0, ?, ?)',
            (generate_uuid(), new_id, srow['titlu'], srow['ordine'], now)
        )
    return new_id


@app.route('/api/tasks/<task_id>', methods=['PUT'])
@login_required
def update_task(task_id):
    data = request.json or {}
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM tasks WHERE id = ?', (task_id,))
    existing = cursor.fetchone()
    if existing is None:
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    old_status = existing['status']

    cursor.execute('''
        UPDATE tasks SET
            titlu = COALESCE(?, titlu),
            status = COALESCE(?, status),
            prioritate = COALESCE(?, prioritate),
            data_scadenta = COALESCE(?, data_scadenta),
            data_finalizare = COALESCE(?, data_finalizare),
            ordine = COALESCE(?, ordine),
            descriere = COALESCE(?, descriere),
            recurenta = COALESCE(?, recurenta),
            updated_at = ?
        WHERE id = ?
    ''', (
        data.get('titlu'),
        data.get('status'),
        data.get('prioritate'),
        data.get('data_scadenta'),
        data.get('data_finalizare'),
        data.get('ordine'),
        data.get('descriere'),
        data.get('recurenta'),
        datetime.now().isoformat(),
        task_id
    ))

    # A recurring task just completed -> spawn the next occurrence.
    spawned_id = None
    if (data.get('status') == 'done' and old_status != 'done'
            and (existing['recurenta'] or '').strip()):
        spawned_id = _spawn_recurring_task(cursor, existing, existing['recurenta'].strip())

    conn.commit()
    conn.close()

    resp = {'message': 'Task updated'}
    if spawned_id:
        resp['recurring_spawned'] = spawned_id
    return jsonify(resp)

@app.route('/api/tasks/<task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM task_subtasks WHERE task_id = ?', (task_id,))
    cursor.execute('DELETE FROM tasks WHERE id = ?', (task_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Task deleted'})

# ============ TASK SUBTASKS (lightweight checklist under a task) ============

@app.route('/api/tasks/<task_id>/subtasks', methods=['GET'])
@login_required
def get_subtasks(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM task_subtasks WHERE task_id = ? ORDER BY ordine ASC, created_at ASC', (task_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route('/api/tasks/<task_id>/subtasks', methods=['POST'])
@login_required
def create_subtask(task_id):
    data = request.json or {}
    titlu = (data.get('titlu') or '').strip()
    if not titlu:
        return jsonify({'error': 'Titlu required'}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COALESCE(MAX(ordine), -1) + 1 FROM task_subtasks WHERE task_id = ?', (task_id,))
    next_ordine = cursor.fetchone()[0]
    sid = generate_uuid()
    cursor.execute(
        'INSERT INTO task_subtasks (id, task_id, titlu, done, ordine, created_at) VALUES (?, ?, ?, 0, ?, ?)',
        (sid, task_id, titlu, next_ordine, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    return jsonify({'id': sid}), 201

@app.route('/api/subtasks/<subtask_id>', methods=['PUT'])
@login_required
def update_subtask(subtask_id):
    data = request.json or {}
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'UPDATE task_subtasks SET titlu = COALESCE(?, titlu), done = COALESCE(?, done) WHERE id = ?',
        (data.get('titlu'), data.get('done'), subtask_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/subtasks/<subtask_id>', methods=['DELETE'])
@login_required
def delete_subtask(subtask_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM task_subtasks WHERE id = ?', (subtask_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ============ GLOBAL TASKS ============

@app.route('/api/global-tasks', methods=['GET'])
@login_required
def get_global_tasks():
    conn = get_db()
    cursor = conn.cursor()

    status = request.args.get('status')
    prioritate = request.args.get('prioritate')
    categorie = request.args.get('categorie')
    arhiva = request.args.get('arhiva')

    query = ('SELECT g.*, '
             '(SELECT COUNT(*) FROM task_subtasks s WHERE s.task_id = g.id) AS subtask_total, '
             '(SELECT COUNT(*) FROM task_subtasks s WHERE s.task_id = g.id AND s.done = 1) AS subtask_done, '
             '(SELECT COALESCE(SUM(durata_secunde), 0) FROM global_task_sessions gs '
             'WHERE gs.global_task_id = g.id) AS timp_secunde '
             'FROM global_tasks g WHERE 1=1')
    params = []

    if arhiva == 'true':
        query += " AND status = 'done'"
    else:
        query += " AND status != 'done'"

    if status and arhiva != 'true':
        query += ' AND status = ?'
        params.append(status)
    if prioritate:
        query += ' AND prioritate = ?'
        params.append(prioritate)
    if categorie:
        query += ' AND categorie = ?'
        params.append(categorie)

    query += ' ORDER BY created_at DESC'

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/global-tasks', methods=['POST'])
@login_required
def create_global_task():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()

    now = datetime.now().isoformat()
    task_id = data.get('id') or generate_uuid()

    cursor.execute('''
        INSERT INTO global_tasks (id, titlu, descriere, prioritate, status, categorie,
                                  data_scadenta, data_finalizare, created_at, updated_at, recurenta)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        task_id,
        data.get('titlu', ''),
        data.get('descriere', ''),
        data.get('prioritate', 'Normal'),
        data.get('status', 'to_do'),
        data.get('categorie', 'General'),
        data.get('data_scadenta', ''),
        data.get('data_finalizare', ''),
        now,
        now,
        data.get('recurenta', '')
    ))

    conn.commit()
    conn.close()

    return jsonify({'id': task_id}), 201

@app.route('/api/global-tasks/<task_id>', methods=['GET'])
@login_required
def get_global_task(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''SELECT g.*,
        (SELECT COALESCE(SUM(durata_secunde), 0) FROM global_task_sessions gs
         WHERE gs.global_task_id = g.id) AS timp_secunde
        FROM global_tasks g WHERE g.id = ?''', (task_id,))
    row = cursor.fetchone()
    conn.close()

    if row is None:
        return jsonify({'error': 'Task not found'}), 404

    return jsonify(row_to_dict(row))

def _spawn_recurring_global_task(cursor, existing, recurenta):
    """Spawn the next occurrence of a completed recurring daily task. Copies
    title/priority/category/description + fresh subtasks. Returns the new id."""
    new_id = generate_uuid()
    now = datetime.now().isoformat()
    next_scad = _next_recurrence_date(existing['data_scadenta'] or '', recurenta)
    cursor.execute('''
        INSERT INTO global_tasks (id, titlu, descriere, prioritate, status, categorie,
                                  data_scadenta, data_finalizare, created_at, updated_at, recurenta)
        VALUES (?, ?, ?, ?, 'to_do', ?, ?, '', ?, ?, ?)
    ''', (new_id, existing['titlu'], existing['descriere'] or '', existing['prioritate'],
          existing['categorie'], next_scad, now, now, recurenta))
    cursor.execute('SELECT titlu, ordine FROM task_subtasks WHERE task_id = ? ORDER BY ordine', (existing['id'],))
    for srow in cursor.fetchall():
        cursor.execute(
            'INSERT INTO task_subtasks (id, task_id, titlu, done, ordine, created_at) VALUES (?, ?, ?, 0, ?, ?)',
            (generate_uuid(), new_id, srow['titlu'], srow['ordine'], now)
        )
    return new_id


@app.route('/api/global-tasks/<task_id>', methods=['PUT', 'POST'])
@login_required
def update_global_task(task_id):
    data = request.json or {}
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM global_tasks WHERE id = ?', (task_id,))
    existing = cursor.fetchone()
    if existing is None:
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    old_status = existing['status']
    now = datetime.now().isoformat()

    cursor.execute('''
        UPDATE global_tasks SET
            titlu = COALESCE(?, titlu),
            descriere = COALESCE(?, descriere),
            prioritate = COALESCE(?, prioritate),
            status = COALESCE(?, status),
            categorie = COALESCE(?, categorie),
            data_scadenta = COALESCE(?, data_scadenta),
            data_finalizare = COALESCE(?, data_finalizare),
            recurenta = COALESCE(?, recurenta),
            updated_at = ?
        WHERE id = ?
    ''', (
        data.get('titlu'),
        data.get('descriere'),
        data.get('prioritate'),
        data.get('status'),
        data.get('categorie'),
        data.get('data_scadenta'),
        data.get('data_finalizare'),
        data.get('recurenta'),
        now,
        task_id
    ))

    # A recurring daily task just completed -> spawn the next occurrence.
    spawned_id = None
    if (data.get('status') == 'done' and old_status != 'done'
            and (existing['recurenta'] or '').strip()):
        spawned_id = _spawn_recurring_global_task(cursor, existing, existing['recurenta'].strip())

    conn.commit()
    conn.close()

    resp = {'message': 'Task updated'}
    if spawned_id:
        resp['recurring_spawned'] = spawned_id
    return jsonify(resp)

@app.route('/api/global-tasks/<task_id>', methods=['DELETE'])
@login_required
def delete_global_task(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM global_tasks WHERE id = ?', (task_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Task deleted'})

# ============ CHECKLIST PIF ============

@app.route('/api/proiecte/<project_id>/checklist', methods=['GET'])
@login_required
def get_checklist(project_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM checklist_pif WHERE proiect_id = ? ORDER BY ordine', (project_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/proiecte/<project_id>/checklist', methods=['POST'])
@login_required
def create_checklist_item(project_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()

    item_id = data.get('id') or generate_uuid()
    cat_id = data.get('categorie_id')
    # Accept '' / 'null' / 0 as "no category" -> NULL
    if cat_id in ('', 'null', 0, '0'):
        cat_id = None

    cursor.execute('''
        INSERT INTO checklist_pif (id, proiect_id, titlu, completed, note, ordine, categorie_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (
        item_id,
        project_id,
        data.get('titlu', ''),
        data.get('completed', 0),
        data.get('note', ''),
        data.get('ordine', 0),
        cat_id
    ))

    conn.commit()
    conn.close()

    return jsonify({'id': item_id}), 201

@app.route('/api/checklist/<item_id>', methods=['PUT'])
@login_required
def update_checklist_item(item_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()

    # categorie_id needs an explicit "no value" sentinel so 'UNSET' means "do not touch"
    # while None / 0 means "clear category".
    cat_present = 'categorie_id' in data
    cat_id = data.get('categorie_id')
    if cat_id in ('', 'null', 0, '0'):
        cat_id = None

    if cat_present:
        cursor.execute('''
            UPDATE checklist_pif SET
                titlu = COALESCE(?, titlu),
                completed = COALESCE(?, completed),
                note = COALESCE(?, note),
                ordine = COALESCE(?, ordine),
                categorie_id = ?
            WHERE id = ?
        ''', (
            data.get('titlu'),
            data.get('completed'),
            data.get('note'),
            data.get('ordine'),
            cat_id,
            item_id
        ))
    else:
        cursor.execute('''
            UPDATE checklist_pif SET
                titlu = COALESCE(?, titlu),
                completed = COALESCE(?, completed),
                note = COALESCE(?, note),
                ordine = COALESCE(?, ordine)
            WHERE id = ?
        ''', (
            data.get('titlu'),
            data.get('completed'),
            data.get('note'),
            data.get('ordine'),
            item_id
        ))

    conn.commit()
    conn.close()

    return jsonify({'message': 'Checklist item updated'})

@app.route('/api/checklist/<item_id>', methods=['DELETE'])
@login_required
def delete_checklist_item(item_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM checklist_pif WHERE id = ?', (item_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Checklist item deleted'})

# ============ CHECKLIST CATEGORII (per-project dynamic) ============

@app.route('/api/proiecte/<project_id>/checklist-categorii', methods=['GET'])
@login_required
def list_checklist_categorii(project_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, proiect_id, nume, ordine, created_at FROM checklist_categorii WHERE proiect_id = ? ORDER BY ordine, id', (project_id,))
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/proiecte/<project_id>/checklist-categorii', methods=['POST'])
@login_required
def create_checklist_categorie(project_id):
    data = request.json or {}
    nume = (data.get('nume') or '').strip()
    if not nume:
        return jsonify({'error': 'Nume required'}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COALESCE(MAX(ordine), -1) + 1 FROM checklist_categorii WHERE proiect_id = ?', (project_id,))
    next_ordine = cursor.fetchone()[0]
    cursor.execute(
        'INSERT INTO checklist_categorii(proiect_id, nume, ordine, created_at) VALUES (?, ?, ?, ?)',
        (project_id, nume, next_ordine, datetime.now().isoformat())
    )
    new_id = cursor.lastrowid
    conn.commit()
    cursor.execute('SELECT id, proiect_id, nume, ordine, created_at FROM checklist_categorii WHERE id = ?', (new_id,))
    row = dict(cursor.fetchone())
    conn.close()
    return jsonify(row), 201

@app.route('/api/checklist-categorii/<int:cat_id>', methods=['PUT'])
@login_required
def update_checklist_categorie(cat_id):
    data = request.json or {}
    conn = get_db()
    cursor = conn.cursor()
    if 'nume' in data:
        nume = (data.get('nume') or '').strip()
        if not nume:
            return jsonify({'error': 'Nume required'}), 400
        cursor.execute('UPDATE checklist_categorii SET nume = ? WHERE id = ?', (nume, cat_id))
    if 'ordine' in data:
        cursor.execute('UPDATE checklist_categorii SET ordine = ? WHERE id = ?', (int(data['ordine']), cat_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/checklist-categorii/<int:cat_id>', methods=['DELETE'])
@login_required
def delete_checklist_categorie(cat_id):
    # ?move=1 (default) -> orphan the items into "Fara categorie".
    # ?move=0           -> hard delete the items together with the category.
    move = request.args.get('move', '1') == '1'
    conn = get_db()
    cursor = conn.cursor()
    if move:
        cursor.execute('UPDATE checklist_pif SET categorie_id = NULL WHERE categorie_id = ?', (cat_id,))
    else:
        cursor.execute('DELETE FROM checklist_pif WHERE categorie_id = ?', (cat_id,))
    cursor.execute('DELETE FROM checklist_categorii WHERE id = ?', (cat_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'moved': move})

# ============ JURNAL ============

@app.route('/api/proiecte/<project_id>/jurnal', methods=['GET'])
@login_required
def get_jurnal(project_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM jurnal WHERE proiect_id = ? ORDER BY data DESC', (project_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/proiecte/<project_id>/jurnal', methods=['POST'])
@login_required
def create_jurnal_entry(project_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    entry_id = data.get('id') or generate_uuid()
    
    cursor.execute('''
        INSERT INTO jurnal (id, proiect_id, data, continut, created_at)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        entry_id,
        project_id,
        data.get('data', now[:10]),
        data.get('continut', ''),
        now
    ))
    
    conn.commit()
    conn.close()
    
    return jsonify({'id': entry_id}), 201

@app.route('/api/jurnal/<entry_id>', methods=['DELETE'])
@login_required
def delete_jurnal_entry(entry_id):
    conn = get_db()
    cursor = conn.cursor()

    # If this jurnal entry was created by stop-with-note, there is a paired
    # timer_session row with stop_time within ~2 min of jurnal.created_at.
    # Cascade-delete it so the user doesn't need two clicks (delete the note,
    # then delete the leftover "Timer fără notă" that pops up afterward).
    cursor.execute(
        'SELECT proiect_id, created_at FROM jurnal WHERE id = ?',
        (entry_id,)
    )
    row = cursor.fetchone()
    if row:
        try:
            j_time = datetime.fromisoformat(row['created_at'])
            cursor.execute('''
                SELECT id, stop_time FROM timer_sessions
                WHERE proiect_id = ? AND stop_time IS NOT NULL
            ''', (row['proiect_id'],))
            for s in cursor.fetchall():
                s_stop = datetime.fromisoformat(s['stop_time'])
                if abs((s_stop - j_time).total_seconds()) < 120:
                    cursor.execute('DELETE FROM timer_sessions WHERE id = ?', (s['id'],))
                    break
        except (ValueError, TypeError):
            pass

    cursor.execute('DELETE FROM jurnal WHERE id = ?', (entry_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Jurnal entry deleted'})

@app.route('/api/jurnal/all', methods=['GET'])
@login_required
def get_all_jurnal():
    """Returnează toate intrările de jurnal cu numele proiectului."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT j.id, j.proiect_id, j.data, j.continut, j.created_at, p.nume as project_name
        FROM jurnal j
        JOIN proiecte p ON j.proiect_id = p.id
        ORDER BY j.data DESC
        LIMIT 200
    ''')
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify(rows)

# ============ TIMER ============

@app.route('/api/proiecte/<project_id>/timer/start', methods=['POST'])
@login_required
def start_timer(project_id):
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    session_id = generate_uuid()
    
    cursor.execute('''
        INSERT INTO timer_sessions (id, proiect_id, start_time)
        VALUES (?, ?, ?)
    ''', (session_id, project_id, now))
    
    conn.commit()
    conn.close()
    
    return jsonify({'id': session_id, 'start_time': now})

@app.route('/api/proiecte/<project_id>/timer/stop', methods=['POST'])
@login_required
def stop_timer(project_id):
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    
    # Find active session — only the standalone project timer (task_id IS NULL),
    # never a per-task timer session.
    cursor.execute('''
        SELECT * FROM timer_sessions
        WHERE proiect_id = ? AND stop_time IS NULL AND task_id IS NULL
        ORDER BY start_time DESC LIMIT 1
    ''', (project_id,))
    session = cursor.fetchone()
    
    if session is None:
        conn.close()
        return jsonify({'error': 'No active timer session'}), 404
    
    start_time = datetime.fromisoformat(session['start_time'])
    stop_time = datetime.fromisoformat(now)
    duration = int((stop_time - start_time).total_seconds())
    
    cursor.execute('''
        UPDATE timer_sessions 
        SET stop_time = ?, durata_secunde = ?
        WHERE id = ?
    ''', (now, duration, session['id']))
    
    conn.commit()
    conn.close()
    
    return jsonify({'id': session['id'], 'stop_time': now, 'durata_secunde': duration})

@app.route('/api/proiecte/<project_id>/timer/stop-with-note', methods=['POST'])
@login_required
def stop_timer_with_note(project_id):
    conn = get_db()
    cursor = conn.cursor()

    now = datetime.now().isoformat()

    cursor.execute('''
        SELECT * FROM timer_sessions
        WHERE proiect_id = ? AND stop_time IS NULL AND task_id IS NULL
        ORDER BY start_time DESC LIMIT 1
    ''', (project_id,))
    timer_session = cursor.fetchone()

    if timer_session is None:
        conn.close()
        return jsonify({'error': 'No active timer session'}), 404

    start_time = datetime.fromisoformat(timer_session['start_time'])
    stop_time = datetime.fromisoformat(now)
    duration = int((stop_time - start_time).total_seconds())

    cursor.execute('''
        UPDATE timer_sessions
        SET stop_time = ?, durata_secunde = ?
        WHERE id = ?
    ''', (now, duration, timer_session['id']))

    data = request.json or {}
    titlu = data.get('titlu', 'Activitate')
    note = data.get('note', '')

    hours = duration / 3600
    continut = f"[{titlu}] {note} — {hours:.1f}h ({duration // 60}min)".strip()

    entry_id = generate_uuid()
    cursor.execute('''
        INSERT INTO jurnal (id, proiect_id, data, continut, created_at)
        VALUES (?, ?, ?, ?, ?)
    ''', (entry_id, project_id, now[:10], continut, now))

    conn.commit()
    conn.close()

    return jsonify({
        'timer_id': timer_session['id'],
        'jurnal_id': entry_id,
        'durata_secunde': duration,
        'continut': continut
    })


@app.route('/api/timer/<session_id>', methods=['DELETE'])
@login_required
def delete_timer_session(session_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM timer_sessions WHERE id = ?', (session_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Timer session deleted'})

# ============ PER-TASK TIMER ============
# Time tracked against a specific task. Stored in timer_sessions with task_id
# set, kept isolated from the standalone project timer (which uses task_id NULL).

@app.route('/api/tasks/<task_id>/timer', methods=['GET'])
@login_required
def get_task_timer(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT start_time FROM timer_sessions
        WHERE task_id = ? AND stop_time IS NULL
        ORDER BY start_time DESC LIMIT 1
    ''', (task_id,))
    running = cursor.fetchone()
    cursor.execute(
        'SELECT COALESCE(SUM(durata_secunde), 0) FROM timer_sessions '
        'WHERE task_id = ? AND durata_secunde IS NOT NULL',
        (task_id,)
    )
    total = cursor.fetchone()[0]
    conn.close()
    return jsonify({
        'running': running is not None,
        'running_since': running['start_time'] if running else None,
        'total_secunde': total,
    })

@app.route('/api/tasks/<task_id>/timer/start', methods=['POST'])
@login_required
def start_task_timer(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT proiect_id FROM tasks WHERE id = ?', (task_id,))
    task = cursor.fetchone()
    if not task:
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    now = datetime.now().isoformat()
    # Close any already-running session for this task (avoid duplicates).
    cursor.execute('SELECT id, start_time FROM timer_sessions WHERE task_id = ? AND stop_time IS NULL', (task_id,))
    for r in cursor.fetchall():
        dur = int((datetime.fromisoformat(now) - datetime.fromisoformat(r['start_time'])).total_seconds())
        cursor.execute('UPDATE timer_sessions SET stop_time = ?, durata_secunde = ? WHERE id = ?',
                       (now, dur, r['id']))
    session_id = generate_uuid()
    cursor.execute(
        'INSERT INTO timer_sessions (id, proiect_id, task_id, start_time) VALUES (?, ?, ?, ?)',
        (session_id, task['proiect_id'], task_id, now)
    )
    conn.commit()
    conn.close()
    return jsonify({'id': session_id, 'start_time': now})

@app.route('/api/tasks/<task_id>/timer/stop', methods=['POST'])
@login_required
def stop_task_timer(task_id):
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().isoformat()
    cursor.execute('''
        SELECT id, start_time FROM timer_sessions
        WHERE task_id = ? AND stop_time IS NULL
        ORDER BY start_time DESC LIMIT 1
    ''', (task_id,))
    session = cursor.fetchone()
    if not session:
        conn.close()
        return jsonify({'error': 'No running timer for this task'}), 404
    dur = int((datetime.fromisoformat(now) - datetime.fromisoformat(session['start_time'])).total_seconds())
    cursor.execute('UPDATE timer_sessions SET stop_time = ?, durata_secunde = ? WHERE id = ?',
                   (now, dur, session['id']))
    cursor.execute(
        'SELECT COALESCE(SUM(durata_secunde), 0) FROM timer_sessions '
        'WHERE task_id = ? AND durata_secunde IS NOT NULL',
        (task_id,)
    )
    total = cursor.fetchone()[0]
    conn.commit()
    conn.close()
    return jsonify({'durata_secunde': dur, 'total_secunde': total})

@app.route('/api/proiecte/<project_id>/timer', methods=['GET'])
@login_required
def get_timer_sessions(project_id):
    conn = get_db()
    cursor = conn.cursor()
    # Standalone project timer only — per-task sessions are shown in the task modal.
    cursor.execute('''
        SELECT * FROM timer_sessions
        WHERE proiect_id = ? AND task_id IS NULL
        ORDER BY start_time DESC
    ''', (project_id,))
    rows = cursor.fetchall()
    conn.close()
    
    sessions = [row_to_dict(row) for row in rows]
    total = sum(s['durata_secunde'] or 0 for s in sessions)
    
    return jsonify({'sessions': sessions, 'total_secunde': total})

# ============ GLOBAL TASK TIMER ============
# Time tracking for daily (global) tasks. Stored in global_task_sessions,
# kept separate from timer_sessions (which is project-scoped, proiect_id NOT NULL).

@app.route('/api/global-tasks/<task_id>/timer', methods=['GET'])
@login_required
def get_global_task_timer(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''SELECT start_time FROM global_task_sessions
                      WHERE global_task_id = ? AND stop_time IS NULL
                      ORDER BY start_time DESC LIMIT 1''', (task_id,))
    running = cursor.fetchone()
    cursor.execute('SELECT COALESCE(SUM(durata_secunde), 0) FROM global_task_sessions '
                   'WHERE global_task_id = ? AND durata_secunde IS NOT NULL', (task_id,))
    total = cursor.fetchone()[0]
    cursor.execute('SELECT * FROM global_task_sessions WHERE global_task_id = ? '
                   'ORDER BY start_time DESC', (task_id,))
    sessions = [row_to_dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({
        'running': running is not None,
        'running_since': running['start_time'] if running else None,
        'total_secunde': total,
        'sessions': sessions,
    })


@app.route('/api/global-tasks/<task_id>/timer/start', methods=['POST'])
@login_required
def start_global_task_timer(task_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM global_tasks WHERE id = ?', (task_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    now = datetime.now().isoformat()
    # Close any already-running session for this task (avoid duplicates).
    cursor.execute('SELECT id, start_time FROM global_task_sessions '
                   'WHERE global_task_id = ? AND stop_time IS NULL', (task_id,))
    for r in cursor.fetchall():
        dur = int((datetime.fromisoformat(now) - datetime.fromisoformat(r['start_time'])).total_seconds())
        cursor.execute('UPDATE global_task_sessions SET stop_time = ?, durata_secunde = ? WHERE id = ?',
                       (now, max(0, dur), r['id']))
    session_id = generate_uuid()
    cursor.execute('INSERT INTO global_task_sessions (id, global_task_id, start_time) VALUES (?, ?, ?)',
                   (session_id, task_id, now))
    conn.commit()
    conn.close()
    return jsonify({'id': session_id, 'start_time': now})


@app.route('/api/global-tasks/<task_id>/timer/stop', methods=['POST'])
@login_required
def stop_global_task_timer(task_id):
    conn = get_db()
    cursor = conn.cursor()
    now = datetime.now().isoformat()
    cursor.execute('''SELECT id, start_time FROM global_task_sessions
                      WHERE global_task_id = ? AND stop_time IS NULL
                      ORDER BY start_time DESC LIMIT 1''', (task_id,))
    session = cursor.fetchone()
    if not session:
        conn.close()
        return jsonify({'error': 'No running timer for this task'}), 404
    dur = max(0, int((datetime.fromisoformat(now) - datetime.fromisoformat(session['start_time'])).total_seconds()))
    cursor.execute('UPDATE global_task_sessions SET stop_time = ?, durata_secunde = ? WHERE id = ?',
                   (now, dur, session['id']))
    cursor.execute('SELECT COALESCE(SUM(durata_secunde), 0) FROM global_task_sessions '
                   'WHERE global_task_id = ? AND durata_secunde IS NOT NULL', (task_id,))
    total = cursor.fetchone()[0]
    conn.commit()
    conn.close()
    return jsonify({'durata_secunde': dur, 'total_secunde': total})


@app.route('/api/global-task-timer/<session_id>', methods=['DELETE'])
@login_required
def delete_global_task_session(session_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM global_task_sessions WHERE id = ?', (session_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Session deleted'})


# ============ MANUAL TIMER ENTRIES ============
# Log worked time without having run a live timer (forgot to start it).
# Body: {data: 'YYYY-MM-DD', durata_secunde: int}.

def _manual_session_times(data_str, durata_secunde):
    """Manual time entry -> (start_iso, stop_iso). Anchors the session at noon
    on the given date so timezone math can never shift it to another day."""
    try:
        d = datetime.fromisoformat((data_str or '')[:10])
    except (ValueError, TypeError):
        d = datetime.now()
    start = d.replace(hour=12, minute=0, second=0, microsecond=0)
    stop = start + timedelta(seconds=int(durata_secunde))
    return start.isoformat(), stop.isoformat()


@app.route('/api/proiecte/<project_id>/timer/manual', methods=['POST'])
@login_required
def add_manual_timer(project_id):
    """Manual time entry on the project's standalone timer."""
    data = request.json or {}
    try:
        dur = int(data.get('durata_secunde') or 0)
    except (ValueError, TypeError):
        dur = 0
    if dur <= 0:
        return jsonify({'error': 'Durata invalidă'}), 400
    start, stop = _manual_session_times(data.get('data'), dur)
    conn = get_db()
    cursor = conn.cursor()
    sid = generate_uuid()
    cursor.execute('INSERT INTO timer_sessions (id, proiect_id, start_time, stop_time, durata_secunde) '
                   'VALUES (?, ?, ?, ?, ?)', (sid, project_id, start, stop, dur))
    conn.commit()
    conn.close()
    return jsonify({'id': sid, 'durata_secunde': dur})


@app.route('/api/tasks/<task_id>/timer/manual', methods=['POST'])
@login_required
def add_manual_task_timer(task_id):
    """Manual time entry on a project task."""
    data = request.json or {}
    try:
        dur = int(data.get('durata_secunde') or 0)
    except (ValueError, TypeError):
        dur = 0
    if dur <= 0:
        return jsonify({'error': 'Durata invalidă'}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT proiect_id FROM tasks WHERE id = ?', (task_id,))
    task = cursor.fetchone()
    if not task:
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    start, stop = _manual_session_times(data.get('data'), dur)
    sid = generate_uuid()
    cursor.execute('INSERT INTO timer_sessions (id, proiect_id, task_id, start_time, stop_time, durata_secunde) '
                   'VALUES (?, ?, ?, ?, ?, ?)', (sid, task['proiect_id'], task_id, start, stop, dur))
    cursor.execute('SELECT COALESCE(SUM(durata_secunde), 0) FROM timer_sessions '
                   'WHERE task_id = ? AND durata_secunde IS NOT NULL', (task_id,))
    total = cursor.fetchone()[0]
    conn.commit()
    conn.close()
    return jsonify({'id': sid, 'durata_secunde': dur, 'total_secunde': total})


@app.route('/api/global-tasks/<task_id>/timer/manual', methods=['POST'])
@login_required
def add_manual_global_task_timer(task_id):
    """Manual time entry on a daily (global) task."""
    data = request.json or {}
    try:
        dur = int(data.get('durata_secunde') or 0)
    except (ValueError, TypeError):
        dur = 0
    if dur <= 0:
        return jsonify({'error': 'Durata invalidă'}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM global_tasks WHERE id = ?', (task_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'error': 'Task not found'}), 404
    start, stop = _manual_session_times(data.get('data'), dur)
    sid = generate_uuid()
    cursor.execute('INSERT INTO global_task_sessions (id, global_task_id, start_time, stop_time, durata_secunde) '
                   'VALUES (?, ?, ?, ?, ?)', (sid, task_id, start, stop, dur))
    cursor.execute('SELECT COALESCE(SUM(durata_secunde), 0) FROM global_task_sessions '
                   'WHERE global_task_id = ? AND durata_secunde IS NOT NULL', (task_id,))
    total = cursor.fetchone()[0]
    conn.commit()
    conn.close()
    return jsonify({'id': sid, 'durata_secunde': dur, 'total_secunde': total})

# ============ ATTACHMENTS ============

@app.route('/api/proiecte/<project_id>/atasamente', methods=['POST'])
@login_required
def upload_atasament(project_id):
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    attachment_id = generate_uuid()
    
    # Create project folder if not exists
    project_folder = os.path.join(UPLOAD_FOLDER, project_id)
    os.makedirs(project_folder, exist_ok=True)
    
    # Save file — sanitize the client-supplied name; never trust it for a
    # filesystem path (blocks ../ traversal and overwriting server files).
    original_name = file.filename
    safe_name = secure_filename(original_name) or ('fisier_' + attachment_id[:8])
    filepath = os.path.join(project_folder, safe_name)
    if os.path.exists(filepath):
        safe_name = attachment_id[:8] + '_' + safe_name
        filepath = os.path.join(project_folder, safe_name)
    file.save(filepath)

    # Get file size
    size = os.path.getsize(filepath)

    # Determine file type (from the original name — keeps the real extension)
    ext = os.path.splitext(original_name)[1].lower()
    tip_map = {
        '.pdf': 'PDF',
        '.jpg': 'IMG', '.jpeg': 'IMG', '.png': 'IMG',
        '.gif': 'IMG', '.bmp': 'IMG', '.webp': 'IMG',
        '.msg': 'EMAIL', '.eml': 'EMAIL',
        '.doc': 'DOC', '.docx': 'DOC',
        '.xls': 'XLS', '.xlsx': 'XLS',
        '.zip': 'ZIP', '.rar': 'ZIP'
    }
    tip = tip_map.get(ext, 'ALT')
    
    cursor.execute('''
        INSERT INTO atasamente (id, proiect_id, nume_fisier, tip_fisier, dimensiune, data, cale_locala)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (attachment_id, project_id, original_name, tip, size, now[:10], filepath))
    
    conn.commit()
    conn.close()
    
    return jsonify({'id': attachment_id}), 201

@app.route('/api/proiecte/<project_id>/atasamente', methods=['GET'])
@login_required
def get_atasamente(project_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM atasamente WHERE proiect_id = ? ORDER BY data DESC', (project_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/atasamente/<attachment_id>/download', methods=['GET'])
@login_required
def download_atasament(attachment_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM atasamente WHERE id = ?', (attachment_id,))
    row = cursor.fetchone()
    conn.close()
    
    if row is None:
        return jsonify({'error': 'Attachment not found'}), 404
    
    filepath = row['cale_locala']
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found on disk'}), 404

    ext = os.path.splitext(row['nume_fisier'])[1].lower()
    inline_types = {'.pdf', '.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp'}
    as_attachment = ext not in inline_types
    return send_file(filepath, as_attachment=as_attachment, download_name=row['nume_fisier'])

@app.route('/api/atasamente/<attachment_id>', methods=['DELETE'])
@login_required
def delete_atasament(attachment_id):
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT cale_locala FROM atasamente WHERE id = ?', (attachment_id,))
    row = cursor.fetchone()
    
    if row:
        filepath = row['cale_locala']
        if os.path.exists(filepath):
            os.remove(filepath)
        
        cursor.execute('DELETE FROM atasamente WHERE id = ?', (attachment_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'message': 'Attachment deleted'})

# ============ STATS ============

@app.route('/api/stats', methods=['GET'])
@login_required
def get_stats():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status = 'in_lucru' THEN 1 ELSE 0 END) as active,
            SUM(CASE WHEN status = 'finalizat' THEN 1 ELSE 0 END) as finished
        FROM proiecte
    """)
    row = cursor.fetchone()
    conn.close()
    return jsonify({
        'total': row['total'] or 0,
        'active': row['active'] or 0,
        'finished': row['finished'] or 0
    })

@app.route('/api/stats/extended', methods=['GET'])
@login_required
def get_extended_stats():
    """Extended statistics for Chart.js dashboard"""
    conn = get_db()
    cursor = conn.cursor()

    # Projects by status
    cursor.execute("""
        SELECT status, COUNT(*) as count
        FROM proiecte
        GROUP BY status
    """)
    by_status = [{'status': row['status'], 'count': row['count']} for row in cursor.fetchall()]

    # Projects by manufacturer
    cursor.execute("""
        SELECT producator, COUNT(*) as count
        FROM proiecte
        GROUP BY producator
        ORDER BY count DESC
    """)
    by_manufacturer = [{'producator': row['producator'], 'count': row['count']} for row in cursor.fetchall()]

    # Projects per month (last 12 months)
    cursor.execute("""
        SELECT
            strftime('%Y-%m', created_at) as month,
            COUNT(*) as count
        FROM proiecte
        WHERE created_at >= date('now', '-12 months')
        GROUP BY month
        ORDER BY month ASC
    """)
    by_month = [{'month': row['month'], 'count': row['count']} for row in cursor.fetchall()]

    # Billable hours per project
    cursor.execute("""
        SELECT
            p.id,
            p.nume,
            COALESCE(SUM(t.durata_secunde), 0) as total_seconds
        FROM proiecte p
        LEFT JOIN timer_sessions t ON p.id = t.proiect_id AND t.durata_secunde IS NOT NULL
        GROUP BY p.id
        HAVING total_seconds > 0
        ORDER BY total_seconds DESC
        LIMIT 20
    """)
    hours_per_project = [{
        'id': row['id'],
        'nume': row['nume'],
        'hours': round(row['total_seconds'] / 3600, 2)
    } for row in cursor.fetchall()]

    # Total billable hours
    cursor.execute("""
        SELECT COALESCE(SUM(durata_secunde), 0) as total
        FROM timer_sessions
        WHERE durata_secunde IS NOT NULL
    """)
    total_hours_row = cursor.fetchone()
    total_hours = round(total_hours_row['total'] / 3600, 2) if total_hours_row else 0

    conn.close()

    return jsonify({
        'by_status': by_status,
        'by_manufacturer': by_manufacturer,
        'by_month': by_month,
        'hours_per_project': hours_per_project,
        'total_billable_hours': total_hours
    })

# ============ PHASE 2b: EXCEL EXPORT ============

@app.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    """Export data to Excel format"""
    export_type = request.args.get('type', 'projects')
    
    conn = get_db()
    cursor = conn.cursor()
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    def auto_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    if export_type == 'projects':
        ws = wb.active
        ws.title = 'Proiecte'
        
        headers = ['Nume', 'Client', 'Tip', 'Producător', 'Status', 'Data Start', 'Deadline', 'Locație']
        ws.append(headers)
        style_header(ws)
        
        cursor.execute('''
            SELECT nume, client, tip, producator, status, data_incepere, deadline, locatie
            FROM proiecte ORDER BY created_at DESC
        ''')
        for row in cursor.fetchall():
            # Map status to Romanian
            status_map = {'in_lucru': 'În Lucru', 'finalizat': 'Finalizat', 'blocat': 'Blocat', 'in_așteptare': 'În Așteptare'}
            row_data = list(row)
            row_data[4] = status_map.get(row_data[4], row_data[4])
            ws.append(row_data)
        
        auto_width(ws)
        
    elif export_type == 'tasks':
        ws = wb.active
        ws.title = 'Task-uri'
        
        headers = ['Proiect', 'Task', 'Status', 'Prioritate', 'Data Scadență', 'Data Finalizare']
        ws.append(headers)
        style_header(ws)
        
        cursor.execute('''
            SELECT p.nume, t.titlu, t.status, t.prioritate, t.data_scadenta, t.data_finalizare
            FROM tasks t
            JOIN proiecte p ON t.proiect_id = p.id
            ORDER BY t.created_at DESC
        ''')
        for row in cursor.fetchall():
            # Map status to Romanian
            status_map = {'to_do': 'To Do', 'in_lucru': 'În Lucru', 'done': 'Finalizat'}
            priority_map = {'urgent': 'Urgent', 'normal': 'Normal', 'minor': 'Minor'}
            row_data = list(row)
            row_data[2] = status_map.get(row_data[2], row_data[2])
            row_data[3] = priority_map.get(row_data[3], row_data[3])
            ws.append(row_data)
        
        auto_width(ws)
        
    elif export_type == 'hours':
        ws = wb.active
        ws.title = 'Ore'
        
        headers = ['Proiect', 'Start', 'Stop', 'Durată (ore)']
        ws.append(headers)
        style_header(ws)
        
        cursor.execute('''
            SELECT p.nume, t.start_time, t.stop_time, 
                   ROUND(CAST(t.durata_secunde AS FLOAT) / 3600, 2) as hours
            FROM timer_sessions t
            JOIN proiecte p ON t.proiect_id = p.id
            WHERE t.durata_secunde IS NOT NULL
            ORDER BY t.start_time DESC
        ''')
        for row in cursor.fetchall():
            ws.append(list(row))
        
        auto_width(ws)
    
    conn.close()
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'pif_export_{export_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    logger.info(f"Excel export: {export_type} - {filename}")
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ============ PHASE 2c: PDF EXPORT ============

@app.route('/api/export/pdf', methods=['GET'])
@login_required
def export_pdf():
    """Export project to PDF format"""
    project_id = request.args.get('project_id')
    
    if not project_id:
        return jsonify({'error': 'project_id is required'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get project info
    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (project_id,))
    project = cursor.fetchone()
    if not project:
        conn.close()
        return jsonify({'error': 'Project not found'}), 404
    
    project_dict = row_to_dict(project)
    
    # Get tasks
    cursor.execute('SELECT * FROM tasks WHERE proiect_id = ? ORDER BY ordine ASC', (project_id,))
    tasks = [row_to_dict(row) for row in cursor.fetchall()]
    
    # Get checklist
    cursor.execute('SELECT * FROM checklist_pif WHERE proiect_id = ? ORDER BY ordine ASC', (project_id,))
    checklist = [row_to_dict(row) for row in cursor.fetchall()]
    
    # Get journal entries
    cursor.execute('SELECT * FROM jurnal WHERE proiect_id = ? ORDER BY data DESC', (project_id,))
    jurnal = [row_to_dict(row) for row in cursor.fetchall()]
    
    # Get equipment
    cursor.execute('SELECT * FROM echipamente WHERE proiect_id = ?', (project_id,))
    echipamente = [row_to_dict(row) for row in cursor.fetchall()]

    # Get timer sessions
    cursor.execute('SELECT * FROM timer_sessions WHERE proiect_id = ?', (project_id,))
    timer_sessions = [row_to_dict(row) for row in cursor.fetchall()]

    # Get checklist categories (v5+)
    try:
        cursor.execute('SELECT id, nume, ordine FROM checklist_categorii WHERE proiect_id = ? ORDER BY ordine, id', (project_id,))
        checklist_cat = [row_to_dict(row) for row in cursor.fetchall()]
    except sqlite3.OperationalError:
        checklist_cat = []  # Pre-v5 schema fallback

    conn.close()

    # Calculate total hours
    total_seconds = sum(ts['durata_secunde'] or 0 for ts in timer_sessions)
    total_hours = total_seconds / 3600
    total_h = int(total_hours)
    total_m = int((total_seconds % 3600) / 60)
    total_hours_label = f"{total_h}h {total_m}m" if total_h > 0 else f"{total_m}m"

    # Status mapping
    status_map = {'in_lucru': 'In Lucru', 'finalizat': 'Finalizat', 'blocat': 'Blocat',
                  'in_asteptare': 'In Asteptare', 'in_așteptare': 'In Asteptare'}
    task_status_map = {'to_do': 'To Do', 'in_lucru': 'In Lucru', 'done': 'Finalizat'}

    # PIF design palette (matches the web UI tokens 1:1)
    PIF_BG = colors.HexColor('#11161e')
    PIF_ACCENT = colors.HexColor('#58d1c9')
    PIF_ACCENT_SOFT = colors.HexColor('#1d3835')
    PIF_TEXT = colors.HexColor('#1f2937')
    PIF_TEXT_DIM = colors.HexColor('#5a6473')
    PIF_LINE = colors.HexColor('#dbe1ea')
    PIF_SUCCESS = colors.HexColor('#16a34a')
    PIF_DANGER = colors.HexColor('#dc2626')
    PIF_WARNING = colors.HexColor('#d97706')

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm, topMargin=2*cm, bottomMargin=2*cm,
        title=f"PIF Report — {project_dict.get('nume', '')}",
        author='PIF Dashboard'
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PIFTitle', parent=styles['Heading1'],
        fontSize=18, leading=22, spaceAfter=4,
        textColor=PIF_BG, alignment=TA_LEFT, fontName='Helvetica-Bold')
    subtitle_style = ParagraphStyle(
        'PIFSubtitle', parent=styles['Normal'],
        fontSize=10, leading=14, textColor=PIF_TEXT_DIM, spaceAfter=18, fontName='Helvetica')
    heading_style = ParagraphStyle(
        'PIFHeading', parent=styles['Heading2'],
        fontSize=12, leading=16, spaceBefore=14, spaceAfter=8,
        textColor=PIF_ACCENT, fontName='Helvetica-Bold')
    subheading_style = ParagraphStyle(
        'PIFSubHeading', parent=styles['Heading3'],
        fontSize=10.5, leading=14, spaceBefore=8, spaceAfter=4,
        textColor=PIF_BG, fontName='Helvetica-Bold')
    normal_style = ParagraphStyle(
        'PIFNormal', parent=styles['Normal'],
        fontSize=9.5, leading=14, textColor=PIF_TEXT, fontName='Helvetica')

    def _safe_text(text):
        """Escape HTML special chars and normalize line breaks for ReportLab Paragraph.
        
        Handles mixed content: existing <br>/<div> tags from WYSIWYG are stripped,
        newlines are converted to <br/>, and remaining HTML chars are escaped.
        """
        if not text:
            return ''
        # Normalize existing HTML <br> variants to actual newlines
        t = re.sub(r'<br\s*/?>', '\n', text)
        # Strip <div>...</div> tags (not supported by ReportLab markup)
        t = re.sub(r'</?div[^>]*>', '', t)
        # Escape remaining HTML entities so they're treated as literal text
        t = html.escape(t)
        # Convert newlines to <br/> for ReportLab
        return t.replace('\n', '<br/>')
    small_style = ParagraphStyle(
        'PIFSmall', parent=styles['Normal'],
        fontSize=8, leading=11, textColor=PIF_TEXT_DIM, fontName='Helvetica')

    elements = []

    # Header band — accent strip + title
    is_pif = (project_dict.get('tip') == 'PIF')
    tip_label = 'PIF' if is_pif else 'Service'
    elements.append(Paragraph(f"<font color='#58d1c9'>{tip_label}</font> — {project_dict.get('nume', '')}", title_style))
    meta_bits = []
    if project_dict.get('client'): meta_bits.append(project_dict['client'])
    if project_dict.get('locatie'): meta_bits.append(project_dict['locatie'])
    meta_bits.append(f"export {datetime.now().strftime('%d.%m.%Y')}")
    elements.append(Paragraph(' · '.join(meta_bits), subtitle_style))

    # 1. Detalii Administrative
    elements.append(Paragraph("1. Detalii administrative", heading_style))
    project_info = [
        ['Client', project_dict.get('client') or '-'],
        ['Locație', project_dict.get('locatie') or '-'],
        ['Producător', project_dict.get('producator') or '-'],
        ['Echipament principal', project_dict.get('echipament_principal') or '-'],
        ['Status', status_map.get(project_dict.get('status', ''), project_dict.get('status', '-'))],
        ['Data începere', project_dict.get('data_incepere') or '-'],
        ['Deadline', project_dict.get('deadline') or '-'],
        ['PM', project_dict.get('pm') or '-'],
        ['Nr. comandă', project_dict.get('nr_comanda') or '-'],
        ['Nr. contract', project_dict.get('nr_contract') or '-'],
        ['Cod proiect', project_dict.get('cod_proiect') or '-'],
        ['Total ore lucrate', total_hours_label],
    ]
    info_table = Table(project_info, colWidths=[4.5*cm, 11*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), PIF_TEXT_DIM),
        ('TEXTCOLOR', (1, 0), (1, -1), PIF_TEXT),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (0, 0), (-1, -2), 0.3, PIF_LINE),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))

    # 2. Conținut tehnic (PIF observații / Service before+after)
    if is_pif and project_dict.get('observatii'):
        elements.append(Paragraph("2. Observații tehnice", heading_style))
        elements.append(Paragraph(_safe_text(project_dict.get('observatii', '')), normal_style))
        elements.append(Spacer(1, 6))
    if not is_pif and (project_dict.get('service_before') or project_dict.get('service_after')):
        elements.append(Paragraph("2. Fișă intervenție", heading_style))
        if project_dict.get('service_before'):
            elements.append(Paragraph("Constatări înainte de intervenție", subheading_style))
            elements.append(Paragraph(_safe_text(project_dict.get('service_before', '')), normal_style))
            elements.append(Spacer(1, 4))
        if project_dict.get('service_after'):
            elements.append(Paragraph("Acțiuni efectuate și rezultat", subheading_style))
            elements.append(Paragraph(_safe_text(project_dict.get('service_after', '')), normal_style))
            elements.append(Spacer(1, 4))

    # 3. Checklist PIF cu categorii
    if is_pif and checklist:
        elements.append(Paragraph("3. Checklist PIF", heading_style))
        # Group by category
        by_cat = {}
        for it in checklist:
            key = str(it.get('categorie_id')) if it.get('categorie_id') is not None else '0'
            by_cat.setdefault(key, []).append(it)
        ordered = sorted(checklist_cat, key=lambda c: (c.get('ordine') or 0))

        def _render_cat_block(cat_name, items):
            done = sum(1 for i in items if i.get('completed'))
            elements.append(Paragraph(f"{cat_name} ({done}/{len(items)})", subheading_style))
            rows = [['', 'Item']]
            for it in items:
                mark = '✓' if it.get('completed') else '○'
                rows.append([mark, it.get('titlu', '-')])
            t = Table(rows, colWidths=[0.7*cm, 14*cm])
            t.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('TEXTCOLOR', (0, 0), (-1, 0), PIF_TEXT_DIM),
                ('LINEBELOW', (0, 0), (-1, 0), 0.5, PIF_ACCENT),
                ('LINEBELOW', (0, 1), (-1, -1), 0.25, PIF_LINE),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('TEXTCOLOR', (0, 1), (0, -1), PIF_SUCCESS),
                ('FONTSIZE', (0, 1), (0, -1), 12),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 6))

        for cat in ordered:
            its = by_cat.get(str(cat['id']), [])
            if its:
                _render_cat_block(cat.get('nume', '?'), its)
        uncategorized = by_cat.get('0', [])
        if uncategorized:
            _render_cat_block('Fără categorie', uncategorized)

    # 4. Tasks
    if tasks:
        section_n = 4 if is_pif and checklist else 3
        elements.append(Paragraph(f"{section_n}. Listă taskuri", heading_style))
        task_data = [['#', 'Titlu', 'Status', 'Prioritate', 'Termen']]
        for i, t in enumerate(tasks, 1):
            task_data.append([
                str(i),
                t.get('titlu', '-'),
                task_status_map.get(t.get('status', ''), t.get('status', '-')),
                (t.get('prioritate') or 'Normal').capitalize(),
                t.get('data_scadenta') or '-',
            ])
        task_table = Table(task_data, colWidths=[0.8*cm, 8.5*cm, 2.2*cm, 1.8*cm, 2.2*cm])
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), PIF_ACCENT),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('LINEBELOW', (0, 0), (-1, -1), 0.25, PIF_LINE),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(task_table)
        elements.append(Spacer(1, 10))

    # 5. Echipamente
    if echipamente:
        section_n = 5 if is_pif and checklist else 4
        elements.append(Paragraph(f"{section_n}. Echipamente", heading_style))
        equip_data = [['Nume', 'Producător', 'Model', 'Serial']]
        for eq in echipamente:
            equip_data.append([
                eq.get('nume') or '-',
                eq.get('producator') or '-',
                eq.get('model') or '-',
                eq.get('serial_number') or '-',
            ])
        equip_table = Table(equip_data, colWidths=[5*cm, 3.5*cm, 4*cm, 3*cm])
        equip_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), PIF_ACCENT),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('LINEBELOW', (0, 0), (-1, -1), 0.25, PIF_LINE),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(equip_table)
        elements.append(Spacer(1, 10))

    # 6. Jurnal de lucru (with timer dedupe by ±2 min match on end_time)
    if jurnal or timer_sessions:
        section_n = 6 if is_pif and checklist else 5
        elements.append(Paragraph(f"{section_n}. Jurnal de lucru", heading_style))
        # Match each jurnal entry to a session with matching end_time within 2 min.
        sessions = list(timer_sessions)
        matched = set()
        for j in jurnal:
            jt_raw = j.get('created_at') or j.get('data')
            try: jt = datetime.fromisoformat(jt_raw).timestamp() if jt_raw else 0
            except: jt = 0
            if jt:
                for s in sessions:
                    if s['id'] in matched: continue
                    et_raw = s.get('stop_time')
                    if not et_raw: continue
                    try: et = datetime.fromisoformat(et_raw).timestamp()
                    except: continue
                    if abs(jt - et) < 120:
                        matched.add(s['id'])
                        j['_dur'] = s.get('durata_secunde') or 0
                        break
        for entry in sorted(jurnal, key=lambda e: e.get('data') or '', reverse=False)[-30:]:
            dur = entry.get('_dur')
            dur_suffix = ''
            if dur:
                dh = int(dur // 3600); dm = int((dur % 3600) // 60)
                dur_suffix = f" · <font color='#58d1c9'>{dh}h {dm}m</font>" if dh > 0 else f" · <font color='#58d1c9'>{dm}m</font>"
            elements.append(Paragraph(f"<b>{entry.get('data', '-')}</b>{dur_suffix}", subheading_style))
            elements.append(Paragraph(_safe_text(entry.get('continut') or ''), normal_style))
            elements.append(Spacer(1, 4))
        # Timer fără notă
        unmatched = [s for s in sessions if s['id'] not in matched and s.get('end_time')]
        if unmatched:
            elements.append(Paragraph("Sesiuni timer fără notă", subheading_style))
            for s in unmatched[:20]:
                dur = s.get('durata_secunde') or 0
                dh = int(dur // 3600); dm = int((dur % 3600) // 60)
                date = (s.get('start_time') or '')[:10]
                elements.append(Paragraph(f"{date} · {dh}h {dm}m" if dh > 0 else f"{date} · {dm}m", small_style))

    # Footer
    elements.append(Spacer(1, 16))
    elements.append(Paragraph(
        f"<font color='#5a6473'>Document generat automat din PIF Dashboard · {datetime.now().strftime('%d.%m.%Y %H:%M')} · Ion Ursu</font>",
        small_style))

    # Build PDF
    doc.build(elements)

    buffer.seek(0)
    safe_name = (project_dict.get('nume', 'project') or 'project').replace(' ', '_').replace('/', '_')
    if project_dict.get('cod_proiect'):
        filename = f"{project_dict['cod_proiect']}_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    else:
        filename = f"pif_report_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    logger.info(f"PDF export: {filename}")

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/export/pdf/all', methods=['GET'])
@login_required
def export_all_projects_pdf():
    """Export summary of all projects to PDF"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM proiecte ORDER BY created_at DESC')
    projects = [row_to_dict(row) for row in cursor.fetchall()]
    conn.close()
    
    status_map = {'in_lucru': 'În Lucru', 'finalizat': 'Finalizat', 'blocat': 'Blocat', 'in_așteptare': 'În Așteptare'}
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=20, alignment=TA_CENTER)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=12, spaceBefore=15, spaceAfter=10)
    
    elements = []
    elements.append(Paragraph("Raport Sumar Proiecte PIF", title_style))
    elements.append(Spacer(1, 10))
    
    # Summary table
    summary_data = [['#', 'Nume', 'Client', 'Tip', 'Status', 'Deadline']]
    for i, proj in enumerate(projects, 1):
        summary_data.append([
            str(i),
            proj.get('nume', '-')[:30],
            proj.get('client', '-')[:20],
            proj.get('tip', '-'),
            status_map.get(proj.get('status', ''), proj.get('status', '-')),
            proj.get('deadline', '-')
        ])
    
    summary_table = Table(summary_data, colWidths=[1*cm, 5*cm, 3.5*cm, 2*cm, 3*cm, 2.5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(summary_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"pif_all_projects_{datetime.now().strftime('%Y%m%d')}.pdf"
    logger.info(f"PDF export all projects: {filename}")
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

# ============ END PDF EXPORT ============

# ============ BACKUP / RESTORE ============

@app.route('/api/backup', methods=['GET'])
@login_required
def backup_database():
    conn = get_db()
    cursor = conn.cursor()
    
    backup = {}
    
    tables = ['proiecte', 'tasks', 'checklist_pif', 'jurnal', 'timer_sessions', 'atasamente', 'global_tasks', 'clienti', 'echipamente', 'project_templates']
    for table in tables:
        cursor.execute(f'SELECT * FROM {safe_table(table)}')
        rows = cursor.fetchall()
        backup[table] = [row_to_dict(row) for row in rows]
    
    conn.close()
    
    return jsonify(backup)

@app.route('/api/restore', methods=['POST'])
@login_required
def restore_database():
    data = request.json
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Run the whole clear + restore inside ONE transaction, so a bad payload
        # rolls the deletes back instead of leaving the database wiped.
        conn.execute('BEGIN TRANSACTION')
        # Clear existing data
        tables = ['proiecte', 'tasks', 'checklist_pif', 'jurnal', 'timer_sessions', 'atasamente', 'global_tasks', 'clienti', 'echipamente', 'project_templates']
        for table in tables:
            cursor.execute(f'DELETE FROM {safe_table(table)}')

        # Restore proiecte
        for p in data.get('proiecte', []):
            cursor.execute('''
                INSERT INTO proiecte (id, tip, nume, client, locatie, echipament_principal, producator,
                    cod_proiect, pm, folder_server, data_incepere, deadline, data_crearii,
                    status, observatii, nr_comanda, nr_contract, service_before, service_after,
                    confirmat_client, client_nume_confirmare, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                p.get('id'), p.get('tip'), p.get('nume'), p.get('client'), p.get('locatie'),
                p.get('echipament_principal'), p.get('producator'), p.get('cod_proiect'),
                p.get('pm'), p.get('folder_server'), p.get('data_incepere'), p.get('deadline'),
                p.get('data_crearii'), p.get('status'), p.get('observatii'), p.get('nr_comanda'),
                p.get('nr_contract'), p.get('service_before'), p.get('service_after'),
                p.get('confirmat_client', 0), p.get('client_nume_confirmare'),
                p.get('created_at'), p.get('updated_at')
            ))
        
        # Restore tasks
        for t in data.get('tasks', []):
            cursor.execute('''
                INSERT INTO tasks (id, proiect_id, titlu, status, prioritate, data_scadenta, data_finalizare, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (t.get('id'), t.get('proiect_id'), t.get('titlu'), t.get('status'),
                  t.get('prioritate'), t.get('data_scadenta'), t.get('data_finalizare'), t.get('created_at')))
        
        # Restore checklist
        for c in data.get('checklist_pif', []):
            cursor.execute('''
                INSERT INTO checklist_pif (id, proiect_id, titlu, completed, note, ordine)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (c.get('id'), c.get('proiect_id'), c.get('titlu'), c.get('completed'),
                  c.get('note'), c.get('ordine', 0)))
        
        # Restore jurnal
        for j in data.get('jurnal', []):
            cursor.execute('''
                INSERT INTO jurnal (id, proiect_id, data, continut, created_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (j.get('id'), j.get('proiect_id'), j.get('data'), j.get('continut'), j.get('created_at')))
        
        # Restore timer_sessions
        for ts in data.get('timer_sessions', []):
            cursor.execute('''
                INSERT INTO timer_sessions (id, proiect_id, start_time, stop_time, durata_secunde)
                VALUES (?, ?, ?, ?, ?)
            ''', (ts.get('id'), ts.get('proiect_id'), ts.get('start_time'),
                  ts.get('stop_time'), ts.get('durata_secunde')))
        
        # Restore atasamente (paths need to exist)
        for a in data.get('atasamente', []):
            if os.path.exists(a.get('cale_locala', '')):
                cursor.execute('''
                    INSERT INTO atasamente (id, proiect_id, nume_fisier, tip_fisier, dimensiune, data, cale_locala)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (a.get('id'), a.get('proiect_id'), a.get('nume_fisier'), a.get('tip_fisier'),
                      a.get('dimensiune'), a.get('data'), a.get('cale_locala')))
        
        # Restore global_tasks
        for gt in data.get('global_tasks', []):
            cursor.execute('''
                INSERT INTO global_tasks (id, titlu, descriere, prioritate, status, categorie, data_scadenta, data_finalizare, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (gt.get('id'), gt.get('titlu'), gt.get('descriere'), gt.get('prioritate'),
                  gt.get('status'), gt.get('categorie'), gt.get('data_scadenta'),
                  gt.get('data_finalizare'), gt.get('created_at'), gt.get('updated_at')))
        
        # Restore clienti
        for c in data.get('clienti', []):
            cursor.execute('''
                INSERT INTO clienti (id, nume, adresa, telefon, email, contact_principal, note, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (c.get('id'), c.get('nume'), c.get('adresa'), c.get('telefon'),
                  c.get('email'), c.get('contact_principal'), c.get('note'), c.get('created_at')))
        
        # Restore echipamente
        for e in data.get('echipamente', []):
            cursor.execute('''
                INSERT INTO echipamente (id, proiect_id, nume, producator, model, serial_number, params_json, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (e.get('id'), e.get('proiect_id'), e.get('nume'), e.get('producator'),
                  e.get('model'), e.get('serial_number'), e.get('params_json'),
                  e.get('created_at'), e.get('updated_at')))
        
        # Restore project_templates
        for t in data.get('project_templates', []):
            cursor.execute('''
                INSERT INTO project_templates (id, name, tip, default_checklist_json, default_tasks_json, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (t.get('id'), t.get('name'), t.get('tip'), t.get('default_checklist_json'),
                  t.get('default_tasks_json'), t.get('created_at')))
        
        conn.commit()
        conn.close()
        
        logger.info("Database restored successfully")
        return jsonify({'message': 'Database restored successfully'})
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error(f"Error restoring database: {e}")
        return jsonify({'error': 'Restaurare esuata — modificarile au fost anulate.'}), 500

# ============ PHASE 2a: CLIENTI CRUD ============

@app.route('/api/clienti', methods=['GET'])
@login_required
def get_clienti():
    conn = get_db()
    cursor = conn.cursor()
    
    search = request.args.get('search', '')
    
    if search:
        cursor.execute(
            'SELECT * FROM clienti WHERE nume LIKE ? OR adresa LIKE ? OR telefon LIKE ? ORDER BY nume',
            (f'%{search}%', f'%{search}%', f'%{search}%')
        )
    else:
        cursor.execute('SELECT * FROM clienti ORDER BY nume')
    
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/clienti', methods=['POST'])
@login_required
def create_client():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    client_id = data.get('id') or generate_uuid()
    
    cursor.execute('''
        INSERT INTO clienti (id, nume, adresa, telefon, email, contact_principal, note, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        client_id,
        data.get('nume', ''),
        data.get('adresa', ''),
        data.get('telefon', ''),
        data.get('email', ''),
        data.get('contact_principal', ''),
        data.get('note', ''),
        now
    ))
    
    conn.commit()
    conn.close()
    
    logger.info(f"Client created: {client_id} - {data.get('nume', '')}")
    return jsonify({'id': client_id, 'message': 'Client created'}), 201

@app.route('/api/clienti/<client_id>', methods=['GET'])
@login_required
def get_client(client_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM clienti WHERE id = ?', (client_id,))
    row = cursor.fetchone()
    conn.close()
    
    if row is None:
        return jsonify({'error': 'Client not found'}), 404
    
    return jsonify(row_to_dict(row))

@app.route('/api/clienti/<client_id>', methods=['PUT'])
@login_required
def update_client(client_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE clienti SET
            nume = COALESCE(?, nume),
            adresa = COALESCE(?, adresa),
            telefon = COALESCE(?, telefon),
            email = COALESCE(?, email),
            contact_principal = COALESCE(?, contact_principal),
            note = COALESCE(?, note)
        WHERE id = ?
    ''', (
        data.get('nume'),
        data.get('adresa'),
        data.get('telefon'),
        data.get('email'),
        data.get('contact_principal'),
        data.get('note'),
        client_id
    ))
    
    conn.commit()
    conn.close()
    
    logger.info(f"Client updated: {client_id}")
    return jsonify({'message': 'Client updated'})

@app.route('/api/clienti/<client_id>', methods=['DELETE'])
@login_required
def delete_client(client_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM clienti WHERE id = ?', (client_id,))
    conn.commit()
    conn.close()
    
    logger.info(f"Client deleted: {client_id}")
    return jsonify({'message': 'Client deleted'})

# ============ PHASE 2a: ECHIPAMENTE CRUD ============

@app.route('/api/proiecte/<project_id>/echipamente', methods=['GET'])
@login_required
def get_echipamente(project_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM echipamente WHERE proiect_id = ? ORDER BY created_at DESC', (project_id,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/proiecte/<project_id>/echipamente', methods=['POST'])
@login_required
def create_echipament(project_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    echipament_id = data.get('id') or generate_uuid()
    
    # Parse params_text into key-value pairs
    params_text = data.get('params_text', '')
    params_dict = {}
    if params_text:
        for line in params_text.strip().split('\n'):
            line = line.strip()
            if '=' in line:
                key, value = line.split('=', 1)
                params_dict[key.strip()] = value.strip()
    
    cursor.execute('''
        INSERT INTO echipamente (id, proiect_id, nume, producator, model, serial_number, params_json, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        echipament_id,
        project_id,
        data.get('nume', ''),
        data.get('producator', 'Altul'),
        data.get('model', ''),
        data.get('serial_number', ''),
        json.dumps(params_dict),
        now,
        now
    ))
    
    conn.commit()
    conn.close()
    
    logger.info(f"Equipment created: {echipament_id} for project {project_id}")
    return jsonify({'id': echipament_id, 'message': 'Equipment created'}), 201

@app.route('/api/echipamente/<echipament_id>', methods=['GET'])
@login_required
def get_echipament(echipament_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM echipamente WHERE id = ?', (echipament_id,))
    row = cursor.fetchone()
    conn.close()
    
    if row is None:
        return jsonify({'error': 'Equipment not found'}), 404
    
    return jsonify(row_to_dict(row))

@app.route('/api/echipamente/<echipament_id>', methods=['PUT'])
@login_required
def update_echipament(echipament_id):
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    
    # Parse params_text if provided
    params_json = data.get('params_json')
    if data.get('params_text'):
        params_text = data.get('params_text', '')
        params_dict = {}
        if params_text:
            for line in params_text.strip().split('\n'):
                line = line.strip()
                if '=' in line:
                    key, value = line.split('=', 1)
                    params_dict[key.strip()] = value.strip()
        params_json = json.dumps(params_dict)
    
    cursor.execute('''
        UPDATE echipamente SET
            nume = COALESCE(?, nume),
            producator = COALESCE(?, producator),
            model = COALESCE(?, model),
            serial_number = COALESCE(?, serial_number),
            params_json = COALESCE(?, params_json),
            updated_at = ?
        WHERE id = ?
    ''', (
        data.get('nume'),
        data.get('producator'),
        data.get('model'),
        data.get('serial_number'),
        params_json,
        now,
        echipament_id
    ))
    
    conn.commit()
    conn.close()
    
    logger.info(f"Equipment updated: {echipament_id}")
    return jsonify({'message': 'Equipment updated'})

@app.route('/api/echipamente/<echipament_id>', methods=['DELETE'])
@login_required
def delete_echipament(echipament_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM echipamente WHERE id = ?', (echipament_id,))
    conn.commit()
    conn.close()
    
    logger.info(f"Equipment deleted: {echipament_id}")
    return jsonify({'message': 'Equipment deleted'})

# ============ IMPORT PARAMETRI DIN EXPORT PRODUCATOR ============

def _familie_from_echipament(producator: str, model: str) -> str:
    """Mapeaza producator+model la familie pentru join cu parametri_master.

    Numele familiilor in DB (validat cu scripts/audit_pdf.py FAMILIES):
        ACS580, ACS880, Danfoss_VLT_FC302, Lenze_i550, Lenze_i950,
        SINAMICS_G120, SINAMICS_G130_G150, SINAMICS_S120_S150
    """
    producator = (producator or '').strip()
    model = (model or '').strip()
    p_low = producator.lower()
    m_low = model.lower()
    if 'danfoss' in p_low:
        return 'Danfoss_VLT_FC302'
    if 'abb' in p_low:
        if '880' in m_low: return 'ACS880'
        return 'ACS580'
    if 'siemens' in p_low:
        if 's120' in m_low or 's150' in m_low: return 'SINAMICS_S120_S150'
        if 'g130' in m_low or 'g150' in m_low: return 'SINAMICS_G130_G150'
        return 'SINAMICS_G120'
    if 'lenze' in p_low:
        if '950' in m_low: return 'Lenze_i950'
        return 'Lenze_i550'
    return ''


@app.route('/api/import-params/preview', methods=['POST'])
@login_required
def preview_import_params():
    """Parseaza un export de parametri modificati (producator software) si returneaza preview.

    Form fields:
      file: fisierul exportat (text Danfoss .txt, in viitor PDF Lenze/Siemens)
      producator: numele producatorului (Danfoss, Lenze, Siemens, ABB)
      model: optional, pentru determinarea familiei (ACS580 vs ACS880 etc)

    Preview-ul nu modifica DB. Persistarea se face prin PUT /api/echipamente/<id>
    cu params_json deja merge-uit pe client.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Fisier lipsa (field "file")'}), 400
    upload = request.files['file']
    if not upload.filename:
        return jsonify({'error': 'Fisier gol'}), 400

    producator = (request.form.get('producator') or '').strip()
    model = (request.form.get('model') or '').strip()
    if not producator:
        return jsonify({'error': 'producator lipsa'}), 400

    try:
        raw = upload.read()
    except Exception as e:
        logger.exception("Eroare citire fisier import params")
        return jsonify({'error': f'Eroare citire fisier: {e}'}), 400

    detected, parsed = parse_for_producator(producator, raw, upload.filename or '')
    if detected is None or parsed is None:
        return jsonify({
            'error': f'Producator "{producator}" nu este inca suportat pentru import. Suportate: Danfoss, Lenze, Siemens, ABB.'
        }), 400

    # Imbogateste cu descriere_scurta din parametri_master
    familie = _familie_from_echipament(producator, model)
    if familie and parsed:
        conn = get_db()
        cursor = conn.cursor()
        codes = [p['db_id'] for p in parsed]
        placeholders = ','.join('?' * len(codes))
        cursor.execute(
            f'SELECT parametru, descriere_scurta FROM parametri_master WHERE familie = ? AND parametru IN ({placeholders})',
            [familie] + codes
        )
        desc_map = {r['parametru']: r['descriere_scurta'] for r in cursor.fetchall()}
        conn.close()
        for p in parsed:
            ds = desc_map.get(p['db_id'])
            if ds:
                p['descriere_db'] = ds

    return jsonify({
        'producator_detected': detected,
        'familie': familie,
        'filename': upload.filename,
        'count': len(parsed),
        'conflicts': sum(1 for p in parsed if p.get('conflict')),
        'params': parsed,
    })


# ============ PHASE 2a: PROJECT TEMPLATES CRUD ============

@app.route('/api/templates', methods=['GET'])
@login_required
def get_templates():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM project_templates ORDER BY name')
    rows = cursor.fetchall()
    conn.close()
    return jsonify([row_to_dict(row) for row in rows])

@app.route('/api/templates', methods=['POST'])
@login_required
def create_template():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    template_id = data.get('id') or generate_uuid()
    
    cursor.execute('''
        INSERT INTO project_templates (id, name, tip, default_checklist_json, default_tasks_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (
        template_id,
        data.get('name', ''),
        data.get('tip', 'PIF'),
        json.dumps(data.get('default_checklist', [])),
        json.dumps(data.get('default_tasks', [])),
        now
    ))
    
    conn.commit()
    conn.close()
    
    logger.info(f"Template created: {template_id} - {data.get('name', '')}")
    return jsonify({'id': template_id, 'message': 'Template created'}), 201

@app.route('/api/templates/<template_id>', methods=['GET'])
@login_required
def get_template(template_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM project_templates WHERE id = ?', (template_id,))
    row = cursor.fetchone()
    conn.close()
    
    if row is None:
        return jsonify({'error': 'Template not found'}), 404
    
    return jsonify(row_to_dict(row))

@app.route('/api/templates/<template_id>', methods=['DELETE'])
@login_required
def delete_template(template_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM project_templates WHERE id = ?', (template_id,))
    conn.commit()
    conn.close()
    
    logger.info(f"Template deleted: {template_id}")
    return jsonify({'message': 'Template deleted'})

@app.route('/api/templates/create-from-project/<project_id>', methods=['POST'])
@login_required
def create_template_from_project(project_id):
    """Create a new template from an existing project's checklist and tasks"""
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    
    now = datetime.now().isoformat()
    template_id = generate_uuid()
    
    # Get project info
    cursor.execute('SELECT tip FROM proiecte WHERE id = ?', (project_id,))
    project = cursor.fetchone()
    if not project:
        conn.close()
        return jsonify({'error': 'Project not found'}), 404
    
    # Get checklist items
    cursor.execute('SELECT titlu FROM checklist_pif WHERE proiect_id = ?', (project_id,))
    checklist_items = [row['titlu'] for row in cursor.fetchall()]
    
    # Get tasks
    cursor.execute('SELECT titlu, prioritate FROM tasks WHERE proiect_id = ?', (project_id,))
    task_items = [{'titlu': row['titlu'], 'prioritate': row['prioritate']} for row in cursor.fetchall()]
    
    cursor.execute('''
        INSERT INTO project_templates (id, name, tip, default_checklist_json, default_tasks_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (
        template_id,
        data.get('name', 'Template from Project'),
        project['tip'],
        json.dumps(checklist_items),
        json.dumps(task_items),
        now
    ))
    
    conn.commit()
    conn.close()
    
    logger.info(f"Template created from project {project_id}: {template_id}")
    return jsonify({'id': template_id, 'message': 'Template created from project'}), 201

# ============ ADMIN: DB DOWNLOAD / UPLOAD ============

@app.route('/admin/db-upload', methods=['GET'])
@login_required
def admin_db_upload_page():
    """Minimal HTML form for uploading a local DB file."""
    return '''<!doctype html>
<html><head><meta charset="utf-8"><title>DB upload</title>
<style>
body{background:#0a0d12;color:#e3e8ef;font-family:system-ui,sans-serif;
     max-width:560px;margin:60px auto;padding:0 24px}
h1{font-size:20px;margin:0 0 16px}
.box{background:#161c26;border:1px solid #232a36;border-radius:10px;padding:24px}
input[type=file]{margin:16px 0;color:#e3e8ef}
button{background:#58d1c9;color:#0a0d12;border:0;padding:10px 18px;
       border-radius:8px;font-weight:600;cursor:pointer}
button:disabled{opacity:.5;cursor:wait}
.note{color:#9aa4b2;font-size:13px;margin-top:14px}
pre{background:#0a0d12;border:1px solid #232a36;border-radius:6px;
    padding:10px;font-size:12px;overflow:auto;max-height:280px}
.ok{color:#66d19e}.err{color:#f97066}
</style></head>
<body><div class="box">
<h1>Upload DB SQLite</h1>
<form id="f">
<input type="file" name="db" accept=".db" required>
<br><button id="submit" type="submit">Upload (inlocuieste DB serverului)</button>
</form>
<div class="note">DB-ul curent va fi salvat in <code>backups/</code> automat inainte de inlocuire.</div>
<pre id="out"></pre>
</div>
<script>
const f=document.getElementById('f'),btn=document.getElementById('submit'),out=document.getElementById('out');
f.addEventListener('submit',async e=>{
  e.preventDefault();btn.disabled=true;out.textContent='Uploading...';
  const fd=new FormData(f);
  try{
    const r=await fetch('/api/admin/db-upload',{method:'POST',body:fd});
    const j=await r.json();
    out.textContent=JSON.stringify(j,null,2);
    out.className=r.ok?'ok':'err';
  }catch(err){out.textContent=err.message;out.className='err';}
  btn.disabled=false;
});
</script>
</body></html>'''


@app.route('/api/admin/db-upload', methods=['POST'])
@login_required
def admin_db_upload():
    """Replace the server SQLite DB with an uploaded copy.
    Expects a multipart form field named 'db' with a .db file.
    Validates SQLite header, backs up the existing DB, then atomic-replaces it.
    """
    import shutil
    import tempfile
    from database import DATABASE_PATH

    if 'db' not in request.files:
        return jsonify({'error': "missing form field 'db'"}), 400
    f = request.files['db']
    if not f or f.filename == '':
        return jsonify({'error': 'empty file'}), 400

    dir_name = os.path.dirname(DATABASE_PATH) or '.'
    fd, tmp_path = tempfile.mkstemp(prefix='dbupload_', suffix='.db', dir=dir_name)
    os.close(fd)
    try:
        f.save(tmp_path)
        with open(tmp_path, 'rb') as fh:
            magic = fh.read(16)
        if not magic.startswith(b'SQLite format 3'):
            os.unlink(tmp_path)
            return jsonify({'error': 'not a valid SQLite database'}), 400

        # Beyond the magic header: verify the file is a sound SQLite DB and
        # actually looks like a PIF database before replacing the live one.
        try:
            _chk = sqlite3.connect(tmp_path)
            _integrity = _chk.execute('PRAGMA integrity_check').fetchone()
            _has_core = _chk.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='proiecte'"
            ).fetchone()
            _chk.close()
        except Exception:
            os.unlink(tmp_path)
            return jsonify({'error': 'fisierul nu este o baza SQLite utilizabila'}), 400
        if not _integrity or _integrity[0] != 'ok':
            os.unlink(tmp_path)
            return jsonify({'error': 'baza incarcata a esuat integrity_check'}), 400
        if not _has_core:
            os.unlink(tmp_path)
            return jsonify({'error': 'baza incarcata nu contine structura PIF (tabelul proiecte)'}), 400

        backups_dir = os.path.join(dir_name, 'backups')
        os.makedirs(backups_dir, exist_ok=True)
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(backups_dir, f'pif_dashboard_pre_upload_{stamp}.db')
        if os.path.exists(DATABASE_PATH):
            shutil.copy2(DATABASE_PATH, backup_path)

        os.replace(tmp_path, DATABASE_PATH)

        return jsonify({
            'status': 'ok',
            'backup': backup_path,
            'replaced': DATABASE_PATH,
            'size': os.path.getsize(DATABASE_PATH),
        })
    except Exception as e:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        logger.error(f"db-upload failed: {e}")
        return jsonify({'error': 'Incarcarea bazei a esuat.'}), 500


@app.route('/api/admin/db-dump', methods=['GET'])
@login_required
def admin_db_dump():
    """Stream o copie consistenta a DB-ului SQLite pentru audit local.
    Foloseste sqlite3 backup API ca sa nu blocheze scrieri concurente.
    """
    import tempfile
    import sqlite3 as _sql3
    from database import DATABASE_PATH

    if not os.path.exists(DATABASE_PATH):
        return jsonify({'error': 'DB not found on server'}), 404

    # Hot backup la un fisier temporar (safe vs WAL)
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp.close()
    src = _sql3.connect(DATABASE_PATH)
    dst = _sql3.connect(tmp.name)
    try:
        src.backup(dst)
    finally:
        dst.close()
        src.close()

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=f'pif_dashboard_{timestamp}.db',
        mimetype='application/octet-stream',
    )


# ============ PARAMETRI API ============

@app.route('/api/parametri/familii', methods=['GET'])
@login_required
def get_parametri_familii():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT familie, COUNT(*) as count FROM parametri_master GROUP BY familie ORDER BY familie")
    families = [{'familie': row['familie'], 'count': row['count']} for row in cursor.fetchall()]
    conn.close()
    return jsonify({'families': families})

@app.route('/api/parametri', methods=['GET'])
@login_required
def get_parametri():
    conn = get_db()
    cursor = conn.cursor()
    
    search = request.args.get('search', '')
    familie = request.args.get('familie', '')
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 50, type=int)
    offset = (page - 1) * limit
    
    # Build count query first
    count_query = "SELECT COUNT(*) FROM parametri_master WHERE 1=1"
    query = "SELECT id, familie, parametru, descriere_scurta, descriere, acces, tip_date, valoare_default, valoare_default_str, min, max, unitate, pagina, creat_la FROM parametri_master WHERE 1=1"
    params = []
    
    if search:
        clause = " AND (parametru LIKE ? OR descriere LIKE ?)"
        count_query += clause
        query += clause
        params.extend([f'%{search}%', f'%{search}%'])
    
    if familie:
        clause = " AND familie = ?"
        count_query += clause
        query += clause
        params.append(familie)
    
    # Get total count
    cursor.execute(count_query, params)
    total = cursor.fetchone()[0]
    
    # Get paginated results
    query += " ORDER BY familie, parametru LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    
    cursor.execute(query, params)
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    
    return jsonify({
        'params': rows,
        'total': total,
        'page': page,
        'limit': limit,
        'totalPages': max(1, (total + limit - 1) // limit)
    })

# ============ BULK PARAMS (lightweight, for mobile cache) ============

@app.route('/api/parametri/bulk', methods=['GET'])
@login_required
def get_parametri_bulk():
    """Returnează toți parametrii FĂRĂ explicatie/influenteaza (lightweight)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, familie, parametru, descriere_scurta, descriere, acces, tip_date,
               valoare_default, valoare_default_str, min, max, unitate,
               pagina, creat_la
        FROM parametri_master
        ORDER BY familie, parametru
    ''')
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    
    # Sanitize: JSON nu suportă Infinity/NaN — înlocuiește cu null
    import math
    for row in rows:
        for key in ('min', 'max'):
            if row.get(key) is not None:
                try:
                    v = float(row[key])
                    if math.isinf(v) or math.isnan(v):
                        row[key] = None
                except (ValueError, TypeError):
                    pass
    
    return jsonify(rows)

PRODUCATOR_FAMILII = {
    'ABB': ['ACS580', 'ACS880'],
    'Danfoss': ['Danfoss_VLT_FC302'],
    'Lenze': ['Lenze_i550', 'Lenze_i950'],
    'Siemens': ['SINAMICS_G120', 'SINAMICS_G130_G150', 'SINAMICS_S120_S150'],
}

@app.route('/api/parametri/by-producator/<producator>', methods=['GET'])
@login_required
def get_parametri_by_producator(producator):
    """Returnează parametrii pentru toate familiile unui producător."""
    familii = PRODUCATOR_FAMILII.get(producator, [])
    if not familii:
        return jsonify([])
    conn = get_db()
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(familii))
    cursor.execute(
        f'SELECT id, parametru, descriere_scurta, familie FROM parametri_master WHERE familie IN ({placeholders})',
        familii
    )
    rows = cursor.fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ============ FAULT CODES API ============
# Drive fault / alarm / warning codes extracted from the manufacturer manuals.
# Sibling dataset to parametri_master; same producator -> familie -> list shape.

def _fault_row(row):
    """Row -> dict, decoding extra_json into an `extra` object."""
    d = dict(row)
    raw = d.pop('extra_json', None)
    if raw:
        try:
            d['extra'] = json.loads(raw)
        except (ValueError, TypeError):
            d['extra'] = None
    else:
        d['extra'] = None
    return d


@app.route('/api/fault-codes/familii', methods=['GET'])
@login_required
def get_fault_familii():
    """Producator + familie + count, for the fault-code navigation picker."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''SELECT producator, familie, COUNT(*) AS count
                      FROM fault_codes
                      GROUP BY producator, familie
                      ORDER BY producator, familie''')
    families = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'families': families})


@app.route('/api/fault-codes', methods=['GET'])
@login_required
def get_fault_codes():
    """List/filter fault codes. Query: familie, producator, search, tip, page, limit."""
    conn = get_db()
    cursor = conn.cursor()

    search = request.args.get('search', '').strip()
    familie = request.args.get('familie', '').strip()
    producator = request.args.get('producator', '').strip()
    tip = request.args.get('tip', '').strip()
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 50, type=int)
    offset = (page - 1) * limit

    where = ' WHERE 1=1'
    params = []
    if familie:
        where += ' AND familie = ?'
        params.append(familie)
    if producator:
        where += ' AND producator = ?'
        params.append(producator)
    if tip:
        where += ' AND tip = ?'
        params.append(tip)
    if search:
        where += ' AND (cod LIKE ? OR cod_secundar LIKE ? OR nume LIKE ? OR cauza LIKE ?)'
        s = f'%{search}%'
        params.extend([s, s, s, s])

    cursor.execute(f'SELECT COUNT(*) FROM {safe_table("fault_codes")}{where}', params)
    total = cursor.fetchone()[0]

    cursor.execute(
        f'''SELECT id, producator, familie, cod, cod_secundar, tip, nume, pagina
            FROM {safe_table("fault_codes")}{where} ORDER BY cod LIMIT ? OFFSET ?''',
        params + [limit, offset])
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()

    return jsonify({
        'codes': rows,
        'total': total,
        'page': page,
        'limit': limit,
        'totalPages': max(1, (total + limit - 1) // limit),
    })


@app.route('/api/fault-codes/lookup', methods=['GET'])
@login_required
def lookup_fault_code():
    """Quick lookup by code. Query: cod (required), producator/familie (optional).
    Matches cod or cod_secundar, case-insensitive."""
    cod = request.args.get('cod', '').strip()
    if not cod:
        return jsonify({'error': 'cod lipsa'}), 400
    producator = request.args.get('producator', '').strip()
    familie = request.args.get('familie', '').strip()

    conn = get_db()
    cursor = conn.cursor()
    where = ' WHERE (cod = ? COLLATE NOCASE OR cod_secundar = ? COLLATE NOCASE)'
    params = [cod, cod]
    if producator:
        where += ' AND producator = ?'
        params.append(producator)
    if familie:
        where += ' AND familie = ?'
        params.append(familie)
    cursor.execute(f'SELECT * FROM {safe_table("fault_codes")}{where} ORDER BY producator, familie', params)
    matches = [_fault_row(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'matches': matches, 'count': len(matches)})


@app.route('/api/fault-codes/<int:code_id>', methods=['GET'])
@login_required
def get_fault_code(code_id):
    """Full detail for one fault code."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM fault_codes WHERE id = ?', (code_id,))
    row = cursor.fetchone()
    conn.close()
    if row is None:
        return jsonify({'error': 'Codul nu a fost găsit'}), 404
    return jsonify(_fault_row(row))


@app.route('/api/parametri/audit', methods=['GET'])
@login_required
def parametri_audit():
    """Audit raport calitate DB parametri_master.
    Detectează anomalii grupate pe categorii, cu sample-uri.
    Optional: ?familie=ACS580 restricționează la o familie.
    """
    import re
    familie_filter = request.args.get('familie', '').strip()
    conn = get_db()
    cursor = conn.cursor()

    # Build WHERE for optional familie scope
    where = ''
    params = []
    if familie_filter:
        where = ' WHERE familie = ?'
        params = [familie_filter]

    # Total count
    cursor.execute(f'SELECT COUNT(*) FROM {safe_table("parametri_master")}{where}', params)
    total = cursor.fetchone()[0]

    # Per-familie breakdown
    cursor.execute(f'''
        SELECT familie, COUNT(*) as total,
            SUM(CASE WHEN descriere_scurta IS NULL OR TRIM(descriere_scurta)='' THEN 1 ELSE 0 END) as missing_desc,
            SUM(CASE WHEN explicatie IS NULL OR TRIM(explicatie)='' THEN 1 ELSE 0 END) as missing_explicatie,
            SUM(CASE WHEN pagina IS NULL THEN 1 ELSE 0 END) as missing_pagina,
            SUM(CASE WHEN unitate IS NULL OR TRIM(unitate)='' THEN 1 ELSE 0 END) as missing_unitate,
            SUM(CASE WHEN acces IS NULL OR TRIM(acces)='' THEN 1 ELSE 0 END) as missing_acces
        FROM {safe_table("parametri_master")}{where}
        GROUP BY familie ORDER BY familie
    ''', params)
    per_familie = [dict(r) for r in cursor.fetchall()]

    issues = {}

    # 1. Missing descriere_scurta
    cursor.execute(f'''
        SELECT id, parametru, familie FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} (descriere_scurta IS NULL OR TRIM(descriere_scurta)='')
        LIMIT 30
    ''', params)
    rows = [dict(r) for r in cursor.fetchall()]
    cursor.execute(f'''
        SELECT COUNT(*) FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} (descriere_scurta IS NULL OR TRIM(descriere_scurta)='')
    ''', params)
    issues['missing_descriere_scurta'] = {
        'count': cursor.fetchone()[0],
        'label': 'Parametri fără descriere scurtă',
        'severity': 'high',
        'samples': rows,
    }

    # 2. Descriere_scurta foarte scurtă (probabil placeholder)
    cursor.execute(f'''
        SELECT id, parametru, descriere_scurta, familie FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} LENGTH(TRIM(descriere_scurta)) BETWEEN 1 AND 10
        LIMIT 30
    ''', params)
    rows = [dict(r) for r in cursor.fetchall()]
    cursor.execute(f'''
        SELECT COUNT(*) FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} LENGTH(TRIM(descriere_scurta)) BETWEEN 1 AND 10
    ''', params)
    issues['short_descriere'] = {
        'count': cursor.fetchone()[0],
        'label': 'Descriere foarte scurtă (≤10 caractere)',
        'severity': 'medium',
        'samples': rows,
    }

    # 3. Descriere = parametru (placeholder evident)
    cursor.execute(f'''
        SELECT id, parametru, descriere_scurta, familie FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} TRIM(descriere_scurta) = TRIM(parametru)
        LIMIT 30
    ''', params)
    rows = [dict(r) for r in cursor.fetchall()]
    cursor.execute(f'''
        SELECT COUNT(*) FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} TRIM(descriere_scurta) = TRIM(parametru)
    ''', params)
    issues['descriere_equals_code'] = {
        'count': cursor.fetchone()[0],
        'label': 'Descriere = codul parametrului (placeholder)',
        'severity': 'high',
        'samples': rows,
    }

    # 4. Lipsă explicație tehnică
    cursor.execute(f'''
        SELECT id, parametru, descriere_scurta, familie FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} (explicatie IS NULL OR TRIM(explicatie)='')
        LIMIT 30
    ''', params)
    rows = [dict(r) for r in cursor.fetchall()]
    cursor.execute(f'''
        SELECT COUNT(*) FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} (explicatie IS NULL OR TRIM(explicatie)='')
    ''', params)
    issues['missing_explicatie'] = {
        'count': cursor.fetchone()[0],
        'label': 'Fără explicație tehnică detaliată',
        'severity': 'medium',
        'samples': rows,
    }

    # 5. Lipsă pagină manual
    cursor.execute(f'''
        SELECT id, parametru, descriere_scurta, familie FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} pagina IS NULL
        LIMIT 30
    ''', params)
    rows = [dict(r) for r in cursor.fetchall()]
    cursor.execute(f'''
        SELECT COUNT(*) FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} pagina IS NULL
    ''', params)
    issues['missing_pagina'] = {
        'count': cursor.fetchone()[0],
        'label': 'Fără referință la pagina manualului',
        'severity': 'medium',
        'samples': rows,
    }

    # 6. Cod parametru suspect (doar numere sau prea scurt)
    cursor.execute(f'''
        SELECT id, parametru, descriere_scurta, familie FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} (LENGTH(TRIM(parametru)) < 3 OR parametru NOT LIKE '%.%')
        LIMIT 30
    ''', params)
    rows = [dict(r) for r in cursor.fetchall()]
    cursor.execute(f'''
        SELECT COUNT(*) FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} (LENGTH(TRIM(parametru)) < 3 OR parametru NOT LIKE '%.%')
    ''', params)
    issues['suspect_code'] = {
        'count': cursor.fetchone()[0],
        'label': 'Cod parametru suspect (prea scurt sau format neașteptat)',
        'severity': 'low',
        'samples': rows,
    }

    # 7. Descrieri identice la 5+ parametri (probabil placeholder)
    cursor.execute(f'''
        SELECT descriere_scurta, COUNT(*) as cnt
        FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')} TRIM(descriere_scurta) != ''
        GROUP BY descriere_scurta HAVING cnt >= 5
        ORDER BY cnt DESC LIMIT 20
    ''', params)
    dupes = [dict(r) for r in cursor.fetchall()]
    issues['duplicate_descrieri'] = {
        'count': sum(d['cnt'] for d in dupes),
        'label': 'Descrieri duplicate (5+ parametri cu același text)',
        'severity': 'medium',
        'samples': dupes,
    }

    # 8. Lipsă unitate când titlul sugerează una (SQLite n-are REGEXP, folosim LIKE)
    cursor.execute(f'''
        SELECT id, parametru, descriere_scurta, familie FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')}
        (unitate IS NULL OR TRIM(unitate)='')
        AND (descriere_scurta LIKE '%Hz%' OR descriere_scurta LIKE '%kW%' OR descriere_scurta LIKE '%rpm%' OR descriere_scurta LIKE '%°C%' OR descriere_scurta LIKE '%percent%')
        LIMIT 30
    ''', params)
    rows = [dict(r) for r in cursor.fetchall()]
    cursor.execute(f'''
        SELECT COUNT(*) FROM {safe_table("parametri_master")}
        {where + (' AND ' if where else ' WHERE ')}
        (unitate IS NULL OR TRIM(unitate)='')
        AND (descriere_scurta LIKE '%Hz%' OR descriere_scurta LIKE '%kW%' OR descriere_scurta LIKE '%rpm%' OR descriere_scurta LIKE '%°C%' OR descriere_scurta LIKE '%percent%')
    ''', params)
    issues['missing_unitate_suggested'] = {
        'count': cursor.fetchone()[0],
        'label': 'Lipsă unitate când descrierea sugerează una',
        'severity': 'low',
        'samples': rows,
    }

    conn.close()

    # Health score: cât din total NU are probleme grave (missing desc + descriere=code + missing explicatie)
    severe_count = (
        issues['missing_descriere_scurta']['count'] +
        issues['descriere_equals_code']['count']
    )
    health_pct = round(100 * (total - severe_count) / total, 1) if total else 0

    return jsonify({
        'total': total,
        'familie_filter': familie_filter or None,
        'health_pct': health_pct,
        'per_familie': per_familie,
        'issues': issues,
    })


@app.route('/api/parametri/<int:param_id>', methods=['GET'])
@login_required
def get_parametru_detail(param_id):
    """Returnează detaliul complet al unui parametru.

    Include:
      - toate coloanele din parametri_master
      - `influentat_de`: lista params din aceeași familie care îl referențiază
        în câmpul lor `influenteaza` (computed reverse, nu coloană separată).
    """
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM parametri_master WHERE id = ?', (param_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    result = dict(row)

    # Reverse lookup: who references this parameter's code?
    code = result.get('parametru')
    familie = result.get('familie')
    if code and familie:
        # influenteaza is stored as comma-separated codes (e.g. "30.12, 21.13").
        # Match the code as a standalone token to avoid p100 catching p1000.
        # SQLite REGEXP isn't built-in by default, so use LIKE with delimiters.
        like1 = f'%{code}%'
        cursor.execute(
            'SELECT id, parametru, descriere_scurta, influenteaza FROM parametri_master '
            'WHERE familie = ? AND id != ? AND influenteaza IS NOT NULL AND influenteaza LIKE ?',
            (familie, param_id, like1)
        )
        candidates = cursor.fetchall()
        # Refine: code appears as a comma/space-delimited token (avoid p100 catching p1000)
        import re as _re
        token_pat = _re.compile(r'(?:^|[,\s])' + _re.escape(code) + r'(?:$|[,\s])')
        influentat_de = [
            {'id': c['id'], 'parametru': c['parametru'], 'descriere_scurta': c['descriere_scurta']}
            for c in candidates
            if c['influenteaza'] and token_pat.search(c['influenteaza'])
        ]
        result['influentat_de'] = influentat_de
    else:
        result['influentat_de'] = []

    conn.close()
    return jsonify(result)

# ============ MANUALS API ============

MANUALS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'manuals')

@app.route('/api/manuals', methods=['GET'])
@login_required
def get_manuals():
    """List available PDF manuals with metadata"""
    manuals = []
    if os.path.isdir(MANUALS_DIR):
        for fname in sorted(os.listdir(MANUALS_DIR)):
            if fname.endswith('.pdf'):
                fpath = os.path.join(MANUALS_DIR, fname)
                size_kb = round(os.path.getsize(fpath) / 1024, 1)
                # Derive family from filename
                name_display = fname.replace('.pdf', '').replace('_', ' ')
                manuals.append({
                    'filename': fname,
                    'name': name_display,
                    'size_kb': size_kb,
                    'url': f'/manuals/{fname}'
                })
    return jsonify({'manuals': manuals})

@app.route('/manuals/<path:filename>', methods=['GET'])
def serve_manual(filename):
    """Serve a PDF manual file (no auth — public technical docs)"""
    safe_name = os.path.basename(filename)
    fpath = os.path.join(MANUALS_DIR, safe_name)
    if not os.path.isfile(fpath):
        return 'Manual not found', 404
    return send_file(fpath)

# ============ APP SETTINGS (key/value) ============

def get_app_setting(key, default=None):
    """Read a value from the app_settings key/value table."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT value FROM app_settings WHERE key = ?', (key,))
    row = cursor.fetchone()
    conn.close()
    return row['value'] if row else default


def set_app_setting(key, value):
    """Upsert a value into the app_settings key/value table."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute(
        'INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?) '
        'ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = excluded.updated_at',
        (key, value, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()


# ============ OBSIDIAN VAULT INTEGRATION (read-only) ============
# Ion keeps study notes in an Obsidian vault synced to the laptop-server via
# InSync (Google Drive). The app reads .md files directly — never writes — so
# there is no risk of sync conflicts. The vault path is configured at runtime
# from the Administrativ tab and stored in app_settings.

OBSIDIAN_SETTING_KEY = 'obsidian_vault_path'
OBSIDIAN_FOLDERS_KEY = 'obsidian_folders'
_obsidian_cache = {'path': None, 'sig': None, 'notes': None}
_OBSIDIAN_SKIP_DIRS = {'.obsidian', '.trash', '.git', '.smart-env', 'node_modules'}


def _obsidian_vault():
    """Return the configured vault path, or None if unset/missing."""
    path = (get_app_setting(OBSIDIAN_SETTING_KEY) or '').strip()
    if not path or not os.path.isdir(path):
        return None
    return path


def _obsidian_allowed_folders():
    """Comma-separated top-level folder tokens to include. Empty list => all."""
    raw = (get_app_setting(OBSIDIAN_FOLDERS_KEY) or '').strip()
    return [t.strip() for t in raw.split(',') if t.strip()]


def _obsidian_top_allowed(top, allowed):
    """True if a top-level folder name matches one of the allowed tokens.
    Token '10' matches '10', '10_Library', '10 Library', '10-Library'."""
    if not allowed:
        return True
    if not top:
        return False  # root-level note — excluded once a filter is active
    for tok in allowed:
        if top == tok or top.startswith(tok + '_') \
                or top.startswith(tok + ' ') or top.startswith(tok + '-'):
            return True
    return False


def _obsidian_safe_path(vault, rel):
    """Resolve a vault-relative path safely (defends against path traversal).
    Returns an absolute path only if it stays inside the vault AND is a .md file."""
    if not vault or not rel:
        return None
    rel = str(rel).replace('\\', '/').lstrip('/')
    candidate = os.path.realpath(os.path.join(vault, rel))
    vault_real = os.path.realpath(vault)
    inside = candidate == vault_real or candidate.startswith(vault_real + os.sep)
    if not inside or not candidate.lower().endswith('.md'):
        return None
    return candidate


def _obsidian_walk(vault):
    """Walk the vault, yielding (relpath, abspath, mtime, size) for every .md
    file, skipping Obsidian/system folders and honouring the top-level folder
    filter (Ion only cares about folders 10/30/99)."""
    vault_real = os.path.realpath(vault)
    allowed = _obsidian_allowed_folders()
    out = []
    for root, dirs, files in os.walk(vault_real):
        dirs[:] = [d for d in dirs if d not in _OBSIDIAN_SKIP_DIRS and not d.startswith('.')]
        # At the vault root, prune to allowed top-level folders so we don't even
        # descend into the excluded ones.
        if allowed and os.path.realpath(root) == vault_real:
            dirs[:] = [d for d in dirs if _obsidian_top_allowed(d, allowed)]
        for fname in files:
            if not fname.lower().endswith('.md'):
                continue
            abspath = os.path.join(root, fname)
            try:
                st = os.stat(abspath)
            except OSError:
                continue
            rel = os.path.relpath(abspath, vault_real).replace('\\', '/')
            top = rel.split('/', 1)[0] if '/' in rel else ''
            if not _obsidian_top_allowed(top, allowed):
                continue
            out.append((rel, abspath, st.st_mtime, st.st_size))
    return out


def _obsidian_index(vault):
    """Return a cached list of note dicts {path, title, folder, content, mtime,
    size}. Rebuilds only when the set of files or their mtimes changed."""
    files = _obsidian_walk(vault)
    sig = tuple(sorted((rel, mtime) for rel, _a, mtime, _s in files))
    if (_obsidian_cache['path'] == vault and _obsidian_cache['sig'] == sig
            and _obsidian_cache['notes'] is not None):
        return _obsidian_cache['notes']

    notes = []
    for rel, abspath, mtime, size in files:
        try:
            with open(abspath, 'r', encoding='utf-8', errors='ignore') as fh:
                content = fh.read()
        except OSError:
            content = ''
        folder = os.path.dirname(rel)
        title = os.path.basename(rel)[:-3]  # strip .md
        notes.append({
            'path': rel, 'title': title, 'folder': folder,
            'content': content, 'mtime': mtime, 'size': size,
        })
    notes.sort(key=lambda n: n['path'].lower())
    _obsidian_cache.update({'path': vault, 'sig': sig, 'notes': notes})
    return notes


def _obsidian_config_dict():
    """Current Obsidian config as a JSON-serialisable dict."""
    raw = (get_app_setting(OBSIDIAN_SETTING_KEY) or '').strip()
    folders = (get_app_setting(OBSIDIAN_FOLDERS_KEY) or '').strip()
    vault = _obsidian_vault()
    note_count = len(_obsidian_index(vault)) if vault else 0
    return {
        'configured': bool(raw),
        'vault_path': raw,
        'folders': folders,
        'valid': vault is not None,
        'note_count': note_count,
    }


@app.route('/api/obsidian/config', methods=['GET'])
@login_required
def obsidian_config_get():
    return jsonify(_obsidian_config_dict())


@app.route('/api/obsidian/config', methods=['PUT'])
@login_required
def obsidian_config_set():
    data = request.get_json(silent=True) or {}
    if 'vault_path' in data:
        path = (data.get('vault_path') or '').strip()
        if path and not os.path.isdir(path):
            return jsonify({'error': 'Calea nu există sau nu este un folder', 'valid': False}), 400
        set_app_setting(OBSIDIAN_SETTING_KEY, path)
    if 'folders' in data:
        set_app_setting(OBSIDIAN_FOLDERS_KEY, (data.get('folders') or '').strip())
    _obsidian_cache.update({'path': None, 'sig': None, 'notes': None})  # invalidate
    return jsonify(_obsidian_config_dict())


@app.route('/api/obsidian/notes', methods=['GET'])
@login_required
def obsidian_notes_list():
    vault = _obsidian_vault()
    if not vault:
        return jsonify({'error': 'Vault Obsidian neconfigurat', 'notes': []}), 200
    notes = [
        {'path': n['path'], 'title': n['title'], 'folder': n['folder'],
         'mtime': n['mtime'], 'size': n['size']}
        for n in _obsidian_index(vault)
    ]
    return jsonify({'notes': notes})


@app.route('/api/obsidian/note', methods=['GET'])
@login_required
def obsidian_note_get():
    vault = _obsidian_vault()
    if not vault:
        return jsonify({'error': 'Vault Obsidian neconfigurat'}), 400
    rel = request.args.get('path', '')
    abspath = _obsidian_safe_path(vault, rel)
    if not abspath or not os.path.isfile(abspath):
        return jsonify({'error': 'Nota nu a fost găsită'}), 404
    try:
        with open(abspath, 'r', encoding='utf-8', errors='ignore') as fh:
            content = fh.read()
    except OSError as e:
        return jsonify({'error': f'Eroare la citire: {e}'}), 500
    rel_norm = str(rel).replace('\\', '/').lstrip('/')
    return jsonify({
        'path': rel_norm,
        'title': os.path.basename(rel_norm)[:-3],
        'content': content,
    })


@app.route('/api/obsidian/search', methods=['GET'])
@login_required
def obsidian_search():
    vault = _obsidian_vault()
    if not vault:
        return jsonify({'error': 'Vault Obsidian neconfigurat', 'results': []}), 200
    query = (request.args.get('q') or '').strip()
    if not query:
        return jsonify({'results': []})
    q_low = query.lower()
    results = []
    for n in _obsidian_index(vault):
        content_low = n['content'].lower()
        title_low = n['title'].lower()
        in_title = q_low in title_low
        hits = content_low.count(q_low)
        if not in_title and hits == 0:
            continue
        # Snippet: context window around the first content match.
        snippet = ''
        idx = content_low.find(q_low)
        if idx >= 0:
            start = max(0, idx - 60)
            end = min(len(n['content']), idx + len(query) + 90)
            snippet = ('…' if start > 0 else '') + n['content'][start:end].strip() \
                      + ('…' if end < len(n['content']) else '')
        score = hits + (50 if in_title else 0)
        results.append({
            'path': n['path'], 'title': n['title'], 'folder': n['folder'],
            'snippet': snippet, 'hits': hits, 'score': score,
        })
    results.sort(key=lambda r: r['score'], reverse=True)
    return jsonify({'results': results[:50], 'query': query})


@app.route('/api/obsidian/mentions', methods=['GET'])
@login_required
def obsidian_mentions():
    """Notes that mention a given term — used to auto-link notes to a parameter
    or project from their detail view. Token-boundary match to avoid noise."""
    vault = _obsidian_vault()
    if not vault:
        return jsonify({'mentions': []}), 200
    term = (request.args.get('term') or '').strip()
    if not term or len(term) < 2:
        return jsonify({'mentions': []})
    import re as _re
    pat = _re.compile(r'(?<![\w.])' + _re.escape(term) + r'(?![\w.])', _re.IGNORECASE)
    mentions = []
    for n in _obsidian_index(vault):
        if pat.search(n['content']) or pat.search(n['title']):
            mentions.append({'path': n['path'], 'title': n['title'], 'folder': n['folder']})
    return jsonify({'mentions': mentions[:25], 'term': term})


# ============ AI ASSISTANT (Hermes, in-app) ============
# A chat assistant backed by the MiniMax gateway. It can search the app's data
# and perform actions through tool/function calling. Gateway credentials live
# in .assistant_config on the server (written by Hermes, never committed).

ASSISTANT_CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.assistant_config')

ASSISTANT_SYSTEM_PROMPT = """Ești Hermes — asistentul personal, specializat, integrat în
PIF Dashboard. Lucrezi exclusiv pentru Ion și ești singurul lui asistent în aplicație.

CINE E ION
- Inginer de punere în funcțiune (commissioning) pentru convertoare de frecvență
  industriale. Lucrează pe teren: PIF (puneri în funcțiune) și Service.
- Single user al aplicației — tot ce e aici e al lui.

CUM COMUNICI CU EL
- Mereu în română. Direct, concis, la obiect. Fără emoji, fără introduceri
  politicoase lungi, fără să repeți întrebarea lui.
- Înțelege-l din jumătate de cuvânt: folosește contextul conversației, MEMORIA ta
  și datele reale din aplicație ca să deduci ce vrea. Nu pune întrebări inutile.
- Dacă cererea e cu adevărat ambiguă (ex: nu se știe la ce proiect), întreabă
  scurt. Altfel acționează pe interpretarea cea mai rezonabilă și spune ce ai
  presupus.
- Ești colaborativ și proactiv: dacă vezi ceva util (task scadent, inconsistență,
  un pas care lipsește), spune-i.
- După fiecare acțiune, confirmă scurt și concret ce ai făcut.
- Confirmă în chat ÎNAINTE de ștergeri sau acțiuni ireversibile.

DOMENIUL
- PIF = Punere În Funcțiune. Service = mentenanță/intervenții.
- Familii de convertoare: ABB ACS580 / ACS880, Siemens SINAMICS G120 /
  SINAMICS_G130_G150 / SINAMICS_S120_S150, Danfoss VLT FC302, Lenze i550 / i950.

MODELUL APLICAȚIEI
- Proiecte (tip PIF sau Service), fiecare cu: taskuri, checklist, jurnal de
  activitate, cronometru, echipamente. Status proiect: in_lucru, in_asteptare,
  blocat, finalizat.
- Taskuri: status to_do/in_lucru/done, prioritate normal/urgent/minor, scadență,
  descriere, subtaskuri, recurență (zilnic/saptamanal/lunar).
- Taskuri zilnice (globale) — independente de proiect.
- Bază de ~14.700 parametri de convertoare, cu explicații și influențe.
- Notițele tehnice Obsidian ale lui Ion.
- Clienți și echipamente.

REGULI DE LUCRU
- Caută mereu informația reală cu uneltele — nu inventa date. Dacă o unealtă nu
  găsește nimic, spune clar.
- Pentru parametri: search_parametri / get_parametru.
- Pentru notițele lui Ion: search_obsidian / read_obsidian_note.
- Ai acces complet: poți crea, modifica și șterge proiecte, taskuri, taskuri
  zilnice, subtaskuri, checklist, jurnal, echipamente, clienți.

MEMORIA TA
- Ai o memorie persistentă, vizibilă mai jos. Conține preferințele lui Ion,
  context recurent, decizii, clienți/echipamente uzuale.
- Când Ion îți spune o preferință sau un detaliu de context care contează pe
  termen lung, salvează-l cu save_memory. Nu salva lucruri triviale sau ușor de
  regăsit din datele aplicației.
- Folosește memoria activ ca să-l înțelegi mai bine de fiecare dată."""


def _load_assistant_memory():
    """Return the assistant's persistent memory entries (oldest first)."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, continut FROM assistant_memory ORDER BY created_at ASC')
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return rows


def _build_assistant_system_prompt():
    """System prompt + the current memory section."""
    mem = _load_assistant_memory()
    if not mem:
        return ASSISTANT_SYSTEM_PROMPT + '\n\nMEMORIE: (goală încă)'
    lines = '\n'.join(f"- [{m['id'][:8]}] {m['continut']}" for m in mem)
    return ASSISTANT_SYSTEM_PROMPT + '\n\nMEMORIE (lucruri reținute despre Ion și context):\n' + lines

# Tool schemas (OpenAI-compatible function calling).
ASSISTANT_TOOLS = [
    {"type": "function", "function": {
        "name": "search_parametri",
        "description": "Caută parametri de convertor în baza de date după text (cod sau descriere), opțional filtrat pe familie.",
        "parameters": {"type": "object", "properties": {
            "query": {"type": "string", "description": "cod parametru (ex 30.12) sau cuvinte din descriere"},
            "familie": {"type": "string", "description": "opțional: ACS580, ACS880, SINAMICS_G120, SINAMICS_G130_G150, SINAMICS_S120_S150, Danfoss_VLT_FC302, Lenze_i550, Lenze_i950"}
        }, "required": ["query"]}
    }},
    {"type": "function", "function": {
        "name": "get_parametru",
        "description": "Returnează detaliul complet al unui parametru (descriere, acces, default, min/max, explicație, influențe).",
        "parameters": {"type": "object", "properties": {
            "familie": {"type": "string"},
            "cod": {"type": "string", "description": "codul exact al parametrului"}
        }, "required": ["familie", "cod"]}
    }},
    {"type": "function", "function": {
        "name": "search_obsidian",
        "description": "Caută în notițele tehnice Obsidian ale utilizatorului (full-text).",
        "parameters": {"type": "object", "properties": {
            "query": {"type": "string"}
        }, "required": ["query"]}
    }},
    {"type": "function", "function": {
        "name": "read_obsidian_note",
        "description": "Citește conținutul complet al unei notițe Obsidian după calea ei relativă.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string", "description": "calea relativă a notiței (din search_obsidian)"}
        }, "required": ["path"]}
    }},
    {"type": "function", "function": {
        "name": "list_proiecte",
        "description": "Listează proiectele, opțional filtrate pe status.",
        "parameters": {"type": "object", "properties": {
            "status": {"type": "string", "description": "opțional: in_lucru, in_asteptare, blocat, finalizat"}
        }}
    }},
    {"type": "function", "function": {
        "name": "get_proiect",
        "description": "Detaliul unui proiect după nume (parțial) sau id: date proiect, taskuri, checklist, jurnal.",
        "parameters": {"type": "object", "properties": {
            "nume": {"type": "string", "description": "nume parțial sau id de proiect"}
        }, "required": ["nume"]}
    }},
    {"type": "function", "function": {
        "name": "create_proiect",
        "description": "Creează un proiect nou.",
        "parameters": {"type": "object", "properties": {
            "nume": {"type": "string"},
            "tip": {"type": "string", "description": "PIF sau Service"},
            "client": {"type": "string"},
            "producator": {"type": "string", "description": "ex ABB, Siemens, Danfoss, Lenze"},
            "locatie": {"type": "string"}
        }, "required": ["nume"]}
    }},
    {"type": "function", "function": {
        "name": "create_task",
        "description": "Adaugă un task într-un proiect.",
        "parameters": {"type": "object", "properties": {
            "proiect": {"type": "string", "description": "nume parțial sau id de proiect"},
            "titlu": {"type": "string"},
            "prioritate": {"type": "string", "description": "normal, urgent, minor"},
            "scadenta": {"type": "string", "description": "data YYYY-MM-DD, opțional"},
            "descriere": {"type": "string"},
            "recurenta": {"type": "string", "description": "opțional: zilnic, saptamanal, lunar"}
        }, "required": ["proiect", "titlu"]}
    }},
    {"type": "function", "function": {
        "name": "add_checklist_item",
        "description": "Adaugă un punct în checklist-ul unui proiect.",
        "parameters": {"type": "object", "properties": {
            "proiect": {"type": "string"},
            "titlu": {"type": "string"}
        }, "required": ["proiect", "titlu"]}
    }},
    {"type": "function", "function": {
        "name": "add_jurnal",
        "description": "Adaugă o intrare în jurnalul unui proiect.",
        "parameters": {"type": "object", "properties": {
            "proiect": {"type": "string"},
            "continut": {"type": "string"}
        }, "required": ["proiect", "continut"]}
    }},
    {"type": "function", "function": {
        "name": "update_task",
        "description": "Modifică un task dintr-un proiect (găsit după titlu parțial). Trimite doar câmpurile de schimbat.",
        "parameters": {"type": "object", "properties": {
            "proiect": {"type": "string"},
            "task": {"type": "string", "description": "titlu parțial al taskului de modificat"},
            "status": {"type": "string", "description": "to_do, in_lucru, done"},
            "prioritate": {"type": "string", "description": "normal, urgent, minor"},
            "scadenta": {"type": "string", "description": "data YYYY-MM-DD"},
            "descriere": {"type": "string"},
            "recurenta": {"type": "string", "description": "zilnic, saptamanal, lunar sau gol"}
        }, "required": ["proiect", "task"]}
    }},
    {"type": "function", "function": {
        "name": "delete_task",
        "description": "Șterge un task dintr-un proiect (găsit după titlu parțial). Confirmă cu Ion înainte.",
        "parameters": {"type": "object", "properties": {
            "proiect": {"type": "string"},
            "task": {"type": "string", "description": "titlu parțial al taskului"}
        }, "required": ["proiect", "task"]}
    }},
    {"type": "function", "function": {
        "name": "add_subtask",
        "description": "Adaugă un subtask la un task dintr-un proiect (taskul găsit după titlu parțial).",
        "parameters": {"type": "object", "properties": {
            "proiect": {"type": "string"},
            "task": {"type": "string", "description": "titlu parțial al taskului părinte"},
            "titlu": {"type": "string", "description": "titlul subtaskului"}
        }, "required": ["proiect", "task", "titlu"]}
    }},
    {"type": "function", "function": {
        "name": "update_proiect",
        "description": "Modifică orice câmp al unui proiect. Trimite doar câmpurile pe care le schimbi. Orice modificare (inclusiv redenumirea prin nume_nou) pastreaza tot continutul proiectului: taskuri, jurnal, checklist, timer, echipamente.",
        "parameters": {"type": "object", "properties": {
            "nume": {"type": "string", "description": "nume parțial sau id al proiectului de modificat (cheie de cautare)"},
            "nume_nou": {"type": "string", "description": "noul nume al proiectului — pentru redenumire"},
            "tip": {"type": "string", "description": "PIF sau Service"},
            "status": {"type": "string", "description": "in_lucru, in_asteptare, blocat, finalizat"},
            "client": {"type": "string"},
            "locatie": {"type": "string"},
            "producator": {"type": "string", "description": "ABB, Siemens, Danfoss, Lenze, Altul"},
            "echipament_principal": {"type": "string"},
            "cod_proiect": {"type": "string"},
            "pm": {"type": "string", "description": "responsabil proiect"},
            "folder_server": {"type": "string"},
            "data_incepere": {"type": "string", "description": "data de inceput, format YYYY-MM-DD"},
            "deadline": {"type": "string", "description": "termen limita, format YYYY-MM-DD"},
            "nr_comanda": {"type": "string"},
            "nr_contract": {"type": "string"},
            "observatii": {"type": "string"},
            "service_before": {"type": "string", "description": "constatari Service (inainte de interventie)"},
            "service_after": {"type": "string", "description": "actiuni Service (dupa interventie)"}
        }, "required": ["nume"]}
    }},
    {"type": "function", "function": {
        "name": "delete_proiect",
        "description": "Șterge complet un proiect cu tot ce conține. Ireversibil — confirmă cu Ion în chat înainte.",
        "parameters": {"type": "object", "properties": {
            "nume": {"type": "string", "description": "nume parțial sau id de proiect"}
        }, "required": ["nume"]}
    }},
    {"type": "function", "function": {
        "name": "toggle_checklist_item",
        "description": "Bifează sau debifează un punct din checklist-ul unui proiect (găsit după titlu parțial).",
        "parameters": {"type": "object", "properties": {
            "proiect": {"type": "string"},
            "item": {"type": "string", "description": "titlu parțial al punctului din checklist"},
            "completed": {"type": "boolean", "description": "true = bifat, false = debifat"}
        }, "required": ["proiect", "item", "completed"]}
    }},
    {"type": "function", "function": {
        "name": "add_echipament",
        "description": "Adaugă un echipament la un proiect.",
        "parameters": {"type": "object", "properties": {
            "proiect": {"type": "string"},
            "nume": {"type": "string"},
            "producator": {"type": "string"},
            "model": {"type": "string"},
            "serial_number": {"type": "string"}
        }, "required": ["proiect", "nume"]}
    }},
    {"type": "function", "function": {
        "name": "list_clienti",
        "description": "Listează clienții.",
        "parameters": {"type": "object", "properties": {}}
    }},
    {"type": "function", "function": {
        "name": "create_client",
        "description": "Creează un client nou.",
        "parameters": {"type": "object", "properties": {
            "nume": {"type": "string"},
            "telefon": {"type": "string"},
            "email": {"type": "string"},
            "adresa": {"type": "string"},
            "contact_principal": {"type": "string"}
        }, "required": ["nume"]}
    }},
    {"type": "function", "function": {
        "name": "list_global_tasks",
        "description": "Listează taskurile zilnice (globale, independente de proiect).",
        "parameters": {"type": "object", "properties": {}}
    }},
    {"type": "function", "function": {
        "name": "create_global_task",
        "description": "Creează un task zilnic (global, independent de proiect).",
        "parameters": {"type": "object", "properties": {
            "titlu": {"type": "string"},
            "prioritate": {"type": "string", "description": "Normal, Urgent, Minor"},
            "categorie": {"type": "string"},
            "scadenta": {"type": "string", "description": "data YYYY-MM-DD"},
            "descriere": {"type": "string"}
        }, "required": ["titlu"]}
    }},
    {"type": "function", "function": {
        "name": "save_memory",
        "description": "Salvează în memoria ta persistentă un fapt despre Ion sau contextul de lucru, ca să-l ții minte data viitoare.",
        "parameters": {"type": "object", "properties": {
            "continut": {"type": "string", "description": "faptul de reținut, o frază scurtă"}
        }, "required": ["continut"]}
    }},
    {"type": "function", "function": {
        "name": "delete_memory",
        "description": "Șterge o intrare din memoria ta persistentă, după id-ul scurt afișat în secțiunea MEMORIE.",
        "parameters": {"type": "object", "properties": {
            "id": {"type": "string", "description": "id-ul scurt (8 caractere) al intrării de memorie"}
        }, "required": ["id"]}
    }},
    {"type": "function", "function": {
        "name": "lookup_fault_code",
        "description": "Cauta un cod de eroare/avarie/avertizare al unui convertor de frecventa (ABB, Siemens, Danfoss, Lenze) si returneaza cauza si remediul. Foloseste cand Ion intreaba ce inseamna un cod afisat de drive (ex: F30001, 2310, A07910, ALARM 14).",
        "parameters": {"type": "object", "properties": {
            "cod": {"type": "string", "description": "codul de eroare cautat (ex: F30001, 2310, A2B4, 14)"},
            "producator": {"type": "string", "description": "optional: ABB, Siemens, Danfoss sau Lenze"},
            "familie": {"type": "string", "description": "optional: familia (ex: ACS580, SINAMICS_G120)"}
        }, "required": ["cod"]}
    }},
]


def _load_assistant_config():
    """Read MiniMax gateway config from .assistant_config (JSON)."""
    try:
        with open(ASSISTANT_CONFIG_FILE, 'r', encoding='utf-8') as fh:
            cfg = json.load(fh)
        if cfg.get('api_url') and cfg.get('api_key') and cfg.get('model'):
            return cfg
    except (OSError, ValueError):
        pass
    return None


def _normalize_assistant_url(url):
    """MiniMax's Anthropic-compatible endpoint is <base>/anthropic/v1/messages.
    Hermes supplied a slightly-off path (/anthropic/v1/chat/completions) — fix it."""
    url = (url or '').strip().rstrip('/')
    if '/anthropic' in url:
        base = url.split('/anthropic')[0]
        return base + '/anthropic/v1/messages'
    return url


def _anthropic_tools():
    """Convert the OpenAI-style tool schemas to Anthropic's tool format."""
    return [{
        'name': t['function']['name'],
        'description': t['function']['description'],
        'input_schema': t['function']['parameters'],
    } for t in ASSISTANT_TOOLS]


def _minimax_call(cfg, system, messages):
    """One round-trip to the MiniMax Anthropic-compatible Messages endpoint.
    Raises RuntimeError with the response body on an HTTP error."""
    import urllib.request
    import urllib.error
    payload = json.dumps({
        'model': cfg['model'],
        'max_tokens': 2048,
        'system': system,
        'messages': messages,
        'tools': _anthropic_tools(),
    }).encode('utf-8')
    req = urllib.request.Request(
        _normalize_assistant_url(cfg['api_url']),
        data=payload,
        headers={
            'Content-Type': 'application/json',
            'Authorization': 'Bearer ' + cfg['api_key'],
            'x-api-key': cfg['api_key'],
            'anthropic-version': '2023-06-01',
        },
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='ignore')[:400]
        raise RuntimeError(f'API {e.code}: {body}')


def _assistant_find_project(cursor, needle):
    """Resolve a project by id or partial name. Returns the Row or None."""
    needle = (needle or '').strip()
    if not needle:
        return None
    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (needle,))
    row = cursor.fetchone()
    if row:
        return row
    cursor.execute('SELECT * FROM proiecte WHERE nume LIKE ? ORDER BY data_crearii DESC LIMIT 1',
                   (f'%{needle}%',))
    return cursor.fetchone()


def _assistant_exec_tool(name, args):
    """Execute one assistant tool. Returns a JSON-serialisable dict."""
    try:
        if name == 'search_parametri':
            conn = get_db(); cur = conn.cursor()
            q = f"%{(args.get('query') or '').strip()}%"
            sql = ('SELECT id, familie, parametru, descriere_scurta, descriere FROM parametri_master '
                   'WHERE (parametru LIKE ? OR descriere LIKE ?)')
            params = [q, q]
            if args.get('familie'):
                sql += ' AND familie = ?'
                params.append(args['familie'])
            sql += ' ORDER BY familie, parametru LIMIT 25'
            cur.execute(sql, params)
            rows = [dict(r) for r in cur.fetchall()]
            conn.close()
            return {'count': len(rows), 'parametri': rows}

        if name == 'get_parametru':
            conn = get_db(); cur = conn.cursor()
            cur.execute('SELECT * FROM parametri_master WHERE familie = ? AND parametru = ? LIMIT 1',
                        (args.get('familie'), args.get('cod')))
            row = cur.fetchone()
            conn.close()
            if not row:
                return {'error': 'Parametrul nu a fost găsit'}
            d = dict(row)
            d.pop('pdf_extra', None)
            return d

        if name == 'search_obsidian':
            vault = _obsidian_vault()
            if not vault:
                return {'error': 'Vault Obsidian neconfigurat'}
            q = (args.get('query') or '').strip().lower()
            out = []
            for n in _obsidian_index(vault):
                if q in n['title'].lower() or q in n['content'].lower():
                    out.append({'path': n['path'], 'title': n['title']})
                if len(out) >= 20:
                    break
            return {'count': len(out), 'note': out}

        if name == 'read_obsidian_note':
            vault = _obsidian_vault()
            if not vault:
                return {'error': 'Vault Obsidian neconfigurat'}
            abspath = _obsidian_safe_path(vault, args.get('path'))
            if not abspath or not os.path.isfile(abspath):
                return {'error': 'Nota nu a fost găsită'}
            with open(abspath, 'r', encoding='utf-8', errors='ignore') as fh:
                return {'path': args.get('path'), 'content': fh.read()[:8000]}

        if name == 'list_proiecte':
            conn = get_db(); cur = conn.cursor()
            sql = 'SELECT id, nume, tip, client, producator, status FROM proiecte'
            params = []
            if args.get('status'):
                sql += ' WHERE status = ?'
                params.append(args['status'])
            sql += ' ORDER BY data_crearii DESC LIMIT 60'
            cur.execute(sql, params)
            rows = [dict(r) for r in cur.fetchall()]
            conn.close()
            return {'count': len(rows), 'proiecte': rows}

        if name == 'get_proiect':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('nume'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            pid = proj['id']
            cur.execute('SELECT titlu, status, prioritate, data_scadenta FROM tasks WHERE proiect_id = ? ORDER BY ordine', (pid,))
            tasks = [dict(r) for r in cur.fetchall()]
            cur.execute('SELECT titlu, completed FROM checklist_pif WHERE proiect_id = ?', (pid,))
            checklist = [dict(r) for r in cur.fetchall()]
            cur.execute('SELECT data, continut FROM jurnal WHERE proiect_id = ? ORDER BY created_at DESC LIMIT 10', (pid,))
            jurnal = [dict(r) for r in cur.fetchall()]
            cur.execute('SELECT nume, producator, model, serial_number FROM echipamente WHERE proiect_id = ?', (pid,))
            echipamente = [dict(r) for r in cur.fetchall()]
            conn.close()
            return {'proiect': dict(proj), 'taskuri': tasks, 'checklist': checklist,
                    'jurnal': jurnal, 'echipamente': echipamente}

        if name == 'create_proiect':
            conn = get_db(); cur = conn.cursor()
            now = datetime.now().isoformat()
            pid = generate_uuid()
            cur.execute('''
                INSERT INTO proiecte (id, tip, nume, client, producator, locatie, status, data_crearii, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 'in_lucru', ?, ?)
            ''', (pid, args.get('tip', 'PIF'), args.get('nume', ''), args.get('client', ''),
                  args.get('producator', ''), args.get('locatie', ''), now, now))
            conn.commit(); conn.close()
            return {'ok': True, 'id': pid, 'mesaj': f"Proiect creat: {args.get('nume')}"}

        if name == 'create_task':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('proiect'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            now = datetime.now().isoformat()
            tid = generate_uuid()
            cur.execute('SELECT COALESCE(MAX(ordine), 0) FROM tasks WHERE proiect_id = ?', (proj['id'],))
            mo = cur.fetchone()[0]
            cur.execute('''
                INSERT INTO tasks (id, proiect_id, titlu, status, prioritate, data_scadenta,
                                   data_finalizare, ordine, created_at, descriere, recurenta, updated_at)
                VALUES (?, ?, ?, 'to_do', ?, ?, '', ?, ?, ?, ?, ?)
            ''', (tid, proj['id'], args.get('titlu', ''), args.get('prioritate', 'normal'),
                  args.get('scadenta', ''), mo + 1, now, args.get('descriere', ''),
                  args.get('recurenta', ''), now))
            conn.commit(); conn.close()
            return {'ok': True, 'id': tid, 'mesaj': f"Task adăugat în {proj['nume']}: {args.get('titlu')}"}

        if name == 'add_checklist_item':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('proiect'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            now = datetime.now().isoformat()
            cid = generate_uuid()
            cur.execute('''
                INSERT INTO checklist_pif (id, proiect_id, titlu, completed, note, ordine, categorie_id)
                VALUES (?, ?, ?, 0, '', 0, NULL)
            ''', (cid, proj['id'], args.get('titlu', '')))
            conn.commit(); conn.close()
            return {'ok': True, 'id': cid, 'mesaj': f"Checklist item adăugat: {args.get('titlu')}"}

        if name == 'add_jurnal':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('proiect'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            now = datetime.now().isoformat()
            jid = generate_uuid()
            cur.execute('INSERT INTO jurnal (id, proiect_id, data, continut, created_at) VALUES (?, ?, ?, ?, ?)',
                        (jid, proj['id'], now[:10], args.get('continut', ''), now))
            conn.commit(); conn.close()
            return {'ok': True, 'id': jid, 'mesaj': 'Intrare adăugată în jurnal'}

        if name == 'update_task':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('proiect'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            cur.execute('SELECT id, titlu FROM tasks WHERE proiect_id = ? AND titlu LIKE ? LIMIT 1',
                        (proj['id'], f"%{args.get('task') or ''}%"))
            task = cur.fetchone()
            if not task:
                conn.close()
                return {'error': 'Taskul nu a fost găsit'}
            cur.execute('''UPDATE tasks SET
                status = COALESCE(?, status), prioritate = COALESCE(?, prioritate),
                data_scadenta = COALESCE(?, data_scadenta), descriere = COALESCE(?, descriere),
                recurenta = COALESCE(?, recurenta), updated_at = ? WHERE id = ?''',
                (args.get('status'), args.get('prioritate'), args.get('scadenta'),
                 args.get('descriere'), args.get('recurenta'), datetime.now().isoformat(), task['id']))
            conn.commit(); conn.close()
            return {'ok': True, 'mesaj': f"Task actualizat: {task['titlu']}"}

        if name == 'delete_task':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('proiect'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            cur.execute('SELECT id, titlu FROM tasks WHERE proiect_id = ? AND titlu LIKE ? LIMIT 1',
                        (proj['id'], f"%{args.get('task') or ''}%"))
            task = cur.fetchone()
            if not task:
                conn.close()
                return {'error': 'Taskul nu a fost găsit'}
            cur.execute('DELETE FROM task_subtasks WHERE task_id = ?', (task['id'],))
            cur.execute('DELETE FROM tasks WHERE id = ?', (task['id'],))
            conn.commit(); conn.close()
            return {'ok': True, 'mesaj': f"Task șters: {task['titlu']}"}

        if name == 'add_subtask':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('proiect'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            cur.execute('SELECT id, titlu FROM tasks WHERE proiect_id = ? AND titlu LIKE ? LIMIT 1',
                        (proj['id'], f"%{args.get('task') or ''}%"))
            task = cur.fetchone()
            if not task:
                conn.close()
                return {'error': 'Taskul nu a fost găsit'}
            now = datetime.now().isoformat()
            cur.execute('SELECT COALESCE(MAX(ordine), -1) + 1 FROM task_subtasks WHERE task_id = ?', (task['id'],))
            mo = cur.fetchone()[0]
            cur.execute('INSERT INTO task_subtasks (id, task_id, titlu, done, ordine, created_at) VALUES (?, ?, ?, 0, ?, ?)',
                        (generate_uuid(), task['id'], args.get('titlu', ''), mo, now))
            conn.commit(); conn.close()
            return {'ok': True, 'mesaj': f"Subtask adăugat la '{task['titlu']}': {args.get('titlu')}"}

        if name == 'update_proiect':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('nume'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            nume_nou = (args.get('nume_nou') or '').strip() or None
            cur.execute('''UPDATE proiecte SET
                nume = COALESCE(?, nume), tip = COALESCE(?, tip),
                status = COALESCE(?, status), client = COALESCE(?, client),
                locatie = COALESCE(?, locatie), producator = COALESCE(?, producator),
                echipament_principal = COALESCE(?, echipament_principal),
                cod_proiect = COALESCE(?, cod_proiect), pm = COALESCE(?, pm),
                folder_server = COALESCE(?, folder_server),
                data_incepere = COALESCE(?, data_incepere), deadline = COALESCE(?, deadline),
                nr_comanda = COALESCE(?, nr_comanda), nr_contract = COALESCE(?, nr_contract),
                observatii = COALESCE(?, observatii),
                service_before = COALESCE(?, service_before),
                service_after = COALESCE(?, service_after),
                updated_at = ? WHERE id = ?''',
                (nume_nou, args.get('tip'), args.get('status'), args.get('client'),
                 args.get('locatie'), args.get('producator'), args.get('echipament_principal'),
                 args.get('cod_proiect'), args.get('pm'), args.get('folder_server'),
                 args.get('data_incepere'), args.get('deadline'), args.get('nr_comanda'),
                 args.get('nr_contract'), args.get('observatii'), args.get('service_before'),
                 args.get('service_after'), datetime.now().isoformat(), proj['id']))
            conn.commit(); conn.close()
            return {'ok': True, 'mesaj': f"Proiect actualizat: {nume_nou or proj['nume']}"}

        if name == 'delete_proiect':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('nume'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            pid = proj['id']
            cur.execute('SELECT id FROM tasks WHERE proiect_id = ?', (pid,))
            for t in cur.fetchall():
                cur.execute('DELETE FROM task_subtasks WHERE task_id = ?', (t['id'],))
            for tbl in ('tasks', 'checklist_pif', 'checklist_categorii', 'jurnal',
                        'timer_sessions', 'echipamente', 'atasamente'):
                cur.execute(f'DELETE FROM {safe_table(tbl)} WHERE proiect_id = ?', (pid,))
            cur.execute('DELETE FROM proiecte WHERE id = ?', (pid,))
            conn.commit(); conn.close()
            return {'ok': True, 'mesaj': f"Proiect șters complet: {proj['nume']}"}

        if name == 'toggle_checklist_item':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('proiect'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            cur.execute('SELECT id, titlu FROM checklist_pif WHERE proiect_id = ? AND titlu LIKE ? LIMIT 1',
                        (proj['id'], f"%{args.get('item') or ''}%"))
            item = cur.fetchone()
            if not item:
                conn.close()
                return {'error': 'Punctul din checklist nu a fost găsit'}
            done = 1 if args.get('completed') else 0
            cur.execute('UPDATE checklist_pif SET completed = ? WHERE id = ?', (done, item['id']))
            conn.commit(); conn.close()
            return {'ok': True, 'mesaj': f"Checklist '{item['titlu']}': {'bifat' if done else 'debifat'}"}

        if name == 'add_echipament':
            conn = get_db(); cur = conn.cursor()
            proj = _assistant_find_project(cur, args.get('proiect'))
            if not proj:
                conn.close()
                return {'error': 'Proiectul nu a fost găsit'}
            now = datetime.now().isoformat()
            eid = generate_uuid()
            cur.execute('''INSERT INTO echipamente (id, proiect_id, nume, producator, model, serial_number, created_at, updated_at)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                        (eid, proj['id'], args.get('nume', ''), args.get('producator', ''),
                         args.get('model', ''), args.get('serial_number', ''), now, now))
            conn.commit(); conn.close()
            return {'ok': True, 'id': eid, 'mesaj': f"Echipament adăugat la {proj['nume']}: {args.get('nume')}"}

        if name == 'list_clienti':
            conn = get_db(); cur = conn.cursor()
            cur.execute('SELECT id, nume, telefon, email, contact_principal FROM clienti ORDER BY nume')
            rows = [dict(r) for r in cur.fetchall()]
            conn.close()
            return {'count': len(rows), 'clienti': rows}

        if name == 'create_client':
            conn = get_db(); cur = conn.cursor()
            cid = generate_uuid()
            cur.execute('''INSERT INTO clienti (id, nume, adresa, telefon, email, contact_principal, note, created_at)
                           VALUES (?, ?, ?, ?, ?, ?, '', ?)''',
                        (cid, args.get('nume', ''), args.get('adresa', ''), args.get('telefon', ''),
                         args.get('email', ''), args.get('contact_principal', ''), datetime.now().isoformat()))
            conn.commit(); conn.close()
            return {'ok': True, 'id': cid, 'mesaj': f"Client creat: {args.get('nume')}"}

        if name == 'list_global_tasks':
            conn = get_db(); cur = conn.cursor()
            cur.execute("SELECT id, titlu, status, prioritate, categorie, data_scadenta "
                        "FROM global_tasks WHERE status != 'done' ORDER BY created_at DESC LIMIT 60")
            rows = [dict(r) for r in cur.fetchall()]
            conn.close()
            return {'count': len(rows), 'taskuri_zilnice': rows}

        if name == 'create_global_task':
            conn = get_db(); cur = conn.cursor()
            now = datetime.now().isoformat()
            gid = generate_uuid()
            cur.execute('''INSERT INTO global_tasks
                (id, titlu, descriere, prioritate, status, categorie, data_scadenta, data_finalizare, created_at, updated_at)
                VALUES (?, ?, ?, ?, 'to_do', ?, ?, '', ?, ?)''',
                (gid, args.get('titlu', ''), args.get('descriere', ''), args.get('prioritate', 'Normal'),
                 args.get('categorie', 'General'), args.get('scadenta', ''), now, now))
            conn.commit(); conn.close()
            return {'ok': True, 'id': gid, 'mesaj': f"Task zilnic creat: {args.get('titlu')}"}

        if name == 'save_memory':
            continut = (args.get('continut') or '').strip()
            if not continut:
                return {'error': 'Conținut gol'}
            conn = get_db(); cur = conn.cursor()
            cur.execute('INSERT INTO assistant_memory (id, continut, created_at) VALUES (?, ?, ?)',
                        (generate_uuid(), continut, datetime.now().isoformat()))
            conn.commit(); conn.close()
            return {'ok': True, 'mesaj': 'Reținut.'}

        if name == 'delete_memory':
            short = (args.get('id') or '').strip()
            if not short:
                return {'error': 'id gol'}
            conn = get_db(); cur = conn.cursor()
            cur.execute('DELETE FROM assistant_memory WHERE id LIKE ?', (short + '%',))
            n = cur.rowcount
            conn.commit(); conn.close()
            return {'ok': n > 0, 'mesaj': 'Șters din memorie.' if n else 'Nu am găsit intrarea în memorie.'}

        if name == 'lookup_fault_code':
            cod = (args.get('cod') or '').strip()
            if not cod:
                return {'error': 'cod lipsa'}
            prod = (args.get('producator') or '').strip()
            fam = (args.get('familie') or '').strip()
            conn = get_db(); cur = conn.cursor()
            where = ' WHERE (cod = ? COLLATE NOCASE OR cod_secundar = ? COLLATE NOCASE)'
            params = [cod, cod]
            if prod:
                where += ' AND producator = ?'; params.append(prod)
            if fam:
                where += ' AND familie = ?'; params.append(fam)
            cols = ('producator, familie, cod, cod_secundar, tip, nume, '
                    'cauza, remediu, reactie, confirmare')
            cur.execute(f'SELECT {cols} FROM {safe_table("fault_codes")}{where} '
                        'ORDER BY producator, familie LIMIT 12', params)
            rows = [dict(r) for r in cur.fetchall()]
            if not rows:
                cur.execute(f'SELECT {cols} FROM fault_codes WHERE cod LIKE ? '
                            'ORDER BY producator, familie LIMIT 12', (f'%{cod}%',))
                rows = [dict(r) for r in cur.fetchall()]
            conn.close()
            return {'count': len(rows), 'coduri': rows}

        return {'error': f'Unealtă necunoscută: {name}'}
    except Exception as e:
        logger.error(f'Assistant tool {name} failed: {e}')
        return {'error': str(e)}


@app.route('/api/assistant/status', methods=['GET'])
@login_required
def assistant_status():
    return jsonify({'configured': _load_assistant_config() is not None})


@app.route('/api/assistant/chat', methods=['POST'])
@login_required
def assistant_chat():
    cfg = _load_assistant_config()
    if not cfg:
        return jsonify({'error': 'Asistentul nu e configurat. Hermes trebuie să creeze .assistant_config pe server.'}), 503

    data = request.json or {}
    history = data.get('messages', [])
    if not isinstance(history, list) or not history:
        return jsonify({'error': 'Niciun mesaj'}), 400

    # Anthropic Messages format: system is a top-level field, not a message.
    # Client history holds plain-text user/assistant turns; cap to last 20.
    convo = []
    for m in history[-20:]:
        if isinstance(m, dict) and m.get('role') in ('user', 'assistant') and m.get('content'):
            convo.append({'role': m['role'], 'content': str(m['content'])})
    if not convo:
        return jsonify({'error': 'Niciun mesaj'}), 400

    system_prompt = _build_assistant_system_prompt()
    tool_log = []
    try:
        for _ in range(8):  # max tool rounds
            resp = _minimax_call(cfg, system_prompt, convo)
            blocks = resp.get('content') or []
            text = '\n'.join(
                b.get('text', '') for b in blocks if isinstance(b, dict) and b.get('type') == 'text'
            ).strip()
            tool_uses = [b for b in blocks if isinstance(b, dict) and b.get('type') == 'tool_use']
            convo.append({'role': 'assistant', 'content': blocks})
            if not tool_uses:
                return jsonify({'reply': text or '(răspuns gol)', 'tool_log': tool_log})
            results = []
            for tu in tool_uses:
                fname = tu.get('name', '')
                fargs = tu.get('input') or {}
                result = _assistant_exec_tool(fname, fargs)
                tool_log.append({'tool': fname, 'args': fargs, 'ok': 'error' not in result})
                results.append({
                    'type': 'tool_result',
                    'tool_use_id': tu.get('id'),
                    'content': json.dumps(result, ensure_ascii=False),
                })
            convo.append({'role': 'user', 'content': results})
        return jsonify({'reply': 'Am atins limita de pași. Reformulează cererea, te rog.', 'tool_log': tool_log})
    except Exception as e:
        logger.error(f'Assistant chat failed: {e}')
        return jsonify({'error': f'Eroare asistent: {e}'}), 500


# ============ GLOBAL SEARCH (command palette) ============

def _search_snippet(text, query, width=80):
    """A short context window around the first match of `query` in `text`."""
    if not text:
        return ''
    text = str(text)
    low = text.lower()
    idx = low.find(query.lower())
    if idx < 0:
        return text[:width].strip()
    start = max(0, idx - 32)
    end = min(len(text), idx + len(query) + width)
    return ('…' if start > 0 else '') + text[start:end].strip() + ('…' if end < len(text) else '')


@app.route('/api/search', methods=['GET'])
@login_required
def global_search():
    """Unified search across everything in the app for the command palette."""
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'results': [], 'query': q})
    like = f'%{q}%'
    results = []
    conn = get_db()
    cur = conn.cursor()

    cur.execute('SELECT id, nume, client FROM proiecte '
                'WHERE nume LIKE ? OR client LIKE ? OR cod_proiect LIKE ? OR locatie LIKE ? LIMIT 8',
                (like, like, like, like))
    for r in cur.fetchall():
        results.append({'type': 'proiect', 'id': r['id'], 'title': r['nume'],
                        'subtitle': r['client'] or '', 'snippet': '', 'proiect_id': r['id']})

    cur.execute('SELECT id, nume, observatii FROM proiecte WHERE observatii LIKE ? LIMIT 8', (like,))
    for r in cur.fetchall():
        results.append({'type': 'observatie', 'id': r['id'], 'title': f"Observații — {r['nume']}",
                        'subtitle': '', 'snippet': _search_snippet(r['observatii'], q), 'proiect_id': r['id']})

    cur.execute('SELECT t.id, t.titlu, t.descriere, t.proiect_id, p.nume AS pnume FROM tasks t '
                'JOIN proiecte p ON t.proiect_id = p.id '
                'WHERE t.titlu LIKE ? OR t.descriere LIKE ? LIMIT 12', (like, like))
    for r in cur.fetchall():
        results.append({'type': 'task', 'id': r['id'], 'title': r['titlu'],
                        'subtitle': r['pnume'], 'snippet': _search_snippet(r['descriere'], q),
                        'proiect_id': r['proiect_id']})

    cur.execute('SELECT id, titlu, descriere, categorie FROM global_tasks '
                'WHERE titlu LIKE ? OR descriere LIKE ? LIMIT 10', (like, like))
    for r in cur.fetchall():
        results.append({'type': 'global_task', 'id': r['id'], 'title': r['titlu'],
                        'subtitle': r['categorie'] or 'Task zilnic', 'snippet': _search_snippet(r['descriere'], q)})

    cur.execute('SELECT c.id, c.titlu, c.proiect_id, p.nume AS pnume FROM checklist_pif c '
                'JOIN proiecte p ON c.proiect_id = p.id WHERE c.titlu LIKE ? LIMIT 10', (like,))
    for r in cur.fetchall():
        results.append({'type': 'checklist', 'id': r['id'], 'title': r['titlu'],
                        'subtitle': f"Checklist · {r['pnume']}", 'snippet': '', 'proiect_id': r['proiect_id']})

    cur.execute('SELECT j.id, j.continut, j.proiect_id, p.nume AS pnume FROM jurnal j '
                'JOIN proiecte p ON j.proiect_id = p.id WHERE j.continut LIKE ? LIMIT 10', (like,))
    for r in cur.fetchall():
        results.append({'type': 'jurnal', 'id': r['id'], 'title': _search_snippet(r['continut'], q, 50),
                        'subtitle': f"Jurnal · {r['pnume']}", 'snippet': _search_snippet(r['continut'], q),
                        'proiect_id': r['proiect_id']})

    cur.execute('SELECT e.id, e.nume, e.model, e.proiect_id, p.nume AS pnume FROM echipamente e '
                'JOIN proiecte p ON e.proiect_id = p.id '
                'WHERE e.nume LIKE ? OR e.model LIKE ? OR e.serial_number LIKE ? LIMIT 8',
                (like, like, like))
    for r in cur.fetchall():
        results.append({'type': 'echipament', 'id': r['id'], 'title': r['nume'],
                        'subtitle': f"Echipament · {r['pnume']}", 'snippet': r['model'] or '',
                        'proiect_id': r['proiect_id']})

    cur.execute('SELECT id, nume, telefon FROM clienti WHERE nume LIKE ? OR contact_principal LIKE ? LIMIT 6',
                (like, like))
    for r in cur.fetchall():
        results.append({'type': 'client', 'id': r['id'], 'title': r['nume'],
                        'subtitle': 'Client', 'snippet': r['telefon'] or ''})

    cur.execute('SELECT id, familie, parametru, descriere FROM parametri_master '
                'WHERE parametru LIKE ? OR descriere LIKE ? ORDER BY familie, parametru LIMIT 12',
                (like, like))
    for r in cur.fetchall():
        results.append({'type': 'parametru', 'id': r['id'], 'title': f"{r['parametru']} — {r['familie']}",
                        'subtitle': 'Parametru', 'snippet': _search_snippet(r['descriere'], q),
                        'familie': r['familie'], 'cod': r['parametru']})

    cur.execute('SELECT id, producator, familie, cod, tip, nume, cauza FROM fault_codes '
                'WHERE cod LIKE ? OR cod_secundar LIKE ? OR nume LIKE ? OR cauza LIKE ? '
                'ORDER BY producator, cod LIMIT 12', (like, like, like, like))
    for r in cur.fetchall():
        tip = r['tip'] or 'eroare'
        results.append({'type': 'fault_code', 'id': r['id'],
                        'title': f"{r['cod']} — {r['nume'] or ''}".strip(' —'),
                        'subtitle': f"Cod {tip} · {r['familie']}",
                        'snippet': _search_snippet(r['cauza'], q),
                        'familie': r['familie'], 'producator': r['producator'], 'cod': r['cod']})

    conn.close()

    vault = _obsidian_vault()
    if vault:
        q_low = q.lower()
        n_added = 0
        for n in _obsidian_index(vault):
            if q_low not in n['title'].lower() and q_low not in n['content'].lower():
                continue
            results.append({'type': 'obsidian', 'id': n['path'], 'title': n['title'],
                            'subtitle': n['folder'] or 'Notiță',
                            'snippet': _search_snippet(n['content'], q), 'path': n['path']})
            n_added += 1
            if n_added >= 12:
                break

    return jsonify({'results': results, 'query': q, 'count': len(results)})


# ============ INIT DEFAULT TEMPLATES ============

def init_default_templates():
    """Create default PIF Standard template if none exists"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) as count FROM project_templates')
    count = cursor.fetchone()['count']
    
    if count == 0:
        now = datetime.now().isoformat()
        template_id = generate_uuid()
        
        default_checklist = [
            'Verificare documentație',
            'Verificare conexiuni electrice',
            'Verificare parametri',
            'Test funcțional',
            'Probe de sarcină',
            'Întocmire raport PIF'
        ]
        
        default_tasks = [
            {'titlu': 'Revizuire documentație tehnică', 'prioritate': 'Normal'},
            {'titlu': 'Verificare hardware și conexiuni', 'prioritate': 'Urgent'},
            {'titlu': 'Configurare parametri sistem', 'prioritate': 'Normal'},
            {'titlu': 'Testare funcțională', 'prioritate': 'Normal'},
            {'titlu': 'Elaborare raport PIF', 'prioritate': 'Normal'}
        ]
        
        cursor.execute('''
            INSERT INTO project_templates (id, name, tip, default_checklist_json, default_tasks_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            template_id,
            'PIF Standard',
            'PIF',
            json.dumps(default_checklist),
            json.dumps(default_tasks),
            now
        ))
        
        conn.commit()
        logger.info("Created default 'PIF Standard' template")
    
    conn.close()

# ============ PWA ROUTES ============

@app.route('/m')
def mobile():
    return render_template('mobile.html')

@app.route('/service-worker.js')
def service_worker():
    return app.send_static_file('service-worker.js')

# Make service worker available at root with correct header
@app.after_request
def add_sw_header(response):
    if request.path == '/service-worker.js':
        response.headers['Service-Worker-Allowed'] = '/'
        response.headers['Content-Type'] = 'application/javascript'
        # Never HTTP-cache the worker itself — otherwise update detection is
        # delayed and erratic.
        response.headers['Cache-Control'] = 'no-cache'
    return response

# ============ DASHBOARD HOME ============

@app.route('/api/dashboard/home', methods=['GET'])
@login_required
def dashboard_home():
    conn = get_db()
    cursor = conn.cursor()
    
    # Active projects count
    cursor.execute("SELECT COUNT(*) FROM proiecte WHERE status NOT IN ('finalizat', 'anulat')")
    active_projects = cursor.fetchone()[0]

    # Total projects
    cursor.execute("SELECT COUNT(*) FROM proiecte")
    total_projects = cursor.fetchone()[0]

    # Weekly hours (last 7 days)
    cursor.execute("""
        SELECT COALESCE(SUM(durata_secunde), 0) FROM timer_sessions
        WHERE start_time >= datetime('now', '-7 days')
    """)
    weekly_seconds = cursor.fetchone()[0]
    weekly_hours = round(weekly_seconds / 3600, 1)

    # Hours previous 7-day window for delta
    cursor.execute("""
        SELECT COALESCE(SUM(durata_secunde), 0) FROM timer_sessions
        WHERE start_time >= datetime('now', '-14 days')
          AND start_time < datetime('now', '-7 days')
    """)
    prev_weekly_hours = round(cursor.fetchone()[0] / 3600, 1)
    weekly_delta = round(weekly_hours - prev_weekly_hours, 1)

    # Urgent global tasks
    cursor.execute("""
        SELECT id, titlu, prioritate, data_scadenta, categorie
        FROM global_tasks
        WHERE prioritate = 'Urgent' AND status != 'done'
        ORDER BY data_scadenta IS NULL, data_scadenta, created_at DESC LIMIT 5
    """)
    urgent_tasks = [dict(r) for r in cursor.fetchall()]
    urgent_count = len(urgent_tasks)

    # Upcoming project deadlines (next 7 days)
    cursor.execute("""
        SELECT id, nume, client, deadline
        FROM proiecte
        WHERE deadline IS NOT NULL
          AND deadline >= date('now')
          AND deadline <= date('now', '+7 days')
          AND status NOT IN ('finalizat', 'anulat')
        ORDER BY deadline LIMIT 5
    """)
    upcoming_deadlines = [dict(r) for r in cursor.fetchall()]
    deadline_count = len(upcoming_deadlines)
    
    # Active timer
    cursor.execute("""
        SELECT ts.id, ts.proiect_id as project_id, ts.start_time, p.nume as project_name
        FROM timer_sessions ts JOIN proiecte p ON ts.proiect_id = p.id
        WHERE ts.stop_time IS NULL ORDER BY ts.start_time DESC LIMIT 1
    """)
    active_timer = cursor.fetchone()
    if active_timer:
        active_timer = dict(active_timer)
    
    # Today's tasks
    cursor.execute("""
        SELECT id, titlu, status, prioritate, categorie FROM global_tasks
        WHERE date(created_at) = date('now') OR status = 'to_do'
        ORDER BY 
            CASE prioritate WHEN 'Urgent' THEN 0 WHEN 'Normal' THEN 1 ELSE 2 END,
            created_at DESC LIMIT 5
    """)
    todays_tasks = [dict(r) for r in cursor.fetchall()]
    
    # Recent journal
    cursor.execute("""
        SELECT j.id, j.proiect_id, j.data, j.continut, j.created_at, p.nume as project_name
        FROM jurnal j JOIN proiecte p ON j.proiect_id = p.id
        ORDER BY j.created_at DESC LIMIT 5
    """)
    recent_journal = [dict(r) for r in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({
        'stats': {
            'active_projects': active_projects,
            'total_projects': total_projects,
            'weekly_hours': weekly_hours,
            'weekly_delta': weekly_delta,
            'urgent_count': urgent_count,
            'deadline_count': deadline_count
        },
        'urgent_tasks': urgent_tasks,
        'upcoming_deadlines': upcoming_deadlines,
        'active_timer': active_timer,
        'todays_tasks': todays_tasks,
        'recent_journal': recent_journal
    })


# ============ AUTO-DEPLOY WEBHOOK ============

DEPLOY_SECRET_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.deploy_secret')

# Recent X-GitHub-Delivery IDs — replay guard for the deploy webhook.
_webhook_seen_deliveries = []

def get_deploy_secret():
    if os.path.exists(DEPLOY_SECRET_FILE):
        with open(DEPLOY_SECRET_FILE, 'r') as f:
            return f.read().strip()
    return None

@app.route('/webhook/deploy', methods=['POST'])
def webhook_deploy():
    secret = get_deploy_secret()
    if not secret:
        return 'Webhook not configured', 500

    signature = request.headers.get('X-Hub-Signature-256', '')
    if not signature.startswith('sha256='):
        return 'Invalid signature', 403

    expected = 'sha256=' + hmac.new(
        secret.encode(), request.data, hashlib.sha256
    ).hexdigest()

    if not hmac.compare_digest(signature, expected):
        return 'Bad signature', 403

    # Replay guard: reject a delivery already processed (a captured request
    # cannot be re-sent to force repeated redeploys/restarts).
    delivery_id = request.headers.get('X-GitHub-Delivery', '')
    if delivery_id:
        if delivery_id in _webhook_seen_deliveries:
            logger.warning(f"Webhook replay rejected: {delivery_id}")
            return 'Duplicate delivery', 409
        _webhook_seen_deliveries.append(delivery_id)
        if len(_webhook_seen_deliveries) > 200:
            del _webhook_seen_deliveries[:-200]

    payload = request.get_json(silent=True) or {}
    ref = payload.get('ref', '')
    if ref != 'refs/heads/master':
        return 'Not master branch, skipping', 200

    project_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        # `git pull` aborts when the server worktree is dirty (audit scripts
        # regenerate tracked JSONs in-place, leaving uncommitted changes), so
        # every deploy after such a run silently failed. fetch + reset --hard
        # makes the deploy idempotent: server always matches origin/master.
        # Tracked-but-regenerated files are recreated by the next audit run.
        fetch = subprocess.run(
            ['git', 'fetch', 'origin', 'master'],
            cwd=project_dir, capture_output=True, text=True, timeout=30
        )
        if fetch.returncode != 0:
            logger.error(f"Auto-deploy git fetch failed: {fetch.stderr}")
            return f'Fetch failed: {fetch.stderr}', 500
        result = subprocess.run(
            ['git', 'reset', '--hard', 'origin/master'],
            cwd=project_dir, capture_output=True, text=True, timeout=30
        )
        logger.info(f"Auto-deploy git reset: {result.stdout.strip()}")
        if result.returncode != 0:
            logger.error(f"Auto-deploy git reset failed: {result.stderr}")
            return f'Reset failed: {result.stderr}', 500
    except Exception as e:
        logger.error(f"Auto-deploy error: {e}")
        return f'Deploy error: {e}', 500

    # Install dependintele noi din requirements.txt daca exista venv
    # (previne "no module X" cand un commit aduce dependinte noi)
    venv_pip = os.path.join(project_dir, 'venv', 'bin', 'pip')
    venv_python = os.path.join(project_dir, 'venv', 'bin', 'python')
    req_file = os.path.join(project_dir, 'requirements.txt')
    if os.path.exists(req_file):
        pip_cmds = []
        if os.path.exists(venv_python):
            pip_cmds.append([venv_python, '-m', 'pip', 'install', '-r', req_file, '--quiet'])
        if os.path.exists(venv_pip):
            pip_cmds.append([venv_pip, 'install', '-r', req_file, '--quiet'])
        for pip_cmd in pip_cmds:
            try:
                pip_result = subprocess.run(
                    pip_cmd,
                    cwd=project_dir, capture_output=True, text=True, timeout=120
                )
                if pip_result.returncode != 0:
                    logger.error(f"Auto-deploy pip install failed ({pip_cmd[0]}): {pip_result.stderr}")
                    # nu blocam restart-ul — încearcă următorul fallback sau continuă
                else:
                    logger.info(f"Auto-deploy pip install OK ({pip_cmd[0]})")
                    break
            except Exception as e:
                logger.error(f"Auto-deploy pip install error ({pip_cmd[0]}): {e}")

    subprocess.Popen(
        ['sudo', 'systemctl', 'restart', 'pif-dashboard'],
        cwd=project_dir
    )

    return 'Deploy triggered', 200


# ============ PV (Proces Verbal) generation ============

def _parse_pv_request():
    """Parsare body PV: accepta JSON sau multipart cu 'payload' (JSON string)
    + files 'image_0', 'image_1', ... cu metadata-uri 'image_N_caption' / 'image_N_anchor'.
    Returneaza (data_dict, images_list) unde images_list e [{file: BytesIO, caption, anchor}].
    """
    if request.content_type and request.content_type.startswith('multipart/'):
        try:
            data = json.loads(request.form.get('payload') or '{}')
        except Exception:
            data = {}
        images = []
        i = 0
        while True:
            key = f'image_{i}'
            f = request.files.get(key)
            if f is None:
                break
            caption = request.form.get(f'{key}_caption', '')
            anchor = request.form.get(f'{key}_anchor', 'final')
            width = request.form.get(f'{key}_width', 'half')
            images.append({
                'file': BytesIO(f.read()),
                'caption': caption,
                'anchor': anchor,
                'width': width,
                'filename': f.filename,
            })
            i += 1
        return data, images
    else:
        data = request.json or {}
        return data, []




@app.route('/api/proiecte/<project_id>/pv/service/preview', methods=['GET'])
@login_required
def pv_service_preview(project_id):
    """Returneaza datele auto-populate pentru modalul PV Service: project + echipamente."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (project_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({'error': 'Project not found'}), 404
    project = row_to_dict(row)
    cursor.execute('SELECT * FROM echipamente WHERE proiect_id = ? ORDER BY created_at ASC', (project_id,))
    echipamente = [row_to_dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'proiect': project, 'echipamente': echipamente})


@app.route('/api/proiecte/<project_id>/pv/service/generate', methods=['POST'])
@login_required
def pv_service_generate(project_id):
    """Genereaza DOCX PV Service si returneaza ca download."""
    try:
        from services.pv_generator import generate_pv_service
    except Exception as e:
        logger.error(f"PV generator import failed: {e}")
        return jsonify({'error': f'Generator indisponibil: {e}'}), 500

    data, images = _parse_pv_request()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (project_id,))
    row = cursor.fetchone()
    if row is None:
        conn.close()
        return jsonify({'error': 'Project not found'}), 404
    project = row_to_dict(row)

    echipament = None
    echipament_id = data.get('echipament_id')
    if echipament_id:
        cursor.execute('SELECT * FROM echipamente WHERE id = ? AND proiect_id = ?', (echipament_id, project_id))
        e_row = cursor.fetchone()
        if e_row:
            echipament = row_to_dict(e_row)
    if echipament is None:
        cursor.execute('SELECT * FROM echipamente WHERE proiect_id = ? ORDER BY created_at ASC LIMIT 1', (project_id,))
        e_row = cursor.fetchone()
        if e_row:
            echipament = row_to_dict(e_row)
    conn.close()

    form_data = {
        'cod_proiect': data.get('cod_proiect', '').strip(),
        'data_pv': data.get('data_pv', datetime.now().strftime('%d.%m.%Y')),
        'denumire_comanda': data.get('denumire_comanda', '').strip(),
        'observatii': data.get('observatii', ''),
        'nota_facturare': data.get('nota_facturare', ''),
        'reprezentant_eg': data.get('reprezentant_eg') or 'Ion Ursu',
        'ingineri': data.get('ingineri') or [],
        'ore': data.get('ore', []),
        'images': images,
    }

    try:
        docx_bytes = generate_pv_service(project, echipament, form_data)
    except Exception as e:
        logger.exception("PV Service generate failed")
        return jsonify({'error': f'Eroare generare: {e}'}), 500

    cod = form_data['cod_proiect'] or 'PV'
    filename = f"PV_Service_{cod}_{form_data['data_pv'].replace('.', '-')}.docx"
    return send_file(
        BytesIO(docx_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/api/proiecte/<project_id>/pv/pif/preview', methods=['GET'])
@login_required
def pv_pif_preview(project_id):
    """Returneaza datele auto-populate pentru modalul PV PIF."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (project_id,))
    row = cursor.fetchone()
    conn.close()
    if row is None:
        return jsonify({'error': 'Project not found'}), 404
    return jsonify({'proiect': row_to_dict(row)})


@app.route('/api/proiecte/<project_id>/pv/pif/generate', methods=['POST'])
@login_required
def pv_pif_generate(project_id):
    """Genereaza DOCX PV PIF si returneaza ca download."""
    try:
        from services.pv_generator import generate_pv_pif
    except Exception as e:
        logger.error(f"PV PIF generator import failed: {e}")
        return jsonify({'error': f'Generator indisponibil: {e}'}), 500

    data, images = _parse_pv_request()
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM proiecte WHERE id = ?', (project_id,))
    row = cursor.fetchone()
    conn.close()
    if row is None:
        return jsonify({'error': 'Project not found'}), 404
    project = row_to_dict(row)

    form_data = {
        'cod_proiect': data.get('cod_proiect', '').strip(),
        'data_pv': data.get('data_pv', datetime.now().strftime('%d.%m.%Y')),
        'data_convocare': data.get('data_convocare', ''),
        'data_desfasurare': data.get('data_desfasurare', ''),
        'date_probe': data.get('date_probe', []),
        'nr_file': data.get('nr_file', 2),
        'nr_anexe': data.get('nr_anexe', 0),
        'constatari': data.get('constatari', ''),
        'rep_client': data.get('rep_client', ''),
        'rep_eg': data.get('rep_eg') or 'Ing. Ion Ursu',
        'images': images,
    }

    try:
        docx_bytes = generate_pv_pif(project, form_data)
    except Exception as e:
        logger.exception("PV PIF generate failed")
        return jsonify({'error': f'Eroare generare: {e}'}), 500

    cod = form_data['cod_proiect'] or 'PV'
    filename = f"PV_PIF_{cod}_{form_data['data_pv'].replace('.', '-')}.docx"
    return send_file(
        BytesIO(docx_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        as_attachment=True,
        download_name=filename,
    )


if __name__ == '__main__':
    init_db()
    init_default_templates()
    logger.info("PIF Dashboard starting...")
    app.run(host='0.0.0.0', port=5000, debug=False)
